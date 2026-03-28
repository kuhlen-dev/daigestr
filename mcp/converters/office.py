"""
Daigestr — Office Converter

Enthält Konvertierungsfunktionen für Office-Dateiformate:
- Excel (.xlsx/.xls) mit erweiterter Multi-Sheet / Chart / Formel-Unterstützung
- DOCX Extras (Kommentare, Header/Footer, Track Changes)
- Office Document Properties (Core, App, Custom)
- PPTX Hidden Slides + Embedded Objects
- MarkItDown-basierte Konvertierung mit PDF/DOCX/PPTX Erweiterungen
"""

import io
import sys
import zipfile
from pathlib import Path
from typing import Optional, Any

import structlog

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    from img2table.document import PDF as Img2TablePDF
    from img2table.ocr import TesseractOCR
    IMG2TABLE_AVAILABLE = True
except ImportError:
    IMG2TABLE_AVAILABLE = False

try:
    from icalendar import Calendar as ICalendar
    ICALENDAR_AVAILABLE = True
except ImportError:
    ICALENDAR_AVAILABLE = False

import email as _email_stdlib
import email.policy as _email_policy

from markitdown import MarkItDown
from models import ErrorCode
from utils import detect_and_fence_code_blocks
from converters.pdf import (
    extract_tables_with_pdfplumber,
    extract_tables_with_img2table,
    merge_cross_page_tables,
    tables_to_markdown,
    extract_pdf_metadata,
    prepend_pdf_toc,
    append_pdf_annotations,
    append_pdf_form_fields,
    detect_zugferd,
    parse_zugferd_xml,
    extract_xmp_metadata,
    list_embedded_files,
)

from utils import _get, _LOADED_BY_SERVER  # noqa: F401

log = structlog.get_logger()

# MarkItDown-Instanz für convert_with_markitdown
md = MarkItDown()


# =============================================================================
# FR-MKIT-007: Excel Enhanced Converter
# =============================================================================

def convert_excel_enhanced(file_path: Path, show_formulas: bool = False) -> dict[str, Any]:
    """
    Konvertiert eine Excel-Datei (.xlsx/.xls) zu Markdown mit erweiterten Features.

    Features (FR-MKIT-007):
    - AC-007-1: Jedes Worksheet als eigene Sektion (## Sheet: Name)
    - AC-007-2: Charts werden als Datentabellen extrahiert (via openpyxl)
    - AC-007-3: Zellen mit Formeln optional annotiert (z.B. 42 [=SUM(A1:A10)])
    - AC-007-4: Merged Cells werden korrekt aufgelöst

    Args:
        file_path: Pfad zur Excel-Datei (.xlsx oder .xls)
        show_formulas: Wenn True, werden Formel-Annotationen (z.B. [=SUM(A1:A10)]) hinzugefügt.

    Returns:
        Dict mit:
        - success (bool)
        - markdown (str): Konvertierter Markdown-Text
        - sheets_count (int): Anzahl der verarbeiteten Sheets
        - charts_count (int): Anzahl der gefundenen Charts
        - error_code / error: Nur bei Fehler
    """
    if not _get("OPENPYXL_AVAILABLE", OPENPYXL_AVAILABLE):
        log.warning("openpyxl_not_available", file=str(file_path))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": "openpyxl ist nicht installiert (pip install openpyxl)",
        }

    log.info("excel_enhanced_convert_start", file=str(file_path), show_formulas=show_formulas)

    # Use sys.modules lookup so patch("openpyxl.load_workbook") in tests takes effect
    _openpyxl = sys.modules.get("openpyxl", openpyxl)
    try:
        # data_only=False um Formeln zu lesen; ein zweites Mal mit data_only=True für Werte
        wb_formulas = _openpyxl.load_workbook(str(file_path), data_only=False)
        wb_values = _openpyxl.load_workbook(str(file_path), data_only=True)
    except Exception as exc:
        log.error("excel_open_error", file=str(file_path), error=str(exc))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": f"Excel-Datei konnte nicht geöffnet werden: {str(exc)}",
        }

    markdown_parts: list[str] = []
    total_charts = 0
    hidden_sheet_names: list[str] = []

    # T-MKIT-026: Separate visible from hidden/veryHidden sheets
    # Process visible sheets first, then hidden sheets as a secondary section
    visible_sheet_names: list[str] = []
    hidden_state_map: dict[str, str] = {}  # sheet_name → "hidden" | "veryHidden"

    for sheet_name in wb_values.sheetnames:
        ws = wb_values[sheet_name]
        state = getattr(ws, "sheet_state", "visible")
        if state == "visible":
            visible_sheet_names.append(sheet_name)
        else:
            hidden_state_map[sheet_name] = state
            hidden_sheet_names.append(sheet_name)

    # Process visible first, then hidden — preserving original order within each group
    all_ordered_names = visible_sheet_names + hidden_sheet_names

    for sheet_name in all_ordered_names:
        ws_values = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name]

        # T-MKIT-026: Annotate hidden/veryHidden sheets
        state = hidden_state_map.get(sheet_name, "visible")
        if state == "veryHidden":
            sheet_heading = f"## Sheet: {sheet_name} [VERY HIDDEN]"
        elif state == "hidden":
            sheet_heading = f"## Sheet: {sheet_name} [HIDDEN]"
        else:
            sheet_heading = f"## Sheet: {sheet_name}"

        sheet_parts: list[str] = [sheet_heading]

        # AC-007-4: Merged Cells auflösen
        # Baue ein Dict: (row, col) → Wert der Hauptzelle des Merge-Bereichs
        merged_cell_values: dict[tuple[int, int], Any] = {}
        for merge_range in ws_values.merged_cells.ranges:
            # Wert der linken oberen Zelle des Bereichs
            top_left = ws_values.cell(merge_range.min_row, merge_range.min_col)
            top_left_formula = ws_formulas.cell(merge_range.min_row, merge_range.min_col)
            for row_idx in range(merge_range.min_row, merge_range.max_row + 1):
                for col_idx in range(merge_range.min_col, merge_range.max_col + 1):
                    merged_cell_values[(row_idx, col_idx)] = (
                        top_left.value,
                        top_left_formula.value,
                    )

        # Zeilen und Spalten ermitteln
        rows = list(ws_values.iter_rows())
        formula_rows = list(ws_formulas.iter_rows())

        # Leeres Sheet graceful behandeln (AC: test_empty_sheet_handled)
        if not rows:
            sheet_parts.append("*Kein Inhalt*")
            markdown_parts.append("\n".join(sheet_parts))
            log.debug("excel_empty_sheet", sheet=sheet_name)
            continue

        # Maximale Spaltenanzahl bestimmen (über alle Zeilen)
        max_cols = max((len(row) for row in rows), default=0)

        if max_cols == 0:
            sheet_parts.append("*Kein Inhalt*")
            markdown_parts.append("\n".join(sheet_parts))
            continue

        # Tabellen-Daten aufbauen
        table_rows: list[list[str]] = []

        for row_idx, (row_cells, formula_cells) in enumerate(
            zip(rows, formula_rows), start=1
        ):
            row_data: list[str] = []
            for col_idx, (cell, formula_cell) in enumerate(
                zip(row_cells, formula_cells), start=1
            ):
                # AC-007-4: Merged Cell Wert holen
                if (row_idx, col_idx) in merged_cell_values:
                    raw_value, raw_formula = merged_cell_values[(row_idx, col_idx)]
                else:
                    raw_value = cell.value
                    raw_formula = formula_cell.value

                # Zellwert als String
                cell_str = "" if raw_value is None else str(raw_value)

                # AC-007-3: Formel-Annotation
                if (
                    show_formulas
                    and raw_formula is not None
                    and isinstance(raw_formula, str)
                    and raw_formula.startswith("=")
                ):
                    cell_str = f"{cell_str} [{raw_formula}]"

                # Pipe-Zeichen im Zellinhalt escapen (würde Tabelle brechen)
                cell_str = cell_str.replace("|", "\\|").replace("\n", " ")
                row_data.append(cell_str)

            # Fehlende Spalten auffüllen
            while len(row_data) < max_cols:
                row_data.append("")

            table_rows.append(row_data)

        # Prüfen ob Sheet komplett leer ist (alle Zellen None)
        all_empty = all(
            cell_str == ""
            for row in table_rows
            for cell_str in row
        )
        if all_empty:
            sheet_parts.append("*Kein Inhalt*")
            markdown_parts.append("\n".join(sheet_parts))
            continue

        # Markdown-Tabelle aufbauen
        if table_rows:
            header = table_rows[0]
            # Header mit generischen Spaltennamen falls leer
            display_header = [
                h if h else get_column_letter(i + 1)
                for i, h in enumerate(header)
            ]
            table_lines: list[str] = []
            table_lines.append("| " + " | ".join(display_header) + " |")
            table_lines.append("| " + " | ".join(["---"] * max_cols) + " |")
            for data_row in table_rows[1:]:
                padded = data_row + [""] * (max_cols - len(data_row))
                table_lines.append("| " + " | ".join(padded[:max_cols]) + " |")
            sheet_parts.append("\n".join(table_lines))

        # AC-007-2: Charts extrahieren
        sheet_charts = getattr(ws_values, "_charts", [])
        if sheet_charts:
            for chart_idx, chart in enumerate(sheet_charts, start=1):
                total_charts += 1
                chart_title = ""
                try:
                    if hasattr(chart, "title") and chart.title is not None:
                        title_obj = chart.title
                        # openpyxl chart.title kann str, Title-Objekt oder None sein
                        if isinstance(title_obj, str):
                            chart_title = title_obj
                        elif hasattr(title_obj, "tx") and title_obj.tx is not None:
                            # Title-Objekt mit tx.rich.p[].r[].t Struktur
                            try:
                                texts = []
                                for para in title_obj.tx.rich.p:
                                    for run in para.r:
                                        if run.t:
                                            texts.append(run.t)
                                chart_title = " ".join(texts)
                            except Exception:
                                chart_title = f"Chart {chart_idx}"
                        else:
                            chart_title = f"Chart {chart_idx}"
                    else:
                        chart_title = f"Chart {chart_idx}"
                except Exception:
                    chart_title = f"Chart {chart_idx}"

                chart_parts: list[str] = [f"### Chart: {chart_title}"]

                # Datenserien extrahieren
                try:
                    series_list = []
                    if hasattr(chart, "series"):
                        for serie in chart.series:
                            serie_title = ""
                            try:
                                if hasattr(serie, "title") and serie.title is not None:
                                    st = serie.title
                                    if hasattr(st, "v") and st.v is not None:
                                        serie_title = str(st.v)
                                    elif hasattr(st, "strRef") and st.strRef is not None:
                                        cache = getattr(st.strRef, "strCache", None)
                                        if cache and hasattr(cache, "pt") and cache.pt:
                                            serie_title = str(cache.pt[0].v)
                            except Exception:
                                pass

                            # Werte aus dem Cache holen
                            values: list[str] = []
                            try:
                                val_ref = None
                                if hasattr(serie, "val"):
                                    val_ref = serie.val
                                elif hasattr(serie, "yVal"):
                                    val_ref = serie.yVal

                                if val_ref is not None:
                                    num_cache = getattr(val_ref, "numRef", None)
                                    if num_cache:
                                        num_data = getattr(num_cache, "numCache", None)
                                        if num_data and hasattr(num_data, "pt"):
                                            values = [str(pt.v) for pt in num_data.pt]
                            except Exception:
                                pass

                            series_list.append({
                                "title": serie_title or f"Serie {len(series_list) + 1}",
                                "values": values,
                            })

                    if series_list:
                        # Tabelle mit Serien als Spalten
                        headers = [s["title"] for s in series_list]
                        chart_parts.append("| " + " | ".join(headers) + " |")
                        chart_parts.append("| " + " | ".join(["---"] * len(headers)) + " |")
                        max_vals = max((len(s["values"]) for s in series_list), default=0)
                        for vi in range(max_vals):
                            row_vals = []
                            for s in series_list:
                                row_vals.append(s["values"][vi] if vi < len(s["values"]) else "")
                            chart_parts.append("| " + " | ".join(row_vals) + " |")
                    else:
                        chart_parts.append("*Keine Datenserien gefunden*")

                except Exception as chart_exc:
                    log.warning(
                        "excel_chart_extraction_error",
                        sheet=sheet_name,
                        chart=chart_idx,
                        error=str(chart_exc),
                    )
                    chart_parts.append("*Chart-Daten konnten nicht extrahiert werden*")

                sheet_parts.append("\n".join(chart_parts))

        markdown_parts.append("\n\n".join(sheet_parts))

    wb_values.close()
    wb_formulas.close()

    combined_markdown = "\n\n".join(markdown_parts)
    sheets_count = len(wb_values.sheetnames)

    log.info(
        "excel_enhanced_convert_done",
        file=str(file_path),
        sheets=sheets_count,
        charts=total_charts,
        chars=len(combined_markdown),
        hidden_sheets=len(hidden_sheet_names),
    )

    result: dict[str, Any] = {
        "success": True,
        "markdown": combined_markdown,
        "sheets_count": sheets_count,
        "charts_count": total_charts,
    }
    # T-MKIT-026: Include hidden_sheets only when there are hidden sheets
    if hidden_sheet_names:
        result["hidden_sheets"] = hidden_sheet_names
    return result


# =============================================================================
# DOCX Extras: Kommentare, Header/Footer, Track Changes (FR-MKIT-008)
# =============================================================================

def extract_docx_extras(file_path: Path) -> dict:
    """
    Extrahiert erweiterte Metadaten aus einer DOCX-Datei (FR-MKIT-008).

    Verarbeitet:
    - Kommentare aus word/comments.xml (Author, Date, Text)
    - Header und Footer aus allen Dokumentsektionen via python-docx
    - Track Changes (Einfügungen/Löschungen) aus dem DOCX-XML (w:ins, w:del)

    Args:
        file_path: Pfad zur DOCX-Datei.

    Returns:
        Dict mit:
        - comments: Liste von Dicts mit 'author', 'date', 'text'
        - headers: Liste von Header-Texten (nicht leer)
        - footers: Liste von Footer-Texten (nicht leer)
        - track_changes: Liste von Dicts mit 'type' ('insertion'/'deletion'), 'author', 'date', 'text'
    """
    import xml.etree.ElementTree as ET

    result: dict = {
        "comments": [],
        "headers": [],
        "footers": [],
        "track_changes": [],
    }

    W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"

    # --- Kommentare aus word/comments.xml ---
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            if "word/comments.xml" in zf.namelist():
                xml_data = zf.read("word/comments.xml")
                root = ET.fromstring(xml_data)
                for comment in root.findall(f"{{{W_NS}}}comment"):
                    author = comment.get(f"{{{W_NS}}}author", "")
                    date = comment.get(f"{{{W_NS}}}date", "")
                    # Text aus allen w:t Elementen zusammenführen
                    text_parts = [t.text or "" for t in comment.findall(f".//{{{W_NS}}}t")]
                    text = " ".join(text_parts).strip()
                    if text:
                        result["comments"].append({
                            "author": author,
                            "date": date,
                            "text": text,
                        })
                log.info(
                    "docx_comments_extracted",
                    file=str(file_path),
                    count=len(result["comments"]),
                )
    except Exception as e:
        log.warning("docx_comments_error", file=str(file_path), error=str(e))

    # --- Header/Footer via python-docx ---
    try:
        import docx as _docx  # python-docx
        doc = _docx.Document(str(file_path))
        for section in doc.sections:
            for hf_obj, target_list in [
                (section.header, result["headers"]),
                (section.footer, result["footers"]),
            ]:
                try:
                    text = hf_obj.text.strip() if hf_obj and hasattr(hf_obj, "text") else ""
                    if not text:
                        # Manuell aus Paragraphen extrahieren
                        if hf_obj and hasattr(hf_obj, "paragraphs"):
                            parts = [p.text.strip() for p in hf_obj.paragraphs if p.text.strip()]
                            text = " | ".join(parts)
                    if text and text not in target_list:
                        target_list.append(text)
                except Exception as inner_e:
                    log.debug("docx_hf_paragraph_error", error=str(inner_e))
        log.info(
            "docx_headers_footers_extracted",
            file=str(file_path),
            headers=len(result["headers"]),
            footers=len(result["footers"]),
        )
    except Exception as e:
        log.warning("docx_headers_footers_error", file=str(file_path), error=str(e))

    # --- Track Changes aus word/document.xml (w:ins, w:del) ---
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            if "word/document.xml" in zf.namelist():
                xml_data = zf.read("word/document.xml")
                root = ET.fromstring(xml_data)

                for tag, change_type in [("ins", "insertion"), ("del", "deletion")]:
                    for elem in root.findall(f".//{{{W_NS}}}{tag}"):
                        author = elem.get(f"{{{W_NS}}}author", "")
                        date = elem.get(f"{{{W_NS}}}date", "")
                        # w:delText für Löschungen, w:t für Einfügungen
                        if change_type == "deletion":
                            text_parts = [
                                t.text or ""
                                for t in elem.findall(f".//{{{W_NS}}}delText")
                            ]
                        else:
                            text_parts = [
                                t.text or ""
                                for t in elem.findall(f".//{{{W_NS}}}t")
                            ]
                        text = "".join(text_parts)
                        if text:
                            result["track_changes"].append({
                                "type": change_type,
                                "author": author,
                                "date": date,
                                "text": text,
                            })

                log.info(
                    "docx_track_changes_extracted",
                    file=str(file_path),
                    count=len(result["track_changes"]),
                )
    except Exception as e:
        log.warning("docx_track_changes_error", file=str(file_path), error=str(e))

    return result


def append_docx_extras_to_markdown(markdown: str, extras: dict) -> str:
    """
    Fügt DOCX-Extras (Kommentare, Header/Footer, Track Changes) als Markdown-Sektionen an.

    Sektionen werden nur angefügt, wenn die jeweiligen Extras nicht leer sind:
    - ## Kommentare → Blockquotes mit Author und Date
    - ## Header und Footer → Inhalt als Liste
    - ## Änderungsverfolgung → Einfügungen und Löschungen als Diff-Notation

    Args:
        markdown: Bereits konvertierter Markdown-Text.
        extras: Rückgabe-Dict von extract_docx_extras().

    Returns:
        Erweiterter Markdown-String.
    """
    sections: list[str] = []

    # --- Kommentare als Blockquotes ---
    comments = extras.get("comments", [])
    if comments:
        lines = ["## Kommentare", ""]
        for c in comments:
            author = c.get("author", "Unbekannt")
            date = c.get("date", "")
            text = c.get("text", "")
            date_str = f" ({date})" if date else ""
            lines.append(f"> **{author}**{date_str}: {text}")
            lines.append("")
        sections.append("\n".join(lines).rstrip())

    # --- Header und Footer ---
    headers = extras.get("headers", [])
    footers = extras.get("footers", [])
    if headers or footers:
        lines = ["## Header und Footer", ""]
        if headers:
            lines.append("**Header:**")
            for h in headers:
                lines.append(f"- {h}")
            lines.append("")
        if footers:
            lines.append("**Footer:**")
            for f in footers:
                lines.append(f"- {f}")
            lines.append("")
        sections.append("\n".join(lines).rstrip())

    # --- Track Changes als Diff-Notation ---
    track_changes = extras.get("track_changes", [])
    if track_changes:
        lines = ["## Änderungsverfolgung", ""]
        lines.append("```diff")
        for tc in track_changes:
            change_type = tc.get("type", "")
            author = tc.get("author", "")
            date = tc.get("date", "")
            text = tc.get("text", "")
            meta = f"  # {author}" if author else ""
            if date:
                meta += f" ({date})" if author else f"  # ({date})"
            if change_type == "insertion":
                lines.append(f"+ {text}{meta}")
            elif change_type == "deletion":
                lines.append(f"- {text}{meta}")
        lines.append("```")
        sections.append("\n".join(lines))

    if not sections:
        return markdown

    return markdown.rstrip() + "\n\n" + "\n\n".join(sections)


# =============================================================================
# T-MKIT-028: Office Document Properties (DOCX, XLSX, PPTX)
# =============================================================================

def extract_document_properties(file_data: bytes, filename: str) -> Optional[dict[str, Any]]:
    """
    Extrahiert Core-, App- und Custom-Properties aus Office-Dokumenten (DOCX, XLSX, PPTX).

    Office-Dokumente sind ZIP-Archive mit docProps/*.xml Dateien. Diese Funktion
    parst die drei relevanten XML-Dateien und gibt ein strukturiertes Dict zurück.

    Args:
        file_data: Rohe Datei-Bytes der Office-Datei
        filename: Dateiname (wird für Extension-Erkennung genutzt)

    Returns:
        Dict mit Schlüsseln 'core', 'app', 'custom' oder None wenn kein Office-Format.
        Fehlende XML-Dateien → leere Sections (kein Crash).
    """
    import xml.etree.ElementTree as ET

    suffix = Path(filename).suffix.lower()
    if suffix not in {".docx", ".xlsx", ".pptx", ".odt", ".ods", ".odp"}:
        return None

    core_ns = {
        "dc": "http://purl.org/dc/elements/1.1/",
        "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
        "dcterms": "http://purl.org/dc/terms/",
    }
    app_ns = {
        "ap": "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties",
    }
    custom_ns = {
        "vt": "http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes",
    }

    core: dict[str, str] = {}
    app: dict[str, str] = {}
    custom: dict[str, str] = {}

    try:
        with zipfile.ZipFile(io.BytesIO(file_data), "r") as zf:
            zip_names = set(zf.namelist())

            # --- docProps/core.xml ---
            if "docProps/core.xml" in zip_names:
                try:
                    core_xml = zf.read("docProps/core.xml")
                    root = ET.fromstring(core_xml)

                    def _core(tag_ns: str, tag_name: str) -> Optional[str]:
                        el = root.find(f"{{{core_ns[tag_ns]}}}{tag_name}")
                        return el.text.strip() if el is not None and el.text else None

                    for key, ns, tag in [
                        ("title", "dc", "title"),
                        ("subject", "dc", "subject"),
                        ("description", "dc", "description"),
                        ("author", "dc", "creator"),
                        ("last_modified_by", "cp", "lastModifiedBy"),
                        ("created", "dcterms", "created"),
                        ("modified", "dcterms", "modified"),
                        ("revision", "cp", "revision"),
                    ]:
                        val = _core(ns, tag)
                        if val is not None:
                            core[key] = val
                except Exception as e:
                    log.debug("doc_properties_core_parse_failed", filename=filename, error=str(e))

            # --- docProps/app.xml ---
            if "docProps/app.xml" in zip_names:
                try:
                    app_xml = zf.read("docProps/app.xml")
                    root = ET.fromstring(app_xml)
                    ap_ns = app_ns["ap"]

                    def _app(tag_name: str) -> Optional[str]:
                        el = root.find(f"{{{ap_ns}}}{tag_name}")
                        return el.text.strip() if el is not None and el.text else None

                    for key, tag in [
                        ("company", "Company"),
                        ("manager", "Manager"),
                        ("application", "Application"),
                        ("app_version", "AppVersion"),
                        ("pages", "Pages"),
                        ("words", "Words"),
                        ("characters", "Characters"),
                        ("total_time_minutes", "TotalTime"),
                        ("template", "Template"),
                    ]:
                        val = _app(tag)
                        if val is not None:
                            app[key] = val
                except Exception as e:
                    log.debug("doc_properties_app_parse_failed", filename=filename, error=str(e))

            # --- docProps/custom.xml ---
            if "docProps/custom.xml" in zip_names:
                try:
                    custom_xml = zf.read("docProps/custom.xml")
                    root = ET.fromstring(custom_xml)
                    vt_ns = custom_ns["vt"]

                    # Iterate over <property> elements
                    for prop_el in root:
                        name = prop_el.get("name")
                        if not name:
                            continue
                        # Try known value types
                        value: Optional[str] = None
                        for vtype in ("lpwstr", "i4", "bool", "filetime", "r8", "decimal"):
                            val_el = prop_el.find(f"{{{vt_ns}}}{vtype}")
                            if val_el is not None and val_el.text:
                                value = val_el.text.strip()
                                break
                        if value is not None:
                            custom[name] = value
                except Exception as e:
                    log.debug("doc_properties_custom_parse_failed", filename=filename, error=str(e))

    except Exception as e:
        log.debug("doc_properties_zip_open_failed", filename=filename, error=str(e))
        return None

    return {
        "core": core,
        "app": app,
        "custom": custom,
    }


# =============================================================================
# T-MKIT-030: PPTX Hidden Slides + Embedded Excel aus Charts
# =============================================================================

def extract_pptx_hidden_info(file_data: bytes) -> Optional[dict[str, Any]]:
    """
    Analysiert eine PPTX-Datei auf versteckte Slides und eingebettete Excel-Objekte.

    PPTX ist intern ein ZIP-Archiv. Versteckte Slides haben das Attribut show="0"
    im p:sld-Element. Eingebettete Excel-Dateien liegen in ppt/embeddings/.

    Args:
        file_data: Rohe PPTX-Bytes

    Returns:
        Dict mit:
            - hidden_slide_count (int): Anzahl versteckter Slides
            - hidden_slide_numbers (list[int]): Slide-Nummern (1-basiert)
            - embedded_objects (list[str]): Namen eingebetteter Dateien
        Oder None bei Fehler.
    """
    try:
        hidden_slide_numbers: list[int] = []
        embedded_objects: list[str] = []

        with zipfile.ZipFile(io.BytesIO(file_data), "r") as zf:
            all_names = zf.namelist()

            # Slide XMLs parsen: ppt/slides/slide1.xml, slide2.xml, ...
            import re as _re
            slide_files = [
                n for n in all_names
                if _re.match(r"ppt/slides/slide\d+\.xml$", n)
            ]

            for slide_name in sorted(slide_files):
                # Slide-Nummer aus Dateinamen extrahieren
                m = _re.search(r"slide(\d+)\.xml$", slide_name)
                slide_num = int(m.group(1)) if m else 0

                try:
                    xml_bytes = zf.read(slide_name)
                    xml_text = xml_bytes.decode("utf-8", errors="replace")
                    # Prüfe auf show="0" im p:sld-Element
                    # Das Attribut kann mit oder ohne Namespace-Präfix erscheinen
                    if 'show="0"' in xml_text:
                        # Sicherstellen dass es am p:sld Element hängt
                        # (nicht in einem anderen Element)
                        import xml.etree.ElementTree as _ET
                        try:
                            root = _ET.fromstring(xml_bytes)
                            # p:sld Namespace
                            pml_ns = "http://schemas.openxmlformats.org/presentationml/2006/main"
                            tag_local = root.tag.split("}")[-1] if "}" in root.tag else root.tag
                            show_val = root.get(f"{{{pml_ns}}}show") or root.get("show")
                            if tag_local == "sld" and show_val == "0":
                                hidden_slide_numbers.append(slide_num)
                        except Exception:
                            # Fallback: einfache String-Suche
                            if 'show="0"' in xml_text:
                                hidden_slide_numbers.append(slide_num)
                except Exception:
                    pass

            # Eingebettete Objekte in ppt/embeddings/
            embedding_files = [
                n for n in all_names
                if n.startswith("ppt/embeddings/") and not n.endswith("/")
            ]
            for emb_path in sorted(embedding_files):
                emb_name = emb_path.split("/")[-1]
                # Nur Excel-Objekte: *.xlsx oder oleObject*.bin
                if emb_name.endswith(".xlsx") or (
                    emb_name.startswith("oleObject") and emb_name.endswith(".bin")
                ):
                    embedded_objects.append(emb_name)

        return {
            "hidden_slide_count": len(hidden_slide_numbers),
            "hidden_slide_numbers": sorted(hidden_slide_numbers),
            "embedded_objects": embedded_objects,
        }
    except Exception as e:
        log.warning("pptx_hidden_info_failed", error=str(e))
        return None


# =============================================================================
# T-DAI-012: OpenDocument Extras (ODT, ODP, ODS)
# =============================================================================

def extract_odt_extras(file_path: Path) -> dict:
    """
    Extrahiert erweiterte Metadaten aus einer ODT-Datei (T-DAI-012).

    ODT-Dokumente sind ZIP-Archive mit content.xml (ODF XML).
    Verarbeitet:
    - Kommentare (office:annotation Elemente)
    - Track Changes (text:tracked-changes Elemente)
    - Header und Footer (style:header, style:footer in styles.xml)

    Args:
        file_path: Pfad zur ODT-Datei.

    Returns:
        Dict mit:
        - comments: Liste von Dicts mit 'author', 'date', 'text'
        - track_changes: Liste von Dicts mit 'type', 'author', 'date', 'text'
        - headers: Liste von Header-Texten (nicht leer)
        - footers: Liste von Footer-Texten (nicht leer)
    """
    import xml.etree.ElementTree as ET

    result: dict = {
        "comments": [],
        "track_changes": [],
        "headers": [],
        "footers": [],
    }

    # ODF XML-Namespaces
    OFFICE_NS = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"
    TEXT_NS = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"
    DC_NS = "http://purl.org/dc/elements/1.1/"
    META_NS = "urn:oasis:names:tc:opendocument:xmlns:meta:1.0"
    STYLE_NS = "urn:oasis:names:tc:opendocument:xmlns:style:1.0"
    FO_NS = "urn:oasis:names:tc:opendocument:xmlns:xsl-fo-compatible:1.0"

    def _get_text_content(element) -> str:
        """Extrahiert alle Texte rekursiv aus einem XML-Element."""
        parts = []
        if element.text:
            parts.append(element.text.strip())
        for child in element:
            child_text = _get_text_content(child)
            if child_text:
                parts.append(child_text)
            if child.tail:
                parts.append(child.tail.strip())
        return " ".join(p for p in parts if p)

    # --- content.xml: Kommentare und Track Changes ---
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            if "content.xml" in zf.namelist():
                xml_data = zf.read("content.xml")
                root = ET.fromstring(xml_data)

                # Kommentare: office:annotation Elemente
                for annotation in root.iter(f"{{{OFFICE_NS}}}annotation"):
                    # Author aus dc:creator oder office:author
                    author = ""
                    author_el = annotation.find(f"{{{DC_NS}}}creator")
                    if author_el is None:
                        author_el = annotation.find(f"{{{META_NS}}}creator")
                    if author_el is not None and author_el.text:
                        author = author_el.text.strip()

                    # Datum aus dc:date oder office:date
                    date = ""
                    date_el = annotation.find(f"{{{DC_NS}}}date")
                    if date_el is None:
                        date_el = annotation.find(f"{{{META_NS}}}date")
                    if date_el is not None and date_el.text:
                        date = date_el.text.strip()

                    # Text aus text:p Elementen
                    text_parts = []
                    for p in annotation.findall(f"{{{TEXT_NS}}}p"):
                        p_text = _get_text_content(p)
                        if p_text:
                            text_parts.append(p_text)
                    text = " ".join(text_parts).strip()

                    if text or author:
                        result["comments"].append({
                            "author": author,
                            "date": date,
                            "text": text,
                        })

                log.info(
                    "odt_comments_extracted",
                    file=str(file_path),
                    count=len(result["comments"]),
                )

                # Track Changes: text:tracked-changes und text:insertion/text:deletion
                tracked_changes_el = root.find(f".//{{{TEXT_NS}}}tracked-changes")
                if tracked_changes_el is not None:
                    for change_el in tracked_changes_el:
                        tag_local = change_el.tag.split("}")[-1] if "}" in change_el.tag else change_el.tag
                        if tag_local in ("insertion", "deletion"):
                            change_type = "insertion" if tag_local == "insertion" else "deletion"
                            # Autor aus office:change-info
                            author = change_el.get(f"{{{OFFICE_NS}}}chng-author", "")
                            date = change_el.get(f"{{{OFFICE_NS}}}chng-date", "")
                            # Text aus Kind-Elementen
                            text = _get_text_content(change_el).strip()
                            if text or author:
                                result["track_changes"].append({
                                    "type": change_type,
                                    "author": author,
                                    "date": date,
                                    "text": text,
                                })

                log.info(
                    "odt_track_changes_extracted",
                    file=str(file_path),
                    count=len(result["track_changes"]),
                )

    except Exception as e:
        log.warning("odt_content_xml_error", file=str(file_path), error=str(e))

    # --- styles.xml: Header und Footer ---
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            if "styles.xml" in zf.namelist():
                xml_data = zf.read("styles.xml")
                root = ET.fromstring(xml_data)

                for hf_tag, target_list in [
                    (f"{{{STYLE_NS}}}header", result["headers"]),
                    (f"{{{STYLE_NS}}}footer", result["footers"]),
                ]:
                    for hf_el in root.iter(hf_tag):
                        text = _get_text_content(hf_el).strip()
                        if text and text not in target_list:
                            target_list.append(text)

                log.info(
                    "odt_headers_footers_extracted",
                    file=str(file_path),
                    headers=len(result["headers"]),
                    footers=len(result["footers"]),
                )
    except Exception as e:
        log.warning("odt_styles_xml_error", file=str(file_path), error=str(e))

    return result


def extract_odp_hidden_slides(file_data: bytes) -> Optional[dict]:
    """
    Analysiert eine ODP-Datei auf versteckte Slides (T-DAI-012).

    ODP ist ein ZIP-Archiv mit content.xml. Slides sind draw:page Elemente.
    Ein versteckter Slide hat ein zugehöriges Style mit
    presentation:visibility="hidden" oder draw:show="false".

    Args:
        file_data: Rohe ODP-Bytes.

    Returns:
        Dict mit:
        - hidden_slide_count (int): Anzahl versteckter Slides
        - hidden_slides (list[int]): Slide-Nummern (1-basiert)
        Oder None bei Fehler.
    """
    import xml.etree.ElementTree as ET

    DRAW_NS = "urn:oasis:names:tc:opendocument:xmlns:drawing:1.0"
    PRESENTATION_NS = "urn:oasis:names:tc:opendocument:xmlns:presentation:1.0"
    STYLE_NS = "urn:oasis:names:tc:opendocument:xmlns:style:1.0"
    OFFICE_NS = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"

    try:
        with zipfile.ZipFile(io.BytesIO(file_data), "r") as zf:
            if "content.xml" not in zf.namelist():
                return {"hidden_slide_count": 0, "hidden_slides": []}

            xml_data = zf.read("content.xml")
            root = ET.fromstring(xml_data)

            # Alle draw:page Stile aus automatic-styles und styles sammeln
            # Diese können presentation:visibility="hidden" enthalten
            hidden_style_names: set = set()

            for styles_container in root.iter(f"{{{OFFICE_NS}}}automatic-styles"):
                for style_el in styles_container:
                    style_name = style_el.get(f"{{{STYLE_NS}}}name", "")
                    # Prüfe presentation:show attribute am Style selbst
                    visibility = style_el.get(f"{{{PRESENTATION_NS}}}visibility", "")
                    if visibility == "hidden":
                        hidden_style_names.add(style_name)
                    # Prüfe draw:drawing-page-properties Kind-Element
                    for prop_el in style_el:
                        prop_tag_local = prop_el.tag.split("}")[-1] if "}" in prop_el.tag else prop_el.tag
                        if prop_tag_local == "drawing-page-properties":
                            vis = prop_el.get(f"{{{PRESENTATION_NS}}}visibility", "")
                            if vis == "hidden":
                                hidden_style_names.add(style_name)

            hidden_slide_numbers: list[int] = []
            slide_num = 0

            # Finde alle draw:page Elemente
            for page_el in root.iter(f"{{{DRAW_NS}}}page"):
                slide_num += 1
                page_style = page_el.get(f"{{{DRAW_NS}}}style-name", "")
                # Direkte Sichtbarkeit am draw:page Element
                show_val = page_el.get(f"{{{PRESENTATION_NS}}}show", "")
                if show_val == "false" or page_style in hidden_style_names:
                    hidden_slide_numbers.append(slide_num)

        log.info(
            "odp_hidden_slides_extracted",
            hidden_count=len(hidden_slide_numbers),
            total_slides=slide_num,
        )

        return {
            "hidden_slide_count": len(hidden_slide_numbers),
            "hidden_slides": sorted(hidden_slide_numbers),
        }
    except Exception as e:
        log.warning("odp_hidden_slides_failed", error=str(e))
        return None


def convert_ods_enhanced(file_path: Path, show_formulas: bool = False) -> dict:
    """
    Konvertiert eine ODS-Datei zu Markdown mit erweiterten Features (T-DAI-012).

    ODS ist ein ZIP-Archiv mit content.xml (ODF XML). Nutzt direkte XML-Verarbeitung
    (kein openpyxl — unterstützt kein ODS).

    Features:
    - Multi-Sheet: Jedes table:table als eigene ## Sheet: Name Sektion
    - Hidden Sheets: Sheets mit table:display="false" als [HIDDEN] markiert
    - Formeln: table:formula Attribute optional annotiert

    Args:
        file_path: Pfad zur ODS-Datei.
        show_formulas: Wenn True, werden Formel-Annotationen hinzugefügt.

    Returns:
        Dict mit:
        - success (bool)
        - markdown (str): Konvertierter Markdown-Text
        - sheets_count (int): Gesamtanzahl der Sheets
        - hidden_sheets (list[str]): Namen der versteckten Sheets (nur wenn vorhanden)
        - error_code / error: Nur bei Fehler
    """
    import xml.etree.ElementTree as ET

    TABLE_NS = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
    TEXT_NS = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"
    OFFICE_NS = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"
    STYLE_NS = "urn:oasis:names:tc:opendocument:xmlns:style:1.0"

    log.info("ods_enhanced_convert_start", file=str(file_path), show_formulas=show_formulas)

    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            if "content.xml" not in zf.namelist():
                return {
                    "success": False,
                    "error_code": ErrorCode.CONVERSION_FAILED,
                    "error": "Keine content.xml in der ODS-Datei gefunden",
                }
            xml_data = zf.read("content.xml")
    except Exception as exc:
        log.error("ods_open_error", file=str(file_path), error=str(exc))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": f"ODS-Datei konnte nicht geöffnet werden: {str(exc)}",
        }

    try:
        root = ET.fromstring(xml_data)
    except Exception as exc:
        log.error("ods_xml_parse_error", file=str(file_path), error=str(exc))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": f"ODS content.xml konnte nicht geparst werden: {str(exc)}",
        }

    # Sammle Table-Style-Eigenschaften für Hidden-Sheet-Erkennung
    # table:table-properties mit table:display="false"
    hidden_style_names: set = set()
    for auto_styles in root.iter(f"{{{OFFICE_NS}}}automatic-styles"):
        for style_el in auto_styles:
            style_name = style_el.get(f"{{{STYLE_NS}}}name", "")
            for prop_el in style_el:
                prop_tag_local = prop_el.tag.split("}")[-1] if "}" in prop_el.tag else prop_el.tag
                if prop_tag_local == "table-properties":
                    display = prop_el.get(f"{{{TABLE_NS}}}display", "true")
                    if display == "false":
                        hidden_style_names.add(style_name)

    markdown_parts: list[str] = []
    hidden_sheet_names: list[str] = []

    # Alle table:table Elemente finden
    for table_el in root.iter(f"{{{TABLE_NS}}}table"):
        sheet_name = table_el.get(f"{{{TABLE_NS}}}name", "Tabelle")

        # Erkennung Hidden Sheet: table:print="false" oder Style mit display="false"
        table_print = table_el.get(f"{{{TABLE_NS}}}print", "true")
        table_style = table_el.get(f"{{{TABLE_NS}}}style-name", "")
        is_hidden = (table_print == "false") or (table_style in hidden_style_names)

        if is_hidden:
            hidden_sheet_names.append(sheet_name)
            sheet_heading = f"## Sheet: {sheet_name} [HIDDEN]"
        else:
            sheet_heading = f"## Sheet: {sheet_name}"

        sheet_parts: list[str] = [sheet_heading]

        # Zeilen aus table:row → Zellen aus table:cell / table:covered-table-cell
        table_rows: list[list[str]] = []

        for row_el in table_el.findall(f"{{{TABLE_NS}}}table-row"):
            row_data: list[str] = []

            for cell_el in row_el:
                cell_tag_local = cell_el.tag.split("}")[-1] if "}" in cell_el.tag else cell_el.tag
                if cell_tag_local not in ("table-cell", "covered-table-cell"):
                    continue

                # Wiederholungsfaktor berücksichtigen
                repeat = int(cell_el.get(f"{{{TABLE_NS}}}number-columns-repeated", "1"))

                # Formel auslesen
                formula = cell_el.get(f"{{{TABLE_NS}}}formula", "")

                # Zellinhalt: alle text:p Texte
                text_parts = []
                for p_el in cell_el.findall(f"{{{TEXT_NS}}}p"):
                    # Alle Textknoten rekursiv
                    p_text = "".join(p_el.itertext()).strip()
                    if p_text:
                        text_parts.append(p_text)
                cell_str = " ".join(text_parts)

                # Formel-Annotation
                if show_formulas and formula and formula.startswith("of:="):
                    formula_display = "=" + formula[4:]  # "of:=" → "="
                    cell_str = f"{cell_str} [{formula_display}]" if cell_str else f"[{formula_display}]"
                elif show_formulas and formula and formula.startswith("="):
                    cell_str = f"{cell_str} [{formula}]" if cell_str else f"[{formula}]"

                # Pipe-Zeichen escapen
                cell_str = cell_str.replace("|", "\\|").replace("\n", " ")

                # Wiederholungsfaktor: Zelle N-fach einfügen
                for _ in range(min(repeat, 100)):  # max 100 zum Schutz vor riesigen Dateien
                    row_data.append(cell_str)

            if row_data:
                table_rows.append(row_data)

        if not table_rows:
            sheet_parts.append("*Kein Inhalt*")
            markdown_parts.append("\n".join(sheet_parts))
            continue

        # Prüfen ob Sheet komplett leer ist
        all_empty = all(cell == "" for row in table_rows for cell in row)
        if all_empty:
            sheet_parts.append("*Kein Inhalt*")
            markdown_parts.append("\n".join(sheet_parts))
            continue

        # Maximale Spaltenanzahl
        max_cols = max(len(row) for row in table_rows)

        # Markdown-Tabelle
        header = table_rows[0]
        # Fehlende Header-Spalten mit generischen Namen auffüllen
        while len(header) < max_cols:
            header.append(f"Col{len(header) + 1}")
        display_header = [h if h else f"Col{i + 1}" for i, h in enumerate(header)]

        table_lines: list[str] = []
        table_lines.append("| " + " | ".join(display_header[:max_cols]) + " |")
        table_lines.append("| " + " | ".join(["---"] * max_cols) + " |")
        for data_row in table_rows[1:]:
            padded = data_row + [""] * (max_cols - len(data_row))
            table_lines.append("| " + " | ".join(padded[:max_cols]) + " |")

        sheet_parts.append("\n".join(table_lines))
        markdown_parts.append("\n\n".join(sheet_parts))

    combined_markdown = "\n\n".join(markdown_parts)
    sheets_count = len(list(root.iter(f"{{{TABLE_NS}}}table")))

    log.info(
        "ods_enhanced_convert_done",
        file=str(file_path),
        sheets=sheets_count,
        hidden_sheets=len(hidden_sheet_names),
        chars=len(combined_markdown),
    )

    result: dict = {
        "success": True,
        "markdown": combined_markdown,
        "sheets_count": sheets_count,
    }
    if hidden_sheet_names:
        result["hidden_sheets"] = hidden_sheet_names
    return result


# =============================================================================
# MarkItDown-basierte Konvertierung
# =============================================================================

def convert_with_markitdown(file_path: Path, show_formulas: bool = False) -> dict[str, Any]:
    """
    Konvertiert eine Datei mit MarkItDown.

    Für Excel-Dateien (.xlsx/.xls) wird convert_excel_enhanced() verwendet (FR-MKIT-007):
    - Multi-Sheet Ausgabe, Merged Cells, optionale Formel-Annotationen, Chart-Extraktion.

    Für PDF-Dateien wird zusätzlich pdfplumber für Tabellen-Extraktion genutzt:
    - Wenn pdfplumber Tabellen findet, werden diese als Markdown-Tabellen in den
      MarkItDown-Text integriert.
    - Wenn pdfplumber keine Tabellen findet, wird reiner MarkItDown-Output genutzt.

    Für PDF und DOCX wird nach der Konvertierung Code-Block-Erkennung angewendet
    (FR-MKIT-005): Indentierte Blöcke werden erkannt und in Fences gewrappt.

    Für PDF werden zusätzlich PyMuPDF-Metadaten extrahiert (FR-MKIT-009):
    - Bookmarks/TOC werden als Inhaltsverzeichnis vorangestellt.
    - Annotationen werden als Blockquotes angehängt.
    - Formularfelder werden als Key-Value-Tabelle angehängt.

    Args:
        file_path: Pfad zur Datei.
        show_formulas: Excel-Formeln annotieren (nur für .xlsx/.xls, FR-MKIT-007).
    """
    suffix = file_path.suffix.lower()

    # FR-MKIT-007: Excel → enhanced converter
    if suffix in {".xlsx", ".xls"}:
        return convert_excel_enhanced(file_path, show_formulas=show_formulas)

    try:
        log.info("markitdown_convert", file=str(file_path))
        result = _get("md", md).convert(str(file_path))
        markdown_text = result.text_content

        # PDF-spezifisch: pdfplumber für Tabellen-Extraktion
        if file_path.suffix.lower() == ".pdf" and _get("PDFPLUMBER_AVAILABLE", PDFPLUMBER_AVAILABLE):
            page_tables = _get("extract_tables_with_pdfplumber", extract_tables_with_pdfplumber)(file_path)
            if page_tables:
                merged_tables = _get("merge_cross_page_tables", merge_cross_page_tables)(page_tables)
                table_markdown = _get("tables_to_markdown", tables_to_markdown)(merged_tables)
                if table_markdown:
                    log.info(
                        "pdfplumber_tables_integrated",
                        file=str(file_path),
                        table_count=len(merged_tables),
                    )
                    markdown_text = markdown_text + "\n\n## Tabellen\n\n" + table_markdown
            elif _get("IMG2TABLE_AVAILABLE", IMG2TABLE_AVAILABLE):
                # AC-012-1: img2table als Fallback wenn pdfplumber keine Tabellen findet
                img2table_tables = _get("extract_tables_with_img2table", extract_tables_with_img2table)(file_path)
                if img2table_tables:
                    table_markdown = _get("tables_to_markdown", tables_to_markdown)(img2table_tables)
                    if table_markdown:
                        log.info(
                            "img2table_fallback_integrated",
                            file=str(file_path),
                            table_count=len(img2table_tables),
                            tables_source="img2table",
                        )
                        markdown_text = markdown_text + "\n\n## Tabellen\n\n" + table_markdown

        # FR-MKIT-005: Code-Block-Erkennung für PDF und DOCX
        suffix = file_path.suffix.lower()
        if suffix in {".pdf", ".docx", ".doc"}:
            markdown_text = detect_and_fence_code_blocks(markdown_text)
            log.debug("code_fencing_applied", file=str(file_path))

        # FR-MKIT-008: DOCX Extras (Kommentare, Header/Footer, Track Changes)
        if suffix == ".docx":
            try:
                extras = _get("extract_docx_extras", extract_docx_extras)(file_path)
                markdown_text = _get("append_docx_extras_to_markdown", append_docx_extras_to_markdown)(markdown_text, extras)
                log.debug("docx_extras_appended", file=str(file_path))
            except Exception as extras_err:
                log.warning(
                    "docx_extras_failed",
                    file=str(file_path),
                    error=str(extras_err),
                )

        # FR-MKIT-009: PDF-Metadaten (Bookmarks, Annotationen, Formularfelder)
        zugferd_data: Optional[dict[str, Any]] = None
        xmp_metadata_data: Optional[dict[str, Any]] = None
        embedded_files_data: Optional[list[dict[str, Any]]] = None
        if suffix == ".pdf":
            try:
                pdf_meta = _get("extract_pdf_metadata", extract_pdf_metadata)(file_path)
                markdown_text = _get("prepend_pdf_toc", prepend_pdf_toc)(markdown_text, pdf_meta.get("toc", []))
                markdown_text = _get("append_pdf_annotations", append_pdf_annotations)(markdown_text, pdf_meta.get("annotations", []))
                markdown_text = _get("append_pdf_form_fields", append_pdf_form_fields)(markdown_text, pdf_meta.get("form_fields", []))
                log.debug("pdf_metadata_appended", file=str(file_path))
            except Exception as pdf_meta_err:
                log.warning(
                    "pdf_metadata_failed",
                    file=str(file_path),
                    error=str(pdf_meta_err),
                )

            # Datei-Bytes einmalig lesen für T-MKIT-024 + T-MKIT-025
            pdf_file_bytes: Optional[bytes] = None
            try:
                pdf_file_bytes = file_path.read_bytes()
            except Exception as read_err:
                log.warning("pdf_bytes_read_failed", file=str(file_path), error=str(read_err))

            # T-MKIT-024: ZUGFeRD/Factur-X E-Rechnung aus eingebettetem XML extrahieren
            if pdf_file_bytes is not None:
                try:
                    zugferd_xml = _get("detect_zugferd", detect_zugferd)(pdf_file_bytes)
                    if zugferd_xml is not None:
                        zugferd_data = _get("parse_zugferd_xml", parse_zugferd_xml)(zugferd_xml)
                        log.info("zugferd_detected", file=str(file_path))
                except Exception as zugferd_err:
                    log.warning("zugferd_extract_failed", file=str(file_path), error=str(zugferd_err))

            # T-MKIT-025: XMP Metadata + Embedded Files extrahieren
            if pdf_file_bytes is not None:
                try:
                    xmp_metadata_data = _get("extract_xmp_metadata", extract_xmp_metadata)(pdf_file_bytes)
                    if xmp_metadata_data:
                        log.debug("xmp_metadata_extracted", file=str(file_path))
                except Exception as xmp_err:
                    log.warning("xmp_metadata_failed", file=str(file_path), error=str(xmp_err))

                try:
                    embedded_list = _get("list_embedded_files", list_embedded_files)(pdf_file_bytes)
                    if embedded_list:
                        embedded_files_data = embedded_list
                        log.debug("embedded_files_listed", file=str(file_path), count=len(embedded_list))
                except Exception as emb_err:
                    log.warning("embedded_files_failed", file=str(file_path), error=str(emb_err))

        # T-MKIT-028: Office Document Properties (DOCX, XLSX, PPTX)
        document_properties_data: Optional[dict[str, Any]] = None
        if suffix in {".docx", ".xlsx", ".pptx", ".odt", ".ods", ".odp"}:
            try:
                file_bytes_for_props = file_path.read_bytes()
                document_properties_data = extract_document_properties(
                    file_bytes_for_props, file_path.name
                )
                if document_properties_data:
                    log.debug("document_properties_extracted", file=str(file_path))
            except Exception as props_err:
                log.warning(
                    "document_properties_failed",
                    file=str(file_path),
                    error=str(props_err),
                )

        # T-MKIT-029: Email Metadata (Routing, Threading, Calendar Events)
        email_metadata_data: Optional[dict[str, Any]] = None
        if suffix == ".eml":
            try:
                eml_bytes = file_path.read_bytes()
                email_metadata_data = extract_email_metadata(eml_bytes)
                if email_metadata_data:
                    log.debug("email_metadata_extracted", file=str(file_path))
            except Exception as eml_err:
                log.warning(
                    "email_metadata_failed",
                    file=str(file_path),
                    error=str(eml_err),
                )

        # T-MKIT-030: PPTX Hidden Slides + Embedded Excel aus Charts
        pptx_hidden_info: Optional[dict[str, Any]] = None
        if suffix in {".pptx", ".ppt"}:
            try:
                pptx_bytes = file_path.read_bytes()
                pptx_hidden_info = extract_pptx_hidden_info(pptx_bytes)
                if pptx_hidden_info and pptx_hidden_info.get("hidden_slide_count", 0) > 0:
                    count = pptx_hidden_info["hidden_slide_count"]
                    numbers = pptx_hidden_info["hidden_slide_numbers"]
                    slides_str = ", ".join(str(n) for n in numbers)
                    markdown_text += (
                        f"\n\n---\n*Note: This presentation contains {count} "
                        f"hidden slide(s) (slides: {slides_str}).*"
                    )
                    log.info(
                        "pptx_hidden_slides_detected",
                        file=str(file_path),
                        count=count,
                        slides=numbers,
                    )
            except Exception as pptx_err:
                log.warning(
                    "pptx_hidden_info_failed",
                    file=str(file_path),
                    error=str(pptx_err),
                )

        return {
            "success": True,
            "markdown": markdown_text,
            "title": getattr(result, "title", None),
            "zugferd": zugferd_data,
            "xmp_metadata": xmp_metadata_data,
            "embedded_files": embedded_files_data,
            "document_properties": document_properties_data,
            "email_metadata": email_metadata_data,
            "pptx_hidden_info": pptx_hidden_info,
        }
    except Exception as e:
        log.error("markitdown_error", file=str(file_path), error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": f"MarkItDown Fehler: {str(e)}"
        }


# NOTE: convert_with_markitdown calls several helper functions that are defined
# in server.py and not yet extracted to this module:
#   - extract_tables_with_pdfplumber(file_path)
#   - merge_cross_page_tables(page_tables)
#   - tables_to_markdown(merged_tables)
#   - extract_tables_with_img2table(file_path)
#   - detect_and_fence_code_blocks(markdown_text)
#   - extract_pdf_metadata(file_path)
#   - prepend_pdf_toc(markdown_text, toc)
#   - append_pdf_annotations(markdown_text, annotations)
#   - append_pdf_form_fields(markdown_text, form_fields)
#   - detect_zugferd(pdf_file_bytes)
#   - parse_zugferd_xml(zugferd_xml)
#   - extract_xmp_metadata(pdf_file_bytes)
#   - list_embedded_files(pdf_file_bytes)
#   - extract_email_metadata(eml_bytes)
# These must be imported from server.py or extracted into their own modules
# before this file can be used standalone.
