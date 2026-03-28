#!/usr/bin/env python3
# Aktiviere uvloop für bessere Performance
try:
    import uvloop
    uvloop.install()
except ImportError:
    pass

"""
Daigestr — Document Intelligence Service v3.1.0

Bietet zwei Schnittstellen:
- MCP (Port 8080): Für Claude und andere MCP-Clients
- REST (Port 8081): Für n8n und andere HTTP-Clients

Features:
- Auto-Routing: Bilder → Vision, Dokumente → MarkItDown
- Bild-Resize vor Vision (spart Tokens)
- Folder-Konvertierung (alle Dateien in einem Ordner)
- URL-Konvertierung
- Retry-Logik für API-Calls
- Strukturiertes Logging
"""

import asyncio
import os
import io
import json
import base64
import hashlib
import mimetypes
import re
import sqlite3
import threading
import time
import logging
import zipfile
from datetime import datetime
from typing import Optional, Any
from pathlib import Path

import subprocess

import httpx
import uvicorn
import magic
import structlog
try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False
try:
    from pdf2image import convert_from_path
    PDF2IMAGE_AVAILABLE = True
except ImportError:
    PDF2IMAGE_AVAILABLE = False
from PIL import Image
from fastapi import FastAPI, HTTPException, Request
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse
from fastmcp import FastMCP
from markitdown import MarkItDown
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type

try:
    import jsonschema
    JSONSCHEMA_AVAILABLE = True
except ImportError:
    JSONSCHEMA_AVAILABLE = False

try:
    from faster_whisper import WhisperModel
    WHISPER_AVAILABLE = True
except ImportError:
    WHISPER_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    from img2table.document import PDF as Img2TablePDF
    from img2table.ocr import TesseractOCR
    IMG2TABLE_AVAILABLE = True
except ImportError:
    IMG2TABLE_AVAILABLE = False

try:
    from icalendar import Calendar as ICalendar
    ICALENDAR_AVAILABLE = True
except ImportError:
    ICALENDAR_AVAILABLE = False

import email as _email_stdlib
import email.policy as _email_policy

from models import (
    ConvertRequest,
    ConvertResponse,
    ConvertFolderRequest,
    AnalyzeRequest,
    ExtractRequest,
    TemplateResponse,
    HealthResponse,
    MetaData,
    ErrorCode,
    create_error_response,
    create_success_response,
)


# =============================================================================
# Konfiguration
# =============================================================================

DATA_DIR = Path(os.getenv("DATA_DIR", "/data"))
TEMP_DIR = Path(os.getenv("TEMP_DIR", "/tmp/markitdown"))
TEMP_DIR.mkdir(parents=True, exist_ok=True)

# Template Registry DB (T-MKIT-035)
TEMPLATES_DB_PATH = Path(os.getenv("TEMPLATES_DB_PATH", str(DATA_DIR / "templates.db")))

# Mistral Vision
MISTRAL_API_KEY = os.getenv("MISTRAL_API_KEY", "")
MISTRAL_API_URL = os.getenv("MISTRAL_API_URL", "https://api.mistral.ai/v1")
MISTRAL_VISION_MODEL = os.getenv("MISTRAL_VISION_MODEL", "mistral-small-2603")
MISTRAL_TIMEOUT = int(os.getenv("MISTRAL_TIMEOUT", "120"))

# Mistral OCR 3
MISTRAL_OCR_MODEL = os.getenv("MISTRAL_OCR_MODEL", "mistral-ocr-2512")
MISTRAL_OCR_ENABLED = os.getenv("MISTRAL_OCR_ENABLED", "true").lower() == "true"

# Server Ports
MCP_PORT = int(os.getenv("MCP_PORT", "8080"))
REST_PORT = int(os.getenv("REST_PORT", "8081"))

# Limits
MAX_FILE_SIZE_MB = int(os.getenv("MAX_FILE_SIZE_MB", "25"))
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024
IMAGE_MAX_WIDTH = int(os.getenv("IMAGE_MAX_WIDTH", "2048"))

# Retry
MAX_RETRIES = int(os.getenv("MAX_RETRIES", "3"))

# Scanned PDF Detection
SCAN_THRESHOLD_CHARS = int(os.getenv("SCAN_THRESHOLD_CHARS", "50"))

# Klassifizierung
DEFAULT_CLASSIFY_CATEGORIES = os.getenv(
    "CLASSIFY_CATEGORIES",
    "invoice,contract,cv,protocol,letter,technical_doc,report,presentation,spreadsheet,other"
).split(",")

# Cache für classify-Kategorien aus der Template-Registry
_classify_categories_cache: dict = {"categories": None, "timestamp": 0}
_CLASSIFY_CACHE_TTL = 300  # 5 Minuten

# Logging
LOG_LEVEL = os.getenv("LOG_LEVEL", "info").upper()
LOG_FORMAT = os.getenv("LOG_FORMAT", "json")

# Dateitypen
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
MARKITDOWN_EXTENSIONS = {".pdf", ".docx", ".doc", ".pptx", ".ppt", ".xlsx", ".xls",
                          ".odt", ".ods", ".odp", ".html", ".htm", ".xml", ".json",
                          ".csv", ".txt", ".md", ".rtf", ".eml"}
SKIP_FILES = set(os.getenv("SKIP_FILES", "email.md,consolidated.md,metadata.json,.DS_Store,Thumbs.db").split(","))

# Audio/Video (FR-MKIT-006)
WHISPER_MODEL_SIZE = os.getenv("WHISPER_MODEL_SIZE", "base")
AUDIO_EXTENSIONS = {".mp3", ".wav", ".ogg", ".flac", ".m4a"}
VIDEO_EXTENSIONS = {".mp4", ".mkv", ".webm", ".avi", ".mov"}

# Modelle — separate Modelle für verschiedene Tasks
MISTRAL_TEXT_MODEL = os.getenv("MISTRAL_TEXT_MODEL", "mistral-small-2603")

# Token-Limits — großzügig, Kosten nur pro verbrauchtem Token
VISION_MAX_TOKENS = int(os.getenv("VISION_MAX_TOKENS", "16384"))
CLASSIFY_MAX_TOKENS = int(os.getenv("CLASSIFY_MAX_TOKENS", "1024"))
EXTRACT_MAX_TOKENS = int(os.getenv("EXTRACT_MAX_TOKENS", "16384"))
OCR_CORRECT_MAX_TOKENS = int(os.getenv("OCR_CORRECT_MAX_TOKENS", "16384"))

# Text-Limits für LLM-Input
CLASSIFY_MAX_CHARS = int(os.getenv("CLASSIFY_MAX_CHARS", "32000"))
EXTRACT_MAX_CHARS = int(os.getenv("EXTRACT_MAX_CHARS", "32000"))

# Sprache
DEFAULT_LANGUAGE = os.getenv("DEFAULT_LANGUAGE", "de")

# Bild-Verarbeitung
MIN_IMAGE_SIZE_PX = int(os.getenv("MIN_IMAGE_SIZE_PX", "50"))
PDF_RENDER_DPI = int(os.getenv("PDF_RENDER_DPI", "200"))

# Timeouts
PDFTOTEXT_TIMEOUT = int(os.getenv("PDFTOTEXT_TIMEOUT", "60"))
PDFINFO_TIMEOUT = int(os.getenv("PDFINFO_TIMEOUT", "30"))
FFMPEG_TIMEOUT = int(os.getenv("FFMPEG_TIMEOUT", "600"))

# Whisper
WHISPER_DEVICE = os.getenv("WHISPER_DEVICE", "cpu")
WHISPER_COMPUTE_TYPE = os.getenv("WHISPER_COMPUTE_TYPE", "int8")

# Whisper-Modell-Cache (wird beim ersten Aufruf geladen)
_whisper_model_cache: dict[str, Any] = {}  # key: model_size → WhisperModel instance

VERSION = "3.0.0"
START_TIME = time.time()


# =============================================================================
# Logging Setup
# =============================================================================

def setup_logging():
    """Konfiguriert strukturiertes Logging."""
    structlog.configure(
        processors=[
            structlog.stdlib.filter_by_level,
            structlog.stdlib.add_logger_name,
            structlog.stdlib.add_log_level,
            structlog.processors.TimeStamper(fmt="iso"),
            structlog.processors.StackInfoRenderer(),
            structlog.processors.format_exc_info,
            structlog.processors.UnicodeDecoder(),
            structlog.processors.JSONRenderer() if LOG_FORMAT == "json" else structlog.dev.ConsoleRenderer(),
        ],
        context_class=dict,
        logger_factory=structlog.stdlib.LoggerFactory(),
        wrapper_class=structlog.stdlib.BoundLogger,
        cache_logger_on_first_use=True,
    )
    logging.basicConfig(
        format="%(message)s",
        level=getattr(logging, LOG_LEVEL, logging.INFO),
    )

setup_logging()
log = structlog.get_logger()


# =============================================================================
# Initialisierung
# =============================================================================

# MarkItDown Instanz
md = MarkItDown()

# FastMCP für MCP-Protokoll
mcp = FastMCP("daigestr")

# FastAPI für REST
app = FastAPI(
    title="Daigestr API",
    description="Konvertiert Dokumente und Bilder zu Markdown",
    version=VERSION,
    docs_url="/docs",
    redoc_url="/redoc",
)


def _safe_encode(obj: Any) -> Any:
    """Recursively converts bytes values to a safe string representation."""
    if isinstance(obj, bytes):
        try:
            return obj.decode("utf-8")
        except UnicodeDecodeError:
            return f"<binary {len(obj)} bytes>"
    if isinstance(obj, dict):
        return {k: _safe_encode(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_safe_encode(v) for v in obj]
    return obj


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError) -> JSONResponse:
    """
    Custom handler for request validation errors.

    FastAPI's default handler uses jsonable_encoder which calls bytes.decode(),
    causing UnicodeDecodeError when binary data (e.g. PDF uploads via multipart)
    appears in the error details.
    """
    safe_errors = _safe_encode(exc.errors())
    return JSONResponse(
        status_code=422,
        content={"detail": safe_errors},
    )


# =============================================================================
# Hilfsfunktionen
# =============================================================================

def resolve_path(path: str) -> Path:
    """Löst relativen Pfad zu absolutem auf."""
    p = Path(path)
    if p.is_absolute():
        return p
    return DATA_DIR / p


def get_file_extension(filename: str) -> str:
    """Extrahiert die Dateiendung."""
    return Path(filename).suffix.lower()


def is_image_file(path: Path) -> bool:
    """Prüft ob eine Datei ein unterstütztes Bild ist."""
    return path.suffix.lower() in IMAGE_EXTENSIONS


def is_markitdown_file(path: Path) -> bool:
    """Prüft ob eine Datei von MarkItDown verarbeitet werden kann."""
    return path.suffix.lower() in MARKITDOWN_EXTENSIONS


def is_audio_file(path: Path) -> bool:
    """Prüft ob eine Datei eine unterstützte Audio-Datei ist (FR-MKIT-006).

    Args:
        path: Pfad zur Datei.

    Returns:
        True wenn die Dateiendung in AUDIO_EXTENSIONS enthalten ist.
    """
    return path.suffix.lower() in AUDIO_EXTENSIONS


def is_video_file(path: Path) -> bool:
    """Prüft ob eine Datei eine unterstützte Video-Datei ist (FR-MKIT-006).

    Args:
        path: Pfad zur Datei.

    Returns:
        True wenn die Dateiendung in VIDEO_EXTENSIONS enthalten ist.
    """
    return path.suffix.lower() in VIDEO_EXTENSIONS


def extract_audio_from_video(video_path: Path) -> Path:
    """Extrahiert den Audio-Track aus einer Video-Datei als WAV via ffmpeg (FR-MKIT-006).

    Args:
        video_path: Pfad zur Video-Datei.

    Returns:
        Pfad zur extrahierten WAV-Datei im TEMP_DIR.

    Raises:
        RuntimeError: Wenn ffmpeg fehlschlägt oder nicht installiert ist.
    """
    wav_filename = f"{video_path.stem}_{hashlib.md5(str(video_path).encode()).hexdigest()[:8]}.wav"
    wav_path = TEMP_DIR / wav_filename

    log.info("extract_audio_from_video_start", video=str(video_path), output=str(wav_path))

    try:
        result = subprocess.run(
            [
                "ffmpeg", "-y",
                "-i", str(video_path),
                "-vn",               # kein Video
                "-acodec", "pcm_s16le",
                "-ar", "16000",      # 16kHz — optimal für Whisper
                "-ac", "1",          # Mono
                str(wav_path),
            ],
            capture_output=True,
            text=True,
            timeout=FFMPEG_TIMEOUT,
        )
        if result.returncode != 0:
            log.error(
                "extract_audio_from_video_failed",
                video=str(video_path),
                returncode=result.returncode,
                stderr=result.stderr,
            )
            raise RuntimeError(f"ffmpeg fehlgeschlagen (returncode={result.returncode}): {result.stderr}")
    except FileNotFoundError as exc:
        raise RuntimeError("ffmpeg ist nicht installiert") from exc
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError("ffmpeg-Timeout beim Audio-Extrahieren") from exc

    log.info("extract_audio_from_video_done", wav=str(wav_path))
    return wav_path


def transcribe_audio(audio_path: Path) -> dict[str, Any]:
    """Transkribiert eine Audio-Datei mit faster-whisper (FR-MKIT-006).

    Nutzt WHISPER_MODEL_SIZE aus der Umgebungsvariable.
    Das Modell wird gecacht — nach dem ersten Laden wird es wiederverwendet.

    Args:
        audio_path: Pfad zur Audio-Datei (WAV bevorzugt).

    Returns:
        Dict mit:
        - success (bool)
        - text (str): Vollständiges Transkript
        - language (str): Erkannte Sprache (z.B. "de", "en")
        - duration (float): Dauer in Sekunden
        - model_size (str): Verwendete Modell-Größe
        - error (str, optional): Fehlermeldung bei Misserfolg
    """
    if not WHISPER_AVAILABLE:
        log.warning("transcribe_audio_whisper_not_available", file=str(audio_path))
        return {
            "success": False,
            "error": "faster-whisper ist nicht installiert (pip install faster-whisper)",
        }

    log.info("transcribe_audio_start", file=str(audio_path), model_size=WHISPER_MODEL_SIZE)

    try:
        # Modell aus Cache laden oder frisch initialisieren
        if WHISPER_MODEL_SIZE not in _whisper_model_cache:
            log.info("whisper_model_load", model_size=WHISPER_MODEL_SIZE)
            _whisper_model_cache[WHISPER_MODEL_SIZE] = WhisperModel(
                WHISPER_MODEL_SIZE,
                device=WHISPER_DEVICE,
                compute_type=WHISPER_COMPUTE_TYPE,
            )
        model = _whisper_model_cache[WHISPER_MODEL_SIZE]

        segments, info = model.transcribe(str(audio_path), beam_size=5)

        # Segmente zusammenführen
        text_parts: list[str] = []
        duration = 0.0
        for segment in segments:
            text_parts.append(segment.text.strip())
            duration = max(duration, segment.end)

        full_text = " ".join(text_parts)
        detected_language = info.language if hasattr(info, "language") else "unknown"

        log.info(
            "transcribe_audio_done",
            file=str(audio_path),
            language=detected_language,
            duration=duration,
            chars=len(full_text),
        )

        return {
            "success": True,
            "text": full_text,
            "language": detected_language,
            "duration": duration,
            "model_size": WHISPER_MODEL_SIZE,
        }

    except Exception as exc:
        log.error("transcribe_audio_error", file=str(audio_path), error=str(exc))
        return {
            "success": False,
            "error": f"Transkription fehlgeschlagen: {str(exc)}",
        }


def should_skip_file(filename: str) -> bool:
    """Prüft ob eine Datei übersprungen werden soll."""
    return filename in SKIP_FILES or filename.startswith(".")


def get_mimetype(path: Path) -> str:
    """Ermittelt den MIME-Type einer Datei."""
    suffix = path.suffix.lower()
    mime_map = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".bmp": "image/bmp",
        ".pdf": "application/pdf",
    }
    return mime_map.get(suffix, "application/octet-stream")


def detect_mimetype_from_bytes(data: bytes) -> Optional[str]:
    """Erkennt MIME-Type aus Magic Bytes."""
    try:
        return magic.from_buffer(data, mime=True)
    except Exception:
        return None


def resize_image_if_needed(image_data: bytes, max_width: int = IMAGE_MAX_WIDTH) -> tuple[bytes, dict]:
    """
    Verkleinert ein Bild falls es zu groß ist.

    Returns:
        tuple: (image_bytes, resize_meta)
    """
    resize_meta = {
        "resized": False,
        "original_width": None,
        "original_height": None,
    }

    try:
        img = Image.open(io.BytesIO(image_data))
        resize_meta["original_width"] = img.width
        resize_meta["original_height"] = img.height

        if img.width > max_width:
            ratio = max_width / img.width
            new_height = int(img.height * ratio)
            img = img.resize((max_width, new_height), Image.Resampling.LANCZOS)
            resize_meta["resized"] = True
            resize_meta["width"] = max_width
            resize_meta["height"] = new_height

            output = io.BytesIO()
            img_format = "JPEG" if img.mode == "RGB" else "PNG"
            if img.mode == "RGBA" and img_format == "JPEG":
                img = img.convert("RGB")
            img.save(output, format=img_format, quality=85)
            return output.getvalue(), resize_meta
        else:
            resize_meta["width"] = img.width
            resize_meta["height"] = img.height
            return image_data, resize_meta

    except Exception as e:
        log.warning("image_resize_failed", error=str(e))
        return image_data, resize_meta


# =============================================================================
# Image EXIF/GPS/IPTC Metadata Extraction (T-MKIT-027)
# =============================================================================

def _dms_to_decimal(dms_tuple, ref: str) -> float:
    """Convert GPS DMS (degrees, minutes, seconds) to decimal degrees."""
    degrees, minutes, seconds = dms_tuple
    decimal = float(degrees) + float(minutes) / 60 + float(seconds) / 3600
    if ref in ("S", "W"):
        decimal = -decimal
    return round(decimal, 6)


def extract_image_metadata(image_data: bytes) -> dict:
    """
    Extrahiert EXIF-, GPS- und IPTC-Metadaten aus Bild-Bytes via Pillow.

    Args:
        image_data: Rohe Bild-Bytes (JPEG, PNG, etc.)

    Returns:
        Dict mit optionalen Schlüsseln 'exif', 'gps', 'iptc'.
        Leeres Dict wenn keine Metadaten vorhanden oder Fehler auftritt.
    """
    result: dict = {}
    try:
        img = Image.open(io.BytesIO(image_data))
        raw_exif = img.getexif()

        # --- EXIF-Basisdaten ---
        exif_data: dict = {}

        camera_make = raw_exif.get(271)   # Make
        if camera_make:
            exif_data["camera_make"] = str(camera_make).strip()

        camera_model = raw_exif.get(272)  # Model
        if camera_model:
            exif_data["camera_model"] = str(camera_model).strip()

        software = raw_exif.get(305)      # Software
        if software:
            exif_data["software"] = str(software).strip()

        datetime_original = raw_exif.get(36867)   # DateTimeOriginal
        if datetime_original:
            exif_data["datetime_original"] = str(datetime_original).strip()

        datetime_digitized = raw_exif.get(36868)  # DateTimeDigitized
        if datetime_digitized:
            exif_data["datetime_digitized"] = str(datetime_digitized).strip()

        # --- GPS ---
        try:
            gps_ifd = raw_exif.get_ifd(34853)  # GPSInfo IFD tag
            if gps_ifd:
                gps_data: dict = {}

                lat_dms = gps_ifd.get(2)   # GPSLatitude
                lat_ref = gps_ifd.get(1)   # GPSLatitudeRef
                lon_dms = gps_ifd.get(4)   # GPSLongitude
                lon_ref = gps_ifd.get(3)   # GPSLongitudeRef
                alt_val = gps_ifd.get(6)   # GPSAltitude
                alt_ref = gps_ifd.get(5)   # GPSAltitudeRef (0=above, 1=below sea level)

                if lat_dms and lat_ref:
                    gps_data["latitude"] = _dms_to_decimal(lat_dms, lat_ref)
                if lon_dms and lon_ref:
                    gps_data["longitude"] = _dms_to_decimal(lon_dms, lon_ref)
                if alt_val is not None:
                    try:
                        alt_decimal = float(alt_val)
                        if alt_ref == 1:
                            alt_decimal = -alt_decimal
                        gps_data["altitude"] = round(alt_decimal, 2)
                    except (TypeError, ValueError):
                        pass

                if gps_data:
                    exif_data.update(gps_data)
        except Exception:
            pass  # GPS gracefully skipped

        if exif_data:
            result["exif"] = exif_data

        # --- IPTC (via APP13 / photoshop block) ---
        try:
            photoshop_data = img.info.get("photoshop") or img.info.get("APP13")
            if photoshop_data and isinstance(photoshop_data, (bytes, bytearray)):
                iptc_data: dict = {}
                data = bytes(photoshop_data)
                i = 0
                while i < len(data) - 4:
                    # IPTC marker: 0x1C followed by record number and dataset number
                    if data[i] == 0x1C and data[i + 1] == 0x02:
                        dataset = data[i + 2]
                        length = (data[i + 3] << 8) | data[i + 4]
                        value_bytes = data[i + 5: i + 5 + length]
                        try:
                            value = value_bytes.decode("utf-8", errors="replace").strip()
                        except Exception:
                            value = ""
                        if dataset == 120 and value:   # Caption/Abstract
                            iptc_data["caption"] = value
                        elif dataset == 25 and value:  # Keywords (can repeat)
                            iptc_data.setdefault("keywords", []).append(value)
                        elif dataset == 116 and value: # Copyright Notice
                            iptc_data["copyright"] = value
                        elif dataset == 90 and value:  # City
                            iptc_data["city"] = value
                        elif dataset == 101 and value: # Country
                            iptc_data["country"] = value
                        i += 5 + length
                    else:
                        i += 1
                if iptc_data:
                    result["iptc"] = iptc_data
        except Exception:
            pass  # IPTC gracefully skipped

    except Exception as e:
        log.debug("extract_image_metadata_failed", error=str(e))

    return result


# =============================================================================
# LLM Artifact Cleanup (T-MKIT-016)
# =============================================================================

# Patterns for LLM preambles that should be stripped (German and English)
_PREAMBLE_PATTERNS = [
    # German patterns
    re.compile(
        r'^(?:hier\s+ist\s+(?:der|die|das)\s+[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'^(?:im\s+folgenden\s+[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'^(?:nachfolgend\s+[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'^(?:gerne[,!.]?\s+[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    # English patterns
    re.compile(
        r'^(?:here\s+is\s+(?:the\s+)?[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'^(?:below\s+is\s+(?:the\s+)?[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'^(?:the\s+following\s+[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'^(?:certainly[!,.]?\s*[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'^(?:sure[!,.]?\s*[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
]

# Pattern for full-output code block wrapping (```markdown ... ``` or ``` ... ```)
_FULL_CODEBLOCK_PATTERN = re.compile(
    r'^```(?:markdown|md)?\s*\n([\s\S]*?)\n```\s*$',
    re.IGNORECASE,
)


def strip_llm_artifacts(text: str) -> str:
    """
    Entfernt typische LLM-Artefakte aus dem Output.

    Entfernt:
    - Einleitende Preamble-Sätze ("Hier ist...", "Here is...", "Im Folgenden...", etc.)
    - ```markdown ... ``` Wrapping wenn der gesamte Output darin eingeschlossen ist
    - ``` ... ``` Wrapping wenn der gesamte Output darin eingeschlossen ist

    Codeblöcke INNERHALB des Textes werden NICHT entfernt.

    Args:
        text: LLM-Output-Text der bereinigt werden soll.

    Returns:
        Bereinigter Text.
    """
    if not text:
        return text

    result = text.strip()

    # Mermaid-Codeblöcke NICHT strippen — sie sind gewollter Output
    if result.startswith("```mermaid"):
        return result

    # 1. Outer code block wrapping entfernen (```markdown ... ``` oder ``` ... ```)
    codeblock_match = _FULL_CODEBLOCK_PATTERN.match(result)
    if codeblock_match:
        result = codeblock_match.group(1).strip()

    # 2. Preamble-Zeilen am Anfang entfernen
    for pattern in _PREAMBLE_PATTERNS:
        result = pattern.sub('', result)
        result = result.strip()

    return result


# =============================================================================
# API Calls mit Retry
# =============================================================================

@retry(
    stop=stop_after_attempt(MAX_RETRIES),
    wait=wait_exponential(multiplier=1, min=1, max=10),
    retry=retry_if_exception_type((httpx.TimeoutException, httpx.HTTPStatusError)),
    reraise=True,
)
async def call_mistral_vision_api(payload: dict) -> dict:
    """Ruft die Mistral Vision API mit Retry-Logik auf."""
    async with httpx.AsyncClient(timeout=float(MISTRAL_TIMEOUT)) as client:
        response = await client.post(
            f"{MISTRAL_API_URL}/chat/completions",
            headers={
                "Authorization": f"Bearer {MISTRAL_API_KEY}",
                "Content-Type": "application/json"
            },
            json=payload
        )
        response.raise_for_status()
        return response.json()


@retry(
    stop=stop_after_attempt(MAX_RETRIES),
    wait=wait_exponential(multiplier=1, min=1, max=10),
    retry=retry_if_exception_type((httpx.TimeoutException, httpx.HTTPStatusError)),
    reraise=True,
)
async def call_mistral_ocr_api(file_data: bytes, filename: str) -> dict:
    """Ruft die Mistral OCR API (/v1/ocr) auf."""
    b64 = base64.b64encode(file_data).decode("utf-8")
    payload = {
        "model": MISTRAL_OCR_MODEL,
        "document": {
            "type": "document_url",
            "document_url": f"data:application/pdf;base64,{b64}",
        },
    }
    async with httpx.AsyncClient(timeout=float(MISTRAL_TIMEOUT)) as client:
        response = await client.post(
            f"{MISTRAL_API_URL}/ocr",
            headers={
                "Authorization": f"Bearer {MISTRAL_API_KEY}",
                "Content-Type": "application/json",
            },
            json=payload,
        )
        response.raise_for_status()
        return response.json()


async def analyze_with_mistral_vision(
    image_data: bytes,
    mimetype: str,
    prompt: str,
    language: str = "de"
) -> dict[str, Any]:
    """Analysiert ein Bild mit Mistral Pixtral Vision API."""
    if not MISTRAL_API_KEY:
        return {
            "success": False,
            "error_code": ErrorCode.API_KEY_INVALID,
            "error": "MISTRAL_API_KEY nicht konfiguriert"
        }

    b64_image = base64.b64encode(image_data).decode("utf-8")
    data_url = f"data:{mimetype};base64,{b64_image}"

    system_prompt = (
        "Du bist ein präziser Assistent für Bild- und Dokumentenanalyse. "
        "Befolge ausschließlich die Anweisungen im User-Prompt. "
        "Antworte NICHT mit Einleitungssätzen, Erklärungen oder Code-Block-Wrapping — nur mit dem angeforderten Ergebnis."
        if language == "de"
        else "You are a precise assistant for image and document analysis. "
        "Follow only the instructions in the user prompt. "
        "Do NOT reply with introductions, explanations, or code-block wrapping — only the requested result."
    )

    payload = {
        "model": MISTRAL_VISION_MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": data_url}}
                ]
            }
        ],
        "max_tokens": VISION_MAX_TOKENS
    }

    try:
        log.info("vision_api_call", model=MISTRAL_VISION_MODEL, image_size=len(image_data))
        result = await call_mistral_vision_api(payload)

        content = result["choices"][0]["message"]["content"]
        content = strip_llm_artifacts(content)
        usage = result.get("usage", {})

        log.info("vision_api_success", tokens=usage.get("total_tokens", 0))
        return {
            "success": True,
            "markdown": content,
            "tokens_prompt": usage.get("prompt_tokens", 0),
            "tokens_completion": usage.get("completion_tokens", 0),
            "tokens_total": usage.get("total_tokens", 0),
            "vision_model": MISTRAL_VISION_MODEL,
        }

    except httpx.TimeoutException:
        log.error("vision_api_timeout", timeout=MISTRAL_TIMEOUT)
        return {
            "success": False,
            "error_code": ErrorCode.TIMEOUT,
            "error": f"Mistral API Timeout nach {MISTRAL_TIMEOUT}s"
        }
    except httpx.HTTPStatusError as e:
        error_detail = str(e)
        try:
            error_detail = e.response.json().get("error", {}).get("message", str(e))
        except Exception:
            pass
        log.error("vision_api_error", error=error_detail)
        return {
            "success": False,
            "error_code": ErrorCode.API_ERROR,
            "error": f"Mistral API Fehler: {error_detail}"
        }
    except Exception as e:
        log.error("vision_api_exception", error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.VISION_FAILED,
            "error": f"Vision-Analyse fehlgeschlagen: {str(e)}"
        }


async def dual_pass_validate(
    markdown: str,
    file_data: bytes,
    mimetype: str,
    language: str = "de",
) -> str:
    """
    Validiert und korrigiert OCR-extrahierten Markdown via Dual-Pass Vision-Vergleich.

    Schickt den OCR-extrahierten Markdown-Text zusammen mit dem Originalbild an die
    Vision-API. Das Modell vergleicht beides und korrigiert Fehler in Struktur,
    Tabellen-Spalten und Inhalt.

    Bei fehlender Vision-API oder Fehler wird der Original-Markdown zurückgegeben
    (graceful degradation).

    Args:
        markdown: Per OCR extrahierter Markdown-Text.
        file_data: Rohe Bytes der Originaldatei (Bild oder PDF-Seite als Bild).
        mimetype: MIME-Type der Datei (z.B. 'image/png', 'image/jpeg').
        language: Sprache für den Vision-Prompt (Standard: "de").

    Returns:
        Korrigierter Markdown-Text oder Original bei Fehler.
    """
    if not MISTRAL_API_KEY:
        log.warning("dual_pass_validate_skipped_no_api_key")
        return markdown

    prompt = (
        f"Hier ist ein per OCR extrahierter Text und das Originalbild. "
        f"Vergleiche beides und korrigiere Fehler in Struktur, Tabellen-Spalten und Inhalt. "
        f"Gib den korrigierten Markdown zurück. Antworte NUR mit dem korrigierten Text.\n\n"
        f"OCR-Text:\n{markdown}"
    )

    log.info("dual_pass_validate_start", mimetype=mimetype, markdown_len=len(markdown))

    try:
        result = await analyze_with_mistral_vision(
            image_data=file_data,
            mimetype=mimetype,
            prompt=prompt,
            language=language,
        )

        if result.get("success"):
            corrected = result.get("markdown", markdown)
            log.info("dual_pass_validate_done", original_len=len(markdown), corrected_len=len(corrected))
            return corrected
        else:
            log.warning(
                "dual_pass_validate_vision_failed",
                error=result.get("error", "unknown"),
            )
            return markdown

    except Exception as exc:
        log.warning("dual_pass_validate_exception", error=str(exc))
        return markdown


def extract_tables_with_pdfplumber(file_path: Path) -> list[dict]:
    """
    Extrahiert Tabellen aus einer PDF-Datei mit pdfplumber.

    Gibt eine Liste von Seitentabellen zurück, wobei jeder Eintrag Seitennummer
    und die gefundenen Tabellen (als Liste von Zeilen-Listen) enthält.

    Args:
        file_path: Pfad zur PDF-Datei.

    Returns:
        Liste mit Dictionaries der Form:
        [{"page": int, "tables": list[list[list[str | None]]]}]
        Leere Liste wenn pdfplumber nicht verfügbar oder keine Tabellen gefunden.
    """
    if not PDFPLUMBER_AVAILABLE:
        log.warning("pdfplumber_not_available")
        return []

    page_tables: list[dict] = []
    try:
        with pdfplumber.open(str(file_path)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                if tables:
                    page_tables.append({"page": page_num, "tables": tables})
        log.info(
            "pdfplumber_extracted",
            file=str(file_path),
            pages_with_tables=len(page_tables),
        )
    except Exception as e:
        log.warning("pdfplumber_error", file=str(file_path), error=str(e))
        return []

    return page_tables


def extract_tables_with_img2table(file_path: Path) -> list[list]:
    """
    Extrahiert Tabellen aus einer gescannten PDF-Datei mit img2table + TesseractOCR.

    Wird als Fallback verwendet wenn pdfplumber keine Tabellen in der Datei findet.
    Nutzt TesseractOCR als OCR-Backend für die Erkennung von Zellinhalten.

    Args:
        file_path: Pfad zur PDF-Datei.

    Returns:
        Flache Liste von Tabellen (jede Tabelle = Liste von Zeilen, jede Zeile = Liste
        von Strings). Leere Liste wenn img2table nicht verfügbar, keine Tabellen
        gefunden oder ein Fehler auftritt.
    """
    if not IMG2TABLE_AVAILABLE:
        log.warning("img2table_not_available")
        return []

    try:
        ocr = TesseractOCR()
        pdf_doc = Img2TablePDF(src=str(file_path))
        extracted = pdf_doc.extract_tables(ocr=ocr)

        all_tables: list[list] = []
        # extracted ist ein Dict: {page_index: [ExtractedTable, ...]}
        for _page_idx, page_tables in extracted.items():
            for extracted_table in page_tables:
                # ExtractedTable.df ist ein pandas DataFrame
                df = extracted_table.df
                # Erste Zeile als Header, Rest als Datenzeilen
                rows: list[list[str]] = []
                header = [str(col) if col is not None else "" for col in df.columns]
                rows.append(header)
                for _, row in df.iterrows():
                    rows.append([str(v) if v is not None else "" for v in row])
                all_tables.append(rows)

        log.info(
            "img2table_extracted",
            file=str(file_path),
            table_count=len(all_tables),
        )
        return all_tables

    except Exception as e:
        log.warning("img2table_error", file=str(file_path), error=str(e))
        return []


def merge_cross_page_tables(page_tables: list[dict]) -> list[list]:
    """
    Führt Tabellen die über Seitengrenzen hinweg gehen zusammen.

    Algorithmus:
    - Vergleicht die letzte Tabelle auf Seite N mit der ersten Tabelle auf Seite N+1.
    - Wenn beide die gleiche Spaltenanzahl haben: Zusammenführen.
    - Header-Deduplizierung: Wenn die erste Zeile auf der Folgeseite identisch mit
      der ersten Zeile der Ausgangstabelle ist, wird sie weggelassen.
    - Unterschiedliche Spaltenanzahl → separate Tabellen.

    Args:
        page_tables: Ausgabe von extract_tables_with_pdfplumber.

    Returns:
        Flache Liste aller (ggf. zusammengeführten) Tabellen als Zeilen-Listen.
    """
    if not page_tables:
        return []

    # Alle Tabellen mit ihrer Seitenreihenfolge sammeln
    # Jede Tabelle bekommt eine (Seite, Tabellen-Index)-Referenz
    all_tables: list[list] = []

    for page_entry in page_tables:
        for table in page_entry["tables"]:
            all_tables.append(table)

    if not all_tables:
        return []

    # Tabellen iterativ zusammenführen
    merged: list[list] = [all_tables[0]]

    for current_table in all_tables[1:]:
        last_merged = merged[-1]

        if not last_merged or not current_table:
            merged.append(current_table)
            continue

        last_col_count = len(last_merged[0]) if last_merged else 0
        curr_col_count = len(current_table[0]) if current_table else 0

        if last_col_count == curr_col_count and last_col_count > 0:
            # Gleiche Spaltenanzahl → potenzieller Merge
            last_header = last_merged[0]
            curr_header = current_table[0]

            # Header-Deduplizierung: erste Zeile identisch → überspringen
            rows_to_add = current_table
            if curr_header == last_header:
                rows_to_add = current_table[1:]

            if rows_to_add:
                merged[-1] = last_merged + rows_to_add
        else:
            # Unterschiedliche Spaltenanzahl → separate Tabelle
            merged.append(current_table)

    return merged


def tables_to_markdown(tables: list[list]) -> str:
    """
    Konvertiert extrahierte Tabellen in Markdown-Format.

    Jede Tabelle wird als Markdown-Tabelle mit Header-Trennzeile formatiert.
    None-Werte in Zellen werden als leere Strings behandelt.

    Args:
        tables: Liste von Tabellen, jede Tabelle ist eine Liste von Zeilen,
                jede Zeile ist eine Liste von Zellwerten (str | None).

    Returns:
        Zusammengefügter Markdown-String aller Tabellen, getrennt durch Leerzeilen.
    """
    if not tables:
        return ""

    markdown_parts: list[str] = []

    for table in tables:
        if not table:
            continue

        # Zeilen normalisieren: None → leerer String, alle Werte zu str
        normalized_rows: list[list[str]] = []
        for row in table:
            normalized_rows.append([str(cell) if cell is not None else "" for cell in row])

        if not normalized_rows:
            continue

        lines: list[str] = []

        # Header (erste Zeile)
        header = normalized_rows[0]
        lines.append("| " + " | ".join(header) + " |")

        # Trennzeile
        lines.append("| " + " | ".join(["---"] * len(header)) + " |")

        # Datenzeilen
        for row in normalized_rows[1:]:
            # Sicherstellen dass die Zeilenlänge mit dem Header übereinstimmt
            padded_row = row + [""] * (len(header) - len(row))
            lines.append("| " + " | ".join(padded_row[: len(header)]) + " |")

        markdown_parts.append("\n".join(lines))

    return "\n\n".join(markdown_parts)


def is_scanned_pdf(file_path: Path) -> bool:
    """
    Prüft ob eine PDF-Datei ein eingescanntes Dokument ist.

    Nutzt pdftotext (poppler-utils) um Text zu extrahieren und berechnet den
    Durchschnitt der Zeichen pro Seite. Wenn dieser Durchschnitt unter dem
    konfigurierten Schwellwert (SCAN_THRESHOLD_CHARS) liegt, gilt die PDF
    als Scan.

    Args:
        file_path: Pfad zur PDF-Datei.

    Returns:
        True wenn die Datei als Scan erkannt wurde, False sonst.
    """
    try:
        result = subprocess.run(
            ["pdftotext", str(file_path), "-"],
            capture_output=True,
            text=True,
            timeout=PDFTOTEXT_TIMEOUT,
        )
        if result.returncode != 0:
            log.warning(
                "pdftotext_failed",
                file=str(file_path),
                returncode=result.returncode,
                stderr=result.stderr,
            )
            return False

        extracted_text = result.stdout

        # Seitenanzahl ermitteln
        page_count_result = subprocess.run(
            ["pdfinfo", str(file_path)],
            capture_output=True,
            text=True,
            timeout=PDFINFO_TIMEOUT,
        )
        pages = 1
        if page_count_result.returncode == 0:
            for line in page_count_result.stdout.splitlines():
                if line.lower().startswith("pages:"):
                    try:
                        pages = int(line.split(":", 1)[1].strip())
                    except ValueError:
                        pass
                    break

        total_chars = len(extracted_text.strip())
        avg_chars_per_page = total_chars / max(pages, 1)

        is_scan = avg_chars_per_page < SCAN_THRESHOLD_CHARS
        log.info(
            "scan_detection",
            file=str(file_path),
            pages=pages,
            total_chars=total_chars,
            avg_chars_per_page=avg_chars_per_page,
            threshold=SCAN_THRESHOLD_CHARS,
            is_scan=is_scan,
        )
        return is_scan

    except FileNotFoundError:
        log.warning("pdftotext_not_found", file=str(file_path))
        return False
    except subprocess.TimeoutExpired:
        log.warning("pdftotext_timeout", file=str(file_path))
        return False
    except Exception as e:
        log.warning("scan_detection_error", file=str(file_path), error=str(e))
        return False


async def convert_scanned_pdf_ocr3(file_path: Path) -> dict[str, Any]:
    """
    Konvertiert ein gescanntes PDF via Mistral OCR 3 API (/v1/ocr).

    Args:
        file_path: Pfad zur PDF-Datei.

    Returns:
        Dict mit folgenden Schlüsseln:
        - success (bool)
        - markdown (str): Zusammengeführter Markdown-Text aller Seiten
        - ocr_model (str): Verwendetes OCR-Modell
        - pages (int): Anzahl verarbeiteter Seiten
        - error_code / error: Nur bei Fehler
    """
    if not MISTRAL_API_KEY:
        return {
            "success": False,
            "error_code": ErrorCode.API_KEY_INVALID,
            "error": "MISTRAL_API_KEY nicht konfiguriert",
        }

    log.info("ocr3_convert_start", file=str(file_path), model=MISTRAL_OCR_MODEL)

    try:
        file_data = file_path.read_bytes()
    except Exception as e:
        log.error("ocr3_read_failed", file=str(file_path), error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": f"Datei konnte nicht gelesen werden: {str(e)}",
        }

    try:
        result = await call_mistral_ocr_api(file_data, file_path.name)
    except httpx.TimeoutException:
        log.error("ocr3_timeout", timeout=MISTRAL_TIMEOUT)
        return {
            "success": False,
            "error_code": ErrorCode.TIMEOUT,
            "error": f"Mistral OCR API Timeout nach {MISTRAL_TIMEOUT}s",
        }
    except httpx.HTTPStatusError as e:
        error_detail = str(e)
        try:
            error_detail = e.response.json().get("error", {}).get("message", str(e))
        except Exception:
            pass
        log.error("ocr3_api_error", error=error_detail)
        return {
            "success": False,
            "error_code": ErrorCode.API_ERROR,
            "error": f"Mistral OCR API Fehler: {error_detail}",
        }
    except Exception as e:
        log.error("ocr3_exception", error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.API_ERROR,
            "error": f"OCR API Fehler: {str(e)}",
        }

    try:
        pages = result.get("pages", [])
    except (AttributeError, TypeError) as e:
        log.error("ocr3_invalid_response", error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.API_ERROR,
            "error": f"Ungültige OCR-API-Antwort: {str(e)}",
        }
    if not pages:
        log.warning("ocr3_no_pages", file=str(file_path))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": "OCR API lieferte keine Seiten zurück",
        }

    markdown_parts = []
    for page in pages:
        page_index = page.get("index", 0) + 1
        page_markdown = page.get("markdown", "")
        markdown_parts.append(f"## Seite {page_index}\n\n{page_markdown}")

    combined_markdown = "\n\n".join(markdown_parts)

    log.info("ocr3_convert_complete", file=str(file_path), pages=len(pages), model=MISTRAL_OCR_MODEL)

    return {
        "success": True,
        "markdown": combined_markdown,
        "ocr_model": MISTRAL_OCR_MODEL,
        "pages": len(pages),
    }


async def convert_scanned_pdf(file_path: Path, language: str = "de") -> dict[str, Any]:
    """
    Konvertiert ein eingescanntes PDF zu Markdown.

    Primär: Mistral OCR 3 API (/v1/ocr) — wenn MISTRAL_OCR_ENABLED=true.
    Fallback: Mistral Vision — Seiten werden als Bild gerendert und einzeln analysiert.

    Args:
        file_path: Pfad zur gescannten PDF-Datei.
        language: Sprache für den Vision-Prompt (Standard: "de", nur für Fallback relevant).

    Returns:
        Dict mit folgenden Schlüsseln:
        - success (bool)
        - markdown (str): Zusammengeführter Markdown-Text aller Seiten
        - scanned (bool): Immer True
        - pages_processed (int): Anzahl erfolgreich verarbeiteter Seiten (Vision-Pfad)
        - pages (int): Anzahl Seiten (OCR3-Pfad)
        - tokens_per_page (list[dict]): Token-Verbrauch pro Seite (Vision-Pfad)
        - tokens_total (int): Gesamter Token-Verbrauch (Vision-Pfad)
        - vision_model (str): Genutztes Vision-Modell (Vision-Pfad)
        - ocr_model (str): Genutztes OCR-Modell (OCR3-Pfad)
        - error_code / error: Nur bei Fehler
    """
    # Primärer Pfad: Mistral OCR 3
    if MISTRAL_OCR_ENABLED:
        log.info("scanned_pdf_using_ocr3", file=str(file_path), model=MISTRAL_OCR_MODEL)
        ocr3_result = await convert_scanned_pdf_ocr3(file_path)
        if ocr3_result.get("success"):
            log.info("scanned_pdf_ocr3_success", file=str(file_path), pages=ocr3_result.get("pages", 0))
            return {
                "success": True,
                "markdown": ocr3_result["markdown"],
                "scanned": True,
                "pages_processed": ocr3_result.get("pages", 0),
                "ocr_model": ocr3_result.get("ocr_model", MISTRAL_OCR_MODEL),
            }
        else:
            log.warning(
                "scanned_pdf_ocr3_failed_fallback_to_vision",
                file=str(file_path),
                error=ocr3_result.get("error", "unknown"),
            )
            # Fallback: Vision-Pfad (weiter unten)
    else:
        log.info("scanned_pdf_ocr3_disabled_using_vision", file=str(file_path))

    # Fallback / direkter Pfad: Mistral Vision (pdf2image)
    if not PDF2IMAGE_AVAILABLE:
        log.error("pdf2image_not_available", file=str(file_path))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": "pdf2image ist nicht installiert (pip install pdf2image)",
        }

    if not MISTRAL_API_KEY:
        return {
            "success": False,
            "error_code": ErrorCode.API_KEY_INVALID,
            "error": "MISTRAL_API_KEY nicht konfiguriert",
        }

    log.info("scanned_pdf_convert_start", file=str(file_path))

    try:
        pages = convert_from_path(str(file_path), dpi=PDF_RENDER_DPI)
    except Exception as e:
        log.error("pdf2image_convert_failed", file=str(file_path), error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": f"PDF-Rendering fehlgeschlagen: {str(e)}",
        }

    vision_prompt = (
        "Extrahiere den gesamten Text aus diesem Scan einer PDF-Seite und gib ihn als Markdown zurück.\n\n"
        "Regeln:\n"
        "- Behalte die Dokumentsprache bei — übersetze NICHT\n"
        "- Überschriften → # ## ### Markdown-Syntax\n"
        "- Tabellen → immer als Markdown-Tabelle mit | Spalte | Spalte | und Trennzeile\n"
        "- Listen → - oder 1. Markdown-Syntax\n"
        "- Fußnoten, Seitenzahlen und Kopfzeilen → kursiv in eckigen Klammern, z.B. *[Seite 3]*\n"
        "- Wenn eine Passage unleserlich ist → schreibe [UNLESERLICH]\n"
        "- Wenn die Seite keine Textinhalte enthält → antworte nur mit: [LEERE SEITE]\n\n"
        "Antworte ausschließlich mit dem Markdown-Text."
    )

    markdown_parts: list[str] = []
    tokens_per_page: list[dict] = []
    total_prompt_tokens = 0
    total_completion_tokens = 0
    total_tokens = 0
    pages_processed = 0
    vision_model_used = MISTRAL_VISION_MODEL

    for page_num, page_image in enumerate(pages, start=1):
        # PIL Image → bytes (PNG)
        img_buffer = io.BytesIO()
        # Resize wenn nötig, um Tokens zu sparen
        if page_image.width > IMAGE_MAX_WIDTH:
            ratio = IMAGE_MAX_WIDTH / page_image.width
            new_height = int(page_image.height * ratio)
            page_image = page_image.resize((IMAGE_MAX_WIDTH, new_height), Image.Resampling.LANCZOS)

        page_image.save(img_buffer, format="PNG")
        image_bytes = img_buffer.getvalue()

        log.info(
            "scanned_pdf_page_start",
            file=str(file_path),
            page=page_num,
            total_pages=len(pages),
            image_size=len(image_bytes),
        )

        vision_result = await analyze_with_mistral_vision(
            image_bytes,
            "image/png",
            vision_prompt,
            language,
        )

        page_token_info = {
            "page": page_num,
            "tokens_prompt": vision_result.get("tokens_prompt", 0),
            "tokens_completion": vision_result.get("tokens_completion", 0),
            "tokens_total": vision_result.get("tokens_total", 0),
            "success": vision_result.get("success", False),
        }
        tokens_per_page.append(page_token_info)

        if vision_result.get("success"):
            pages_processed += 1
            total_prompt_tokens += vision_result.get("tokens_prompt", 0)
            total_completion_tokens += vision_result.get("tokens_completion", 0)
            total_tokens += vision_result.get("tokens_total", 0)
            vision_model_used = vision_result.get("vision_model", MISTRAL_VISION_MODEL)
            markdown_parts.append(f"## Seite {page_num}\n\n{vision_result['markdown']}")
            log.info(
                "scanned_pdf_page_done",
                file=str(file_path),
                page=page_num,
                tokens=vision_result.get("tokens_total", 0),
            )
        else:
            log.warning(
                "scanned_pdf_page_failed",
                file=str(file_path),
                page=page_num,
                error=vision_result.get("error", "unknown"),
            )
            markdown_parts.append(
                f"## Seite {page_num}\n\n*Seite konnte nicht verarbeitet werden: "
                f"{vision_result.get('error', 'Vision-Fehler')}*"
            )

    combined_markdown = "\n\n".join(markdown_parts)

    log.info(
        "scanned_pdf_convert_complete",
        file=str(file_path),
        pages_total=len(pages),
        pages_processed=pages_processed,
        tokens_total=total_tokens,
    )

    return {
        "success": True,
        "markdown": combined_markdown,
        "scanned": True,
        "pages_processed": pages_processed,
        "tokens_per_page": tokens_per_page,
        "tokens_prompt": total_prompt_tokens,
        "tokens_completion": total_completion_tokens,
        "tokens_total": total_tokens,
        "vision_model": vision_model_used,
    }


# =============================================================================
# Code-Block-Erkennung + Sprach-Fences (FR-MKIT-005)
# =============================================================================

# Mapping: Sprachname → Liste von Erkennungs-Patterns (case-sensitive Regex)
_LANGUAGE_PATTERNS: list[tuple[str, list[str]]] = [
    ("python",     [r"\bdef ", r"\bimport ", r"\bclass ", r"if __name__", r"\bprint\(", r"\bself\."]),
    ("javascript", [r"\bfunction ", r"\bconst ", r"\blet ", r"=> ", r"\bconsole\.log"]),
    ("java",       [r"\bpublic class\b", r"\bprivate ", r"\bSystem\.out\b", r"\bvoid "]),
    ("sql",        [r"\bSELECT\b", r"\bFROM\b", r"\bWHERE\b", r"\bINSERT INTO\b"]),
    ("html",       [r"<html", r"<div", r"<body", r"<!DOCTYPE"]),
    ("css",        [r"\{color:", r"\bmargin:", r"\bpadding:", r"\bdisplay:"]),
    ("bash",       [r"#!/bin/", r"\becho ", r"if \[", r"\bfi\b", r"\bdone\b"]),
    ("go",         [r"\bfunc ", r"\bpackage ", r"\bimport \(", r"\bfmt\."]),
    ("rust",       [r"\bfn ", r"\blet mut\b", r"\bimpl ", r"\bpub fn\b"]),
    ("cpp",        [r"#include", r"\bint main\b", r"\bprintf\(", r"\bstd::"]),
]

# Minimum score (number of pattern matches) to identify a language
_MIN_LANG_SCORE = 2


def detect_code_language(text: str) -> str:
    """
    Erkennt die Programmiersprache eines Code-Blocks via Regex-Heuristik.

    Args:
        text: Zu analysierender Text-Block.

    Returns:
        Sprachname in Kleinbuchstaben (z.B. "python", "javascript") oder ""
        wenn die Sprache nicht mit ausreichender Sicherheit erkannt werden kann.
    """
    best_lang = ""
    best_score = 0

    for lang, patterns in _LANGUAGE_PATTERNS:
        score = sum(1 for pat in patterns if re.search(pat, text))
        if score > best_score:
            best_score = score
            best_lang = lang

    return best_lang if best_score >= _MIN_LANG_SCORE else ""


def detect_and_fence_code_blocks(markdown: str) -> str:
    """
    Sucht in Markdown-Text nach nicht-gefenctem Code und wrapp ihn in
    ```language ... ``` Fences.

    Erkennungs-Kriterien:
    1. Indentierte Blöcke (4+ Leerzeichen am Zeilenanfang), mindestens 3 Zeilen.
    2. Blöcke mit mindestens 3 Code-Indikatoren (Klammern, Semikolons, Zuweisungen).

    Bereits vorhandene Fences (``` ... ```) werden nicht erneut gewrappt.

    Args:
        markdown: Markdown-Text nach der Konvertierung.

    Returns:
        Markdown-Text mit gefencten Code-Blöcken.
    """
    if not markdown:
        return markdown

    # Schritt 1: Vorhandene Fences aus dem Text ausblenden, damit wir sie nicht
    # versehentlich als Kandidaten erkennen.
    # Wir ersetzen sie durch Platzhalter und stellen sie am Ende wieder her.
    fenced_blocks: list[str] = []
    fence_pattern = re.compile(r"```[\s\S]*?```", re.MULTILINE)

    def _stash_fence(m: re.Match) -> str:
        idx = len(fenced_blocks)
        fenced_blocks.append(m.group(0))
        return f"\x00FENCE{idx}\x00"

    working = fence_pattern.sub(_stash_fence, markdown)

    # Schritt 2: Kandidaten-Blöcke identifizieren.
    # Ein Block ist eine zusammenhängende Gruppe von Zeilen mit 4+ Spaces Einrückung
    # ODER eine Gruppe von Zeilen mit Code-Indikatoren.
    lines = working.split("\n")
    result_lines: list[str] = []
    i = 0

    while i < len(lines):
        line = lines[i]

        # Sammle zusammenhängende Zeilen mit 4+ Spaces
        if re.match(r"^    ", line) and line.strip():
            block_lines: list[str] = []
            j = i
            while j < len(lines) and (re.match(r"^    ", lines[j]) or lines[j].strip() == ""):
                block_lines.append(lines[j])
                j += 1

            # Trailing leere Zeilen aus dem Block entfernen
            while block_lines and not block_lines[-1].strip():
                block_lines.pop()

            # Mindestens 3 nicht-leere Zeilen
            non_empty = [l for l in block_lines if l.strip()]
            if len(non_empty) >= 3:
                block_text = "\n".join(block_lines)
                lang = detect_code_language(block_text)
                fence_open = f"```{lang}" if lang else "```"
                result_lines.append(fence_open)
                result_lines.extend(block_lines)
                result_lines.append("```")
                log.debug(
                    "code_block_fenced",
                    lines=len(non_empty),
                    language=lang or "unknown",
                )
            else:
                result_lines.extend(block_lines)

            i = j
            continue

        result_lines.append(line)
        i += 1

    working = "\n".join(result_lines)

    # Schritt 3: Platzhalter durch originale Fences ersetzen
    for idx, original in enumerate(fenced_blocks):
        working = working.replace(f"\x00FENCE{idx}\x00", original)

    return working


def convert_excel_enhanced(file_path: Path, show_formulas: bool = False) -> dict[str, Any]:
    """
    Konvertiert eine Excel-Datei (.xlsx/.xls) zu Markdown mit erweiterten Features.

    Features (FR-MKIT-007):
    - AC-007-1: Jedes Worksheet als eigene Sektion (## Sheet: Name)
    - AC-007-2: Charts werden als Datentabellen extrahiert (via openpyxl)
    - AC-007-3: Zellen mit Formeln optional annotiert (z.B. 42 [=SUM(A1:A10)])
    - AC-007-4: Merged Cells werden korrekt aufgelöst

    Args:
        file_path: Pfad zur Excel-Datei (.xlsx oder .xls)
        show_formulas: Wenn True, werden Formel-Annotationen (z.B. [=SUM(A1:A10)]) hinzugefügt.

    Returns:
        Dict mit:
        - success (bool)
        - markdown (str): Konvertierter Markdown-Text
        - sheets_count (int): Anzahl der verarbeiteten Sheets
        - charts_count (int): Anzahl der gefundenen Charts
        - error_code / error: Nur bei Fehler
    """
    if not OPENPYXL_AVAILABLE:
        log.warning("openpyxl_not_available", file=str(file_path))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": "openpyxl ist nicht installiert (pip install openpyxl)",
        }

    log.info("excel_enhanced_convert_start", file=str(file_path), show_formulas=show_formulas)

    try:
        # data_only=False um Formeln zu lesen; ein zweites Mal mit data_only=True für Werte
        wb_formulas = openpyxl.load_workbook(str(file_path), data_only=False)
        wb_values = openpyxl.load_workbook(str(file_path), data_only=True)
    except Exception as exc:
        log.error("excel_open_error", file=str(file_path), error=str(exc))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": f"Excel-Datei konnte nicht geöffnet werden: {str(exc)}",
        }

    markdown_parts: list[str] = []
    total_charts = 0
    hidden_sheet_names: list[str] = []

    # T-MKIT-026: Separate visible from hidden/veryHidden sheets
    # Process visible sheets first, then hidden sheets as a secondary section
    visible_sheet_names: list[str] = []
    hidden_state_map: dict[str, str] = {}  # sheet_name → "hidden" | "veryHidden"

    for sheet_name in wb_values.sheetnames:
        ws = wb_values[sheet_name]
        state = getattr(ws, "sheet_state", "visible")
        if state == "visible":
            visible_sheet_names.append(sheet_name)
        else:
            hidden_state_map[sheet_name] = state
            hidden_sheet_names.append(sheet_name)

    # Process visible first, then hidden — preserving original order within each group
    all_ordered_names = visible_sheet_names + hidden_sheet_names

    for sheet_name in all_ordered_names:
        ws_values = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name]

        # T-MKIT-026: Annotate hidden/veryHidden sheets
        state = hidden_state_map.get(sheet_name, "visible")
        if state == "veryHidden":
            sheet_heading = f"## Sheet: {sheet_name} [VERY HIDDEN]"
        elif state == "hidden":
            sheet_heading = f"## Sheet: {sheet_name} [HIDDEN]"
        else:
            sheet_heading = f"## Sheet: {sheet_name}"

        sheet_parts: list[str] = [sheet_heading]

        # AC-007-4: Merged Cells auflösen
        # Baue ein Dict: (row, col) → Wert der Hauptzelle des Merge-Bereichs
        merged_cell_values: dict[tuple[int, int], Any] = {}
        for merge_range in ws_values.merged_cells.ranges:
            # Wert der linken oberen Zelle des Bereichs
            top_left = ws_values.cell(merge_range.min_row, merge_range.min_col)
            top_left_formula = ws_formulas.cell(merge_range.min_row, merge_range.min_col)
            for row_idx in range(merge_range.min_row, merge_range.max_row + 1):
                for col_idx in range(merge_range.min_col, merge_range.max_col + 1):
                    merged_cell_values[(row_idx, col_idx)] = (
                        top_left.value,
                        top_left_formula.value,
                    )

        # Zeilen und Spalten ermitteln
        rows = list(ws_values.iter_rows())
        formula_rows = list(ws_formulas.iter_rows())

        # Leeres Sheet graceful behandeln (AC: test_empty_sheet_handled)
        if not rows:
            sheet_parts.append("*Kein Inhalt*")
            markdown_parts.append("\n".join(sheet_parts))
            log.debug("excel_empty_sheet", sheet=sheet_name)
            continue

        # Maximale Spaltenanzahl bestimmen (über alle Zeilen)
        max_cols = max((len(row) for row in rows), default=0)

        if max_cols == 0:
            sheet_parts.append("*Kein Inhalt*")
            markdown_parts.append("\n".join(sheet_parts))
            continue

        # Tabellen-Daten aufbauen
        table_rows: list[list[str]] = []

        for row_idx, (row_cells, formula_cells) in enumerate(
            zip(rows, formula_rows), start=1
        ):
            row_data: list[str] = []
            for col_idx, (cell, formula_cell) in enumerate(
                zip(row_cells, formula_cells), start=1
            ):
                # AC-007-4: Merged Cell Wert holen
                if (row_idx, col_idx) in merged_cell_values:
                    raw_value, raw_formula = merged_cell_values[(row_idx, col_idx)]
                else:
                    raw_value = cell.value
                    raw_formula = formula_cell.value

                # Zellwert als String
                cell_str = "" if raw_value is None else str(raw_value)

                # AC-007-3: Formel-Annotation
                if (
                    show_formulas
                    and raw_formula is not None
                    and isinstance(raw_formula, str)
                    and raw_formula.startswith("=")
                ):
                    cell_str = f"{cell_str} [{raw_formula}]"

                # Pipe-Zeichen im Zellinhalt escapen (würde Tabelle brechen)
                cell_str = cell_str.replace("|", "\\|").replace("\n", " ")
                row_data.append(cell_str)

            # Fehlende Spalten auffüllen
            while len(row_data) < max_cols:
                row_data.append("")

            table_rows.append(row_data)

        # Prüfen ob Sheet komplett leer ist (alle Zellen None)
        all_empty = all(
            cell_str == ""
            for row in table_rows
            for cell_str in row
        )
        if all_empty:
            sheet_parts.append("*Kein Inhalt*")
            markdown_parts.append("\n".join(sheet_parts))
            continue

        # Markdown-Tabelle aufbauen
        if table_rows:
            header = table_rows[0]
            # Header mit generischen Spaltennamen falls leer
            display_header = [
                h if h else get_column_letter(i + 1)
                for i, h in enumerate(header)
            ]
            table_lines: list[str] = []
            table_lines.append("| " + " | ".join(display_header) + " |")
            table_lines.append("| " + " | ".join(["---"] * max_cols) + " |")
            for data_row in table_rows[1:]:
                padded = data_row + [""] * (max_cols - len(data_row))
                table_lines.append("| " + " | ".join(padded[:max_cols]) + " |")
            sheet_parts.append("\n".join(table_lines))

        # AC-007-2: Charts extrahieren
        sheet_charts = getattr(ws_values, "_charts", [])
        if sheet_charts:
            for chart_idx, chart in enumerate(sheet_charts, start=1):
                total_charts += 1
                chart_title = ""
                try:
                    if hasattr(chart, "title") and chart.title is not None:
                        title_obj = chart.title
                        # openpyxl chart.title kann str, Title-Objekt oder None sein
                        if isinstance(title_obj, str):
                            chart_title = title_obj
                        elif hasattr(title_obj, "tx") and title_obj.tx is not None:
                            # Title-Objekt mit tx.rich.p[].r[].t Struktur
                            try:
                                texts = []
                                for para in title_obj.tx.rich.p:
                                    for run in para.r:
                                        if run.t:
                                            texts.append(run.t)
                                chart_title = " ".join(texts)
                            except Exception:
                                chart_title = f"Chart {chart_idx}"
                        else:
                            chart_title = f"Chart {chart_idx}"
                    else:
                        chart_title = f"Chart {chart_idx}"
                except Exception:
                    chart_title = f"Chart {chart_idx}"

                chart_parts: list[str] = [f"### Chart: {chart_title}"]

                # Datenserien extrahieren
                try:
                    series_list = []
                    if hasattr(chart, "series"):
                        for serie in chart.series:
                            serie_title = ""
                            try:
                                if hasattr(serie, "title") and serie.title is not None:
                                    st = serie.title
                                    if hasattr(st, "v") and st.v is not None:
                                        serie_title = str(st.v)
                                    elif hasattr(st, "strRef") and st.strRef is not None:
                                        cache = getattr(st.strRef, "strCache", None)
                                        if cache and hasattr(cache, "pt") and cache.pt:
                                            serie_title = str(cache.pt[0].v)
                            except Exception:
                                pass

                            # Werte aus dem Cache holen
                            values: list[str] = []
                            try:
                                val_ref = None
                                if hasattr(serie, "val"):
                                    val_ref = serie.val
                                elif hasattr(serie, "yVal"):
                                    val_ref = serie.yVal

                                if val_ref is not None:
                                    num_cache = getattr(val_ref, "numRef", None)
                                    if num_cache:
                                        num_data = getattr(num_cache, "numCache", None)
                                        if num_data and hasattr(num_data, "pt"):
                                            values = [str(pt.v) for pt in num_data.pt]
                            except Exception:
                                pass

                            series_list.append({
                                "title": serie_title or f"Serie {len(series_list) + 1}",
                                "values": values,
                            })

                    if series_list:
                        # Tabelle mit Serien als Spalten
                        headers = [s["title"] for s in series_list]
                        chart_parts.append("| " + " | ".join(headers) + " |")
                        chart_parts.append("| " + " | ".join(["---"] * len(headers)) + " |")
                        max_vals = max((len(s["values"]) for s in series_list), default=0)
                        for vi in range(max_vals):
                            row_vals = []
                            for s in series_list:
                                row_vals.append(s["values"][vi] if vi < len(s["values"]) else "")
                            chart_parts.append("| " + " | ".join(row_vals) + " |")
                    else:
                        chart_parts.append("*Keine Datenserien gefunden*")

                except Exception as chart_exc:
                    log.warning(
                        "excel_chart_extraction_error",
                        sheet=sheet_name,
                        chart=chart_idx,
                        error=str(chart_exc),
                    )
                    chart_parts.append("*Chart-Daten konnten nicht extrahiert werden*")

                sheet_parts.append("\n".join(chart_parts))

        markdown_parts.append("\n\n".join(sheet_parts))

    wb_values.close()
    wb_formulas.close()

    combined_markdown = "\n\n".join(markdown_parts)
    sheets_count = len(wb_values.sheetnames)

    log.info(
        "excel_enhanced_convert_done",
        file=str(file_path),
        sheets=sheets_count,
        charts=total_charts,
        chars=len(combined_markdown),
        hidden_sheets=len(hidden_sheet_names),
    )

    result: dict[str, Any] = {
        "success": True,
        "markdown": combined_markdown,
        "sheets_count": sheets_count,
        "charts_count": total_charts,
    }
    # T-MKIT-026: Include hidden_sheets only when there are hidden sheets
    if hidden_sheet_names:
        result["hidden_sheets"] = hidden_sheet_names
    return result


# =============================================================================
# DOCX Extras: Kommentare, Header/Footer, Track Changes (FR-MKIT-008)
# =============================================================================

def extract_docx_extras(file_path: Path) -> dict:
    """
    Extrahiert erweiterte Metadaten aus einer DOCX-Datei (FR-MKIT-008).

    Verarbeitet:
    - Kommentare aus word/comments.xml (Author, Date, Text)
    - Header und Footer aus allen Dokumentsektionen via python-docx
    - Track Changes (Einfügungen/Löschungen) aus dem DOCX-XML (w:ins, w:del)

    Args:
        file_path: Pfad zur DOCX-Datei.

    Returns:
        Dict mit:
        - comments: Liste von Dicts mit 'author', 'date', 'text'
        - headers: Liste von Header-Texten (nicht leer)
        - footers: Liste von Footer-Texten (nicht leer)
        - track_changes: Liste von Dicts mit 'type' ('insertion'/'deletion'), 'author', 'date', 'text'
    """
    import xml.etree.ElementTree as ET

    result: dict = {
        "comments": [],
        "headers": [],
        "footers": [],
        "track_changes": [],
    }

    W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"

    # --- Kommentare aus word/comments.xml ---
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            if "word/comments.xml" in zf.namelist():
                xml_data = zf.read("word/comments.xml")
                root = ET.fromstring(xml_data)
                for comment in root.findall(f"{{{W_NS}}}comment"):
                    author = comment.get(f"{{{W_NS}}}author", "")
                    date = comment.get(f"{{{W_NS}}}date", "")
                    # Text aus allen w:t Elementen zusammenführen
                    text_parts = [t.text or "" for t in comment.findall(f".//{{{W_NS}}}t")]
                    text = " ".join(text_parts).strip()
                    if text:
                        result["comments"].append({
                            "author": author,
                            "date": date,
                            "text": text,
                        })
                log.info(
                    "docx_comments_extracted",
                    file=str(file_path),
                    count=len(result["comments"]),
                )
    except Exception as e:
        log.warning("docx_comments_error", file=str(file_path), error=str(e))

    # --- Header/Footer via python-docx ---
    try:
        import docx as _docx  # python-docx
        doc = _docx.Document(str(file_path))
        for section in doc.sections:
            for hf_obj, target_list in [
                (section.header, result["headers"]),
                (section.footer, result["footers"]),
            ]:
                try:
                    text = hf_obj.text.strip() if hf_obj and hasattr(hf_obj, "text") else ""
                    if not text:
                        # Manuell aus Paragraphen extrahieren
                        if hf_obj and hasattr(hf_obj, "paragraphs"):
                            parts = [p.text.strip() for p in hf_obj.paragraphs if p.text.strip()]
                            text = " | ".join(parts)
                    if text and text not in target_list:
                        target_list.append(text)
                except Exception as inner_e:
                    log.debug("docx_hf_paragraph_error", error=str(inner_e))
        log.info(
            "docx_headers_footers_extracted",
            file=str(file_path),
            headers=len(result["headers"]),
            footers=len(result["footers"]),
        )
    except Exception as e:
        log.warning("docx_headers_footers_error", file=str(file_path), error=str(e))

    # --- Track Changes aus word/document.xml (w:ins, w:del) ---
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            if "word/document.xml" in zf.namelist():
                xml_data = zf.read("word/document.xml")
                root = ET.fromstring(xml_data)

                for tag, change_type in [("ins", "insertion"), ("del", "deletion")]:
                    for elem in root.findall(f".//{{{W_NS}}}{tag}"):
                        author = elem.get(f"{{{W_NS}}}author", "")
                        date = elem.get(f"{{{W_NS}}}date", "")
                        # w:delText für Löschungen, w:t für Einfügungen
                        if change_type == "deletion":
                            text_parts = [
                                t.text or ""
                                for t in elem.findall(f".//{{{W_NS}}}delText")
                            ]
                        else:
                            text_parts = [
                                t.text or ""
                                for t in elem.findall(f".//{{{W_NS}}}t")
                            ]
                        text = "".join(text_parts)
                        if text:
                            result["track_changes"].append({
                                "type": change_type,
                                "author": author,
                                "date": date,
                                "text": text,
                            })

                log.info(
                    "docx_track_changes_extracted",
                    file=str(file_path),
                    count=len(result["track_changes"]),
                )
    except Exception as e:
        log.warning("docx_track_changes_error", file=str(file_path), error=str(e))

    return result


def append_docx_extras_to_markdown(markdown: str, extras: dict) -> str:
    """
    Fügt DOCX-Extras (Kommentare, Header/Footer, Track Changes) als Markdown-Sektionen an.

    Sektionen werden nur angefügt, wenn die jeweiligen Extras nicht leer sind:
    - ## Kommentare → Blockquotes mit Author und Date
    - ## Header und Footer → Inhalt als Liste
    - ## Änderungsverfolgung → Einfügungen und Löschungen als Diff-Notation

    Args:
        markdown: Bereits konvertierter Markdown-Text.
        extras: Rückgabe-Dict von extract_docx_extras().

    Returns:
        Erweiterter Markdown-String.
    """
    sections: list[str] = []

    # --- Kommentare als Blockquotes ---
    comments = extras.get("comments", [])
    if comments:
        lines = ["## Kommentare", ""]
        for c in comments:
            author = c.get("author", "Unbekannt")
            date = c.get("date", "")
            text = c.get("text", "")
            date_str = f" ({date})" if date else ""
            lines.append(f"> **{author}**{date_str}: {text}")
            lines.append("")
        sections.append("\n".join(lines).rstrip())

    # --- Header und Footer ---
    headers = extras.get("headers", [])
    footers = extras.get("footers", [])
    if headers or footers:
        lines = ["## Header und Footer", ""]
        if headers:
            lines.append("**Header:**")
            for h in headers:
                lines.append(f"- {h}")
            lines.append("")
        if footers:
            lines.append("**Footer:**")
            for f in footers:
                lines.append(f"- {f}")
            lines.append("")
        sections.append("\n".join(lines).rstrip())

    # --- Track Changes als Diff-Notation ---
    track_changes = extras.get("track_changes", [])
    if track_changes:
        lines = ["## Änderungsverfolgung", ""]
        lines.append("```diff")
        for tc in track_changes:
            change_type = tc.get("type", "")
            author = tc.get("author", "")
            date = tc.get("date", "")
            text = tc.get("text", "")
            meta = f"  # {author}" if author else ""
            if date:
                meta += f" ({date})" if author else f"  # ({date})"
            if change_type == "insertion":
                lines.append(f"+ {text}{meta}")
            elif change_type == "deletion":
                lines.append(f"- {text}{meta}")
        lines.append("```")
        sections.append("\n".join(lines))

    if not sections:
        return markdown

    return markdown.rstrip() + "\n\n" + "\n\n".join(sections)


# =============================================================================
# FR-MKIT-009: PDF-Metadaten (Bookmarks, Annotationen, Formularfelder)
# =============================================================================

def extract_pdf_metadata(file_path: Path) -> dict[str, Any]:
    """
    Extrahiert PDF-Metadaten mit PyMuPDF (fitz).

    Liefert:
    - toc: Inhaltsverzeichnis/Bookmarks als Liste von [level, title, page].
    - annotations: Annotationen/Kommentare mit type, content, author, page.
    - form_fields: Formularfelder mit field_name, field_value, field_type, page.

    Wenn PyMuPDF nicht installiert ist, werden leere Listen zurückgegeben.

    Args:
        file_path: Pfad zur PDF-Datei.

    Returns:
        Dict mit Schlüsseln "toc", "annotations", "form_fields".
    """
    if not PYMUPDF_AVAILABLE:
        log.debug("pymupdf_not_available", file=str(file_path))
        return {"toc": [], "annotations": [], "form_fields": []}

    toc: list[list[Any]] = []
    annotations: list[dict[str, Any]] = []
    form_fields: list[dict[str, Any]] = []

    try:
        doc = fitz.open(str(file_path))

        # Bookmarks / Table of Contents
        toc = doc.get_toc()  # Returns list of [level, title, page]

        # Annotationen und Formularfelder seitenweise
        for page_num, page in enumerate(doc, start=1):
            # Annotationen
            for annot in page.annots():
                info = annot.info
                annot_entry: dict[str, Any] = {
                    "page": page_num,
                    "type": annot.type[1] if annot.type else "Unknown",
                    "content": info.get("content", ""),
                    "author": info.get("title", ""),
                }
                annotations.append(annot_entry)

            # Formularfelder (Widgets)
            for widget in page.widgets():
                field_entry: dict[str, Any] = {
                    "page": page_num,
                    "field_name": widget.field_name or "",
                    "field_value": str(widget.field_value) if widget.field_value is not None else "",
                    "field_type": widget.field_type_string or str(widget.field_type),
                }
                form_fields.append(field_entry)

        doc.close()
        log.debug(
            "pdf_metadata_extracted",
            file=str(file_path),
            toc_entries=len(toc),
            annotations=len(annotations),
            form_fields=len(form_fields),
        )

    except Exception as exc:
        log.warning("pdf_metadata_error", file=str(file_path), error=str(exc))

    return {"toc": toc, "annotations": annotations, "form_fields": form_fields}


# =============================================================================
# ZUGFeRD / Factur-X E-Rechnung (T-MKIT-024)
# =============================================================================

# Bekannte ZUGFeRD/Factur-X Dateinamen in eingebetteten PDF-Anhängen
_ZUGFERD_FILENAMES = [
    "factur-x.xml",       # Factur-X / ZUGFeRD 2.x
    "zugferd-invoice.xml",  # ZUGFeRD 1.x (lowercase variant)
    "ZUGFeRD-invoice.xml",  # ZUGFeRD 1.x (original case)
    "xrechnung.xml",      # XRechnung
]

# Namespaces für ZUGFeRD 2.x / Factur-X (CrossIndustryInvoice:100)
_ZUGFERD_NS_V2 = {
    "rsm": "urn:un:unece:uncefact:data:standard:CrossIndustryInvoice:100",
    "ram": "urn:un:unece:uncefact:data:standard:ReusableAggregateBusinessInformationEntity:100",
    "udt": "urn:un:unece:uncefact:data:standard:UnqualifiedDataType:100",
}

# Namespaces für ZUGFeRD 1.x
_ZUGFERD_NS_V1 = {
    "rsm": "urn:ferd:CrossIndustryDocument:invoice:1p0:comfort",
    "ram": "urn:un:unece:uncefact:data:standard:ReusableAggregateBusinessInformationEntity:12",
    "udt": "urn:un:unece:uncefact:data:standard:UnqualifiedDataType:15",
}


def detect_zugferd(file_data: bytes) -> Optional[bytes]:
    """
    Prüft ob ein PDF eingebettete ZUGFeRD/Factur-X XML-Daten enthält.

    Öffnet das PDF mit PyMuPDF und prüft die eingebetteten Dateien (embfile_names)
    auf bekannte ZUGFeRD-Dateinamen. Gibt die XML-Bytes zurück wenn gefunden.

    Args:
        file_data: Rohe PDF-Bytes.

    Returns:
        XML-Bytes wenn ZUGFeRD-Anhang gefunden, sonst None.
    """
    if not PYMUPDF_AVAILABLE:
        return None

    try:
        doc = fitz.open(stream=file_data, filetype="pdf")
        embedded_names = doc.embfile_names()

        for name in embedded_names:
            # Case-insensitive Prüfung gegen bekannte ZUGFeRD-Dateinamen
            for zugferd_name in _ZUGFERD_FILENAMES:
                if name.lower() == zugferd_name.lower():
                    xml_info = doc.embfile_get(name)
                    doc.close()
                    # embfile_get gibt ein dict zurück mit 'content' (Bytes)
                    if isinstance(xml_info, dict):
                        return xml_info.get("content")
                    # Ältere PyMuPDF-Versionen können direkt Bytes zurückgeben
                    if isinstance(xml_info, bytes):
                        return xml_info
                    return None

        doc.close()
        return None

    except Exception as exc:
        log.warning("zugferd_detect_error", error=str(exc))
        return None


def extract_xmp_metadata(file_data: bytes) -> Optional[dict[str, Any]]:
    """
    Extrahiert XMP-Metadaten aus einem PDF.

    Kombiniert PyMuPDF doc.metadata (Basis-Infos) mit XMP-Stream-Daten:
    - pdf_a_level (aus pdfaid:part + pdfaid:conformance, z.B. "3B")
    - document_id (xmpMM:DocumentID)
    - instance_id (xmpMM:InstanceID)
    - creator_tool (xmp:CreatorTool)

    Args:
        file_data: Rohe PDF-Bytes.

    Returns:
        Dict mit Metadaten-Feldern oder None wenn PyMuPDF nicht verfügbar.
    """
    if not PYMUPDF_AVAILABLE:
        return None

    try:
        doc = fitz.open(stream=file_data, filetype="pdf")

        # Basis-Metadaten aus PyMuPDF
        base_meta = doc.metadata or {}
        result: dict[str, Any] = {}

        for key in ("title", "author", "subject", "creator", "producer", "creationDate", "modDate"):
            val = base_meta.get(key)
            if val:
                result[key] = val

        # XMP-Stream parsen
        xmp_xml = doc.get_xml_metadata()
        doc.close()

        if xmp_xml:
            try:
                from lxml import etree as ET  # noqa: PLC0415

                root = ET.fromstring(xmp_xml.encode("utf-8") if isinstance(xmp_xml, str) else xmp_xml)

                def _find_text(tag_local: str, ns_uri: str) -> Optional[str]:
                    """Sucht ein Element per Namespace-URI und lokalem Namen."""
                    for el in root.iter():
                        if el.tag == f"{{{ns_uri}}}{tag_local}":
                            return (el.text or "").strip() or None
                    return None

                NS_PDFAID = "http://www.aiim.org/pdfa/ns/id/"
                NS_XMPMM = "http://ns.adobe.com/xap/1.0/mm/"
                NS_XMP = "http://ns.adobe.com/xap/1.0/"

                part = _find_text("part", NS_PDFAID)
                conformance = _find_text("conformance", NS_PDFAID)
                if part:
                    result["pdf_a_level"] = part + (conformance or "")

                doc_id = _find_text("DocumentID", NS_XMPMM)
                if doc_id:
                    result["document_id"] = doc_id

                inst_id = _find_text("InstanceID", NS_XMPMM)
                if inst_id:
                    result["instance_id"] = inst_id

                creator_tool = _find_text("CreatorTool", NS_XMP)
                if creator_tool:
                    result["creator_tool"] = creator_tool

            except Exception as xmp_err:
                log.debug("xmp_parse_error", error=str(xmp_err))

        return result if result else None

    except Exception as exc:
        log.warning("xmp_metadata_extract_error", error=str(exc))
        return None


# ZUGFeRD-Dateinamen die NICHT in embedded_files erscheinen sollen
_ZUGFERD_FILENAMES_SET = {name.lower() for name in (
    "factur-x.xml",
    "zugferd-invoice.xml",
    "zugferd-invoice.xml",
    "xrechnung.xml",
    "order-x.xml",
    "zugferd_1p0_comfort.xml",
    "zugferd-invoice.xml",
)}


def list_embedded_files(file_data: bytes) -> list[dict[str, Any]]:
    """
    Listet alle in einem PDF eingebetteten Dateien auf.

    Gibt NICHT die ZUGFeRD-Dateien zurück (die werden von T-MKIT-024 separat
    behandelt). Verwendet PyMuPDF embfile_count(), embfile_names() und embfile_info().

    Args:
        file_data: Rohe PDF-Bytes.

    Returns:
        Liste von Dicts mit name, size, description. Leere Liste wenn keine Dateien
        eingebettet sind oder PyMuPDF nicht verfügbar ist.
    """
    if not PYMUPDF_AVAILABLE:
        return []

    try:
        doc = fitz.open(stream=file_data, filetype="pdf")
        names = doc.embfile_names()
        result: list[dict[str, Any]] = []

        for name in names:
            # ZUGFeRD-Dateien überspringen (werden von T-MKIT-024 separat behandelt)
            if name.lower() in _ZUGFERD_FILENAMES_SET:
                continue

            try:
                info = doc.embfile_info(name)
                entry: dict[str, Any] = {"name": name}
                if isinstance(info, dict):
                    size = info.get("size") or info.get("length")
                    if size is not None:
                        entry["size"] = size
                    desc = info.get("desc") or info.get("description")
                    if desc:
                        entry["description"] = desc
                result.append(entry)
            except Exception as info_err:
                log.debug("embfile_info_error", name=name, error=str(info_err))
                result.append({"name": name})

        doc.close()
        return result

    except Exception as exc:
        log.warning("list_embedded_files_error", error=str(exc))
        return []


def parse_zugferd_xml(xml_data: bytes) -> dict[str, Any]:
    """
    Parst ZUGFeRD/Factur-X XML und extrahiert strukturierte Rechnungsdaten.

    Unterstützt ZUGFeRD 2.x / Factur-X (Namespace CrossIndustryInvoice:100)
    und ZUGFeRD 1.x (ältere Namespaces). Alle Felder sind Optional — fehlendes
    XML liefert None-Werte statt Exceptions.

    Args:
        xml_data: ZUGFeRD/Factur-X XML als Bytes.

    Returns:
        Dict mit extrahierten Rechnungsfeldern (BT-Nummern nach EN 16931).
    """
    try:
        from lxml import etree as ET  # noqa: PLC0415
    except ImportError:
        log.warning("zugferd_lxml_not_available")
        return {"parse_error": "lxml not available"}

    result: dict[str, Any] = {
        "invoice_number": None,
        "invoice_date": None,
        "invoice_type": None,
        "seller_name": None,
        "seller_address": None,
        "seller_vat_id": None,
        "buyer_name": None,
        "buyer_address": None,
        "currency": None,
        "total_net": None,
        "total_vat": None,
        "total_gross": None,
        "due_amount": None,
        "payment_reference": None,
        "iban": None,
        "bic": None,
        "payment_terms": None,
        "due_date": None,
        "line_items": [],
    }

    try:
        root = ET.fromstring(xml_data)
    except Exception as exc:
        log.warning("zugferd_xml_parse_error", error=str(exc))
        result["parse_error"] = f"XML parse error: {str(exc)}"
        return result

    # Namespace-Erkennung: V2 oder V1?
    root_ns = root.nsmap.get("rsm") or ""
    if "CrossIndustryInvoice:100" in root_ns:
        NS = _ZUGFERD_NS_V2
    elif "ferd" in root_ns or "CrossIndustryDocument" in root_ns:
        NS = _ZUGFERD_NS_V1
    else:
        # Versuche V2 als Default
        NS = _ZUGFERD_NS_V2

    def _text(element, xpath: str) -> Optional[str]:
        """Extrahiert den Text des ersten XPath-Treffers oder None."""
        try:
            nodes = element.xpath(xpath, namespaces=NS)
            if nodes:
                node = nodes[0]
                text = node.text if hasattr(node, "text") else str(node)
                return text.strip() if text else None
        except Exception:
            pass
        return None

    def _build_address(party_element) -> Optional[str]:
        """Baut eine lesbare Adresse aus Adressfeldern zusammen."""
        try:
            addr = party_element.xpath(".//ram:PostalTradeAddress", namespaces=NS)
            if not addr:
                return None
            a = addr[0]
            parts = []
            for field in ["ram:LineOne", "ram:LineTwo", "ram:PostcodeCode", "ram:CityName", "ram:CountryID"]:
                val = _text(a, field)
                if val:
                    parts.append(val)
            return ", ".join(parts) if parts else None
        except Exception:
            return None

    try:
        # --- ExchangedDocument (Kopfdaten) ---
        doc_elem = root.find("rsm:ExchangedDocument", NS)
        if doc_elem is not None:
            result["invoice_number"] = _text(doc_elem, "ram:ID")
            result["invoice_type"] = _text(doc_elem, "ram:TypeCode")
            # Datum: BT-2 in DateTimeString Format 102 = YYYYMMDD
            raw_date = _text(doc_elem, "ram:IssueDateTime/udt:DateTimeString")
            if raw_date and len(raw_date) == 8:
                result["invoice_date"] = f"{raw_date[:4]}-{raw_date[4:6]}-{raw_date[6:8]}"
            elif raw_date:
                result["invoice_date"] = raw_date

        # --- SupplyChainTradeTransaction ---
        txn = root.find("rsm:SupplyChainTradeTransaction", NS)
        if txn is None:
            return result

        # --- Seller (BT-27) ---
        seller_party = txn.find(".//ram:SellerTradeParty", NS)
        if seller_party is not None:
            result["seller_name"] = _text(seller_party, "ram:Name")
            result["seller_address"] = _build_address(seller_party)
            # VAT-ID: BT-31 schemeID="VA"
            for tax_reg in seller_party.findall(".//ram:SpecifiedTaxRegistration", NS):
                scheme = tax_reg.find("ram:ID", NS)
                if scheme is not None and scheme.get("schemeID") == "VA":
                    result["seller_vat_id"] = scheme.text.strip() if scheme.text else None
                    break

        # --- Buyer (BT-44) ---
        buyer_party = txn.find(".//ram:BuyerTradeParty", NS)
        if buyer_party is not None:
            result["buyer_name"] = _text(buyer_party, "ram:Name")
            result["buyer_address"] = _build_address(buyer_party)

        # --- Settlement (Zahlungsinfos) ---
        settlement = txn.find(".//ram:ApplicableHeaderTradeSettlement", NS)
        if settlement is not None:
            result["currency"] = _text(settlement, "ram:InvoiceCurrencyCode")
            result["payment_reference"] = _text(settlement, "ram:PaymentReference")

            # Fälligkeitsdatum: BT-9
            raw_due = _text(settlement, ".//ram:DueDateDateTime/udt:DateTimeString")
            if raw_due and len(raw_due) == 8:
                result["due_date"] = f"{raw_due[:4]}-{raw_due[4:6]}-{raw_due[6:8]}"
            elif raw_due:
                result["due_date"] = raw_due

            # Zahlungsbedingungen: BT-20
            result["payment_terms"] = _text(settlement, ".//ram:SpecifiedTradePaymentTerms/ram:Description")

            # IBAN / BIC
            result["iban"] = _text(settlement, ".//ram:PayeePartyCreditorFinancialAccount/ram:IBANID")
            result["bic"] = _text(settlement, ".//ram:PayeeSpecifiedCreditorFinancialInstitution/ram:BICID")

            # Beträge: Monetary Summation
            summation = settlement.find(".//ram:SpecifiedTradeSettlementHeaderMonetarySummation", NS)
            if summation is not None:
                result["total_net"] = _text(summation, "ram:TaxBasisTotalAmount")
                result["total_gross"] = _text(summation, "ram:GrandTotalAmount")
                result["due_amount"] = _text(summation, "ram:DuePayableAmount")

            # Steuerbetrag: BT-110 (mit currencyID Attribut)
            tax_total_elem = settlement.find(".//ram:TaxTotalAmount", NS)
            if tax_total_elem is not None and tax_total_elem.text:
                result["total_vat"] = tax_total_elem.text.strip()

        # --- Line Items ---
        line_items = []
        for item in txn.findall(".//ram:IncludedSupplyChainTradeLineItem", NS):
            li: dict[str, Any] = {
                "description": None,
                "quantity": None,
                "unit_price": None,
                "total": None,
                "vat_rate": None,
            }
            # Beschreibung
            li["description"] = _text(item, ".//ram:SpecifiedTradeProduct/ram:Name")
            # Menge
            li["quantity"] = _text(item, ".//ram:BilledQuantity")
            # Einzelpreis
            li["unit_price"] = _text(item, ".//ram:ChargeAmount")
            # Zeilensumme
            li["total"] = _text(item, ".//ram:LineTotalAmount")
            # MwSt-Satz
            li["vat_rate"] = _text(item, ".//ram:ApplicableTradeTax/ram:RateApplicablePercent")
            line_items.append(li)

        result["line_items"] = line_items

    except Exception as exc:
        log.warning("zugferd_xml_extract_error", error=str(exc))
        result["parse_error"] = f"extraction error: {str(exc)}"

    return result


def prepend_pdf_toc(markdown: str, toc: list[list[Any]]) -> str:
    """
    Fügt ein Markdown-Inhaltsverzeichnis aus PDF-Bookmarks VOR dem Inhalt ein.

    Level-Mapping: Level 1 → ##, Level 2 → ###, Level 3 → ####, etc.

    Args:
        markdown: Bestehender Markdown-Text.
        toc: Liste von [level, title, page] aus fitz.Document.get_toc().

    Returns:
        Markdown mit vorangestelltem Inhaltsverzeichnis.
    """
    if not toc:
        return markdown

    lines = ["## Inhaltsverzeichnis", ""]
    for entry in toc:
        if len(entry) >= 3:
            level, title, page = entry[0], entry[1], entry[2]
        elif len(entry) == 2:
            level, title, page = entry[0], entry[1], None
        else:
            continue
        # Level 1 → ##, Level 2 → ###, ...
        prefix = "#" * (level + 1)
        page_str = f" *(Seite {page})*" if page and page > 0 else ""
        lines.append(f"{prefix} {title}{page_str}")

    toc_block = "\n".join(lines)
    return toc_block + "\n\n" + markdown


def append_pdf_annotations(markdown: str, annotations: list[dict[str, Any]]) -> str:
    """
    Hängt PDF-Annotationen als Blockquotes an den Markdown-Text an.

    Format:
    ## Annotationen
    > **Author** (Seite N, Typ): Content

    Args:
        markdown: Bestehender Markdown-Text.
        annotations: Liste von Dicts mit page, type, content, author.

    Returns:
        Markdown mit angehängter Annotationen-Sektion.
    """
    if not annotations:
        return markdown

    lines = ["## Annotationen", ""]
    for ann in annotations:
        author = ann.get("author", "")
        page = ann.get("page", "")
        ann_type = ann.get("type", "")
        content = ann.get("content", "")

        author_str = f"**{author}**" if author else "**Unbekannt**"
        meta_parts = []
        if page:
            meta_parts.append(f"Seite {page}")
        if ann_type:
            meta_parts.append(ann_type)
        meta_str = f" ({', '.join(meta_parts)})" if meta_parts else ""

        lines.append(f"> {author_str}{meta_str}: {content}")
        lines.append("")

    section = "\n".join(lines).rstrip()
    return markdown.rstrip() + "\n\n" + section


def append_pdf_form_fields(markdown: str, form_fields: list[dict[str, Any]]) -> str:
    """
    Hängt PDF-Formularfelder als Key-Value-Tabelle an den Markdown-Text an.

    Format:
    ## Formularfelder
    | Feld | Wert | Typ | Seite |
    |------|------|-----|-------|
    | name | wert | typ | 1     |

    Args:
        markdown: Bestehender Markdown-Text.
        form_fields: Liste von Dicts mit field_name, field_value, field_type, page.

    Returns:
        Markdown mit angehängter Formularfelder-Sektion.
    """
    if not form_fields:
        return markdown

    lines = [
        "## Formularfelder",
        "",
        "| Feld | Wert | Typ | Seite |",
        "|------|------|-----|-------|",
    ]
    for field in form_fields:
        name = field.get("field_name", "")
        value = field.get("field_value", "")
        ftype = field.get("field_type", "")
        page = field.get("page", "")
        lines.append(f"| {name} | {value} | {ftype} | {page} |")

    section = "\n".join(lines)
    return markdown.rstrip() + "\n\n" + section


# =============================================================================
# Office Document Properties Extraction (T-MKIT-028)
# =============================================================================

def extract_document_properties(file_data: bytes, filename: str) -> Optional[dict[str, Any]]:
    """
    Extrahiert Core-, App- und Custom-Properties aus Office-Dokumenten (DOCX, XLSX, PPTX).

    Office-Dokumente sind ZIP-Archive mit docProps/*.xml Dateien. Diese Funktion
    parst die drei relevanten XML-Dateien und gibt ein strukturiertes Dict zurück.

    Args:
        file_data: Rohe Datei-Bytes der Office-Datei
        filename: Dateiname (wird für Extension-Erkennung genutzt)

    Returns:
        Dict mit Schlüsseln 'core', 'app', 'custom' oder None wenn kein Office-Format.
        Fehlende XML-Dateien → leere Sections (kein Crash).
    """
    import xml.etree.ElementTree as ET

    suffix = Path(filename).suffix.lower()
    if suffix not in {".docx", ".xlsx", ".pptx", ".odt", ".ods", ".odp"}:
        return None

    core_ns = {
        "dc": "http://purl.org/dc/elements/1.1/",
        "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
        "dcterms": "http://purl.org/dc/terms/",
    }
    app_ns = {
        "ap": "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties",
    }
    custom_ns = {
        "vt": "http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes",
    }

    core: dict[str, str] = {}
    app: dict[str, str] = {}
    custom: dict[str, str] = {}

    try:
        with zipfile.ZipFile(io.BytesIO(file_data), "r") as zf:
            zip_names = set(zf.namelist())

            # --- docProps/core.xml ---
            if "docProps/core.xml" in zip_names:
                try:
                    core_xml = zf.read("docProps/core.xml")
                    root = ET.fromstring(core_xml)

                    def _core(tag_ns: str, tag_name: str) -> Optional[str]:
                        el = root.find(f"{{{core_ns[tag_ns]}}}{tag_name}")
                        return el.text.strip() if el is not None and el.text else None

                    for key, ns, tag in [
                        ("title", "dc", "title"),
                        ("subject", "dc", "subject"),
                        ("description", "dc", "description"),
                        ("author", "dc", "creator"),
                        ("last_modified_by", "cp", "lastModifiedBy"),
                        ("created", "dcterms", "created"),
                        ("modified", "dcterms", "modified"),
                        ("revision", "cp", "revision"),
                    ]:
                        val = _core(ns, tag)
                        if val is not None:
                            core[key] = val
                except Exception as e:
                    log.debug("doc_properties_core_parse_failed", filename=filename, error=str(e))

            # --- docProps/app.xml ---
            if "docProps/app.xml" in zip_names:
                try:
                    app_xml = zf.read("docProps/app.xml")
                    root = ET.fromstring(app_xml)
                    ap_ns = app_ns["ap"]

                    def _app(tag_name: str) -> Optional[str]:
                        el = root.find(f"{{{ap_ns}}}{tag_name}")
                        return el.text.strip() if el is not None and el.text else None

                    for key, tag in [
                        ("company", "Company"),
                        ("manager", "Manager"),
                        ("application", "Application"),
                        ("app_version", "AppVersion"),
                        ("pages", "Pages"),
                        ("words", "Words"),
                        ("characters", "Characters"),
                        ("total_time_minutes", "TotalTime"),
                        ("template", "Template"),
                    ]:
                        val = _app(tag)
                        if val is not None:
                            app[key] = val
                except Exception as e:
                    log.debug("doc_properties_app_parse_failed", filename=filename, error=str(e))

            # --- docProps/custom.xml ---
            if "docProps/custom.xml" in zip_names:
                try:
                    custom_xml = zf.read("docProps/custom.xml")
                    root = ET.fromstring(custom_xml)
                    vt_ns = custom_ns["vt"]

                    # Iterate over <property> elements
                    for prop_el in root:
                        name = prop_el.get("name")
                        if not name:
                            continue
                        # Try known value types
                        value: Optional[str] = None
                        for vtype in ("lpwstr", "i4", "bool", "filetime", "r8", "decimal"):
                            val_el = prop_el.find(f"{{{vt_ns}}}{vtype}")
                            if val_el is not None and val_el.text:
                                value = val_el.text.strip()
                                break
                        if value is not None:
                            custom[name] = value
                except Exception as e:
                    log.debug("doc_properties_custom_parse_failed", filename=filename, error=str(e))

    except Exception as e:
        log.debug("doc_properties_zip_open_failed", filename=filename, error=str(e))
        return None

    return {
        "core": core,
        "app": app,
        "custom": custom,
    }


# =============================================================================
# Email Metadata Extraction (T-MKIT-029)
# =============================================================================

def extract_email_metadata(file_data: bytes) -> Optional[dict[str, Any]]:
    """
    Extrahiert E-Mail-Metadaten aus rohen EML-Bytes.

    Prüft ob die Daten eine E-Mail sind (typische Header-Zeilen) und parst
    mit der Python-Stdlib email.message_from_bytes().

    Extrahiert:
    - Routing: Received-Chain, SPF/DKIM/DMARC aus Authentication-Results,
               X-Originating-IP, X-Mailer/User-Agent
    - Threading: Message-ID, In-Reply-To, References, Thread-Index
    - Calendar Events: ICS-Anhänge (text/calendar Parts), parsed via icalendar

    Args:
        file_data: Rohe Datei-Bytes (muss eine gültige E-Mail sein)

    Returns:
        Dict mit 'routing', 'thread', 'calendar_events' oder None wenn
        die Datei keine E-Mail ist.
    """
    if not file_data:
        return None

    # Schnelle Heuristik: E-Mails beginnen typischerweise mit bekannten Headern
    try:
        head = file_data[:2048].decode("utf-8", errors="replace")
    except Exception:
        return None

    first_line = head.split("\n")[0].strip()
    is_email = (
        first_line.startswith("From:")
        or first_line.startswith("Received:")
        or first_line.startswith("Return-Path:")
        or first_line.startswith("MIME-Version:")
        or first_line.startswith("Date:")
        or first_line.startswith("Message-ID:")
    )
    if not is_email:
        return None

    try:
        msg = _email_stdlib.message_from_bytes(file_data, policy=_email_policy.default)
    except Exception as e:
        log.warning("email_parse_failed", error=str(e))
        return None

    # --- Routing ---
    auth_results = str(msg.get("Authentication-Results", ""))
    spf = None
    dkim = None
    dmarc = None
    if "spf=pass" in auth_results:
        spf = "pass"
    elif "spf=fail" in auth_results:
        spf = "fail"
    elif "spf=softfail" in auth_results:
        spf = "softfail"
    elif "spf=neutral" in auth_results:
        spf = "neutral"

    if "dkim=pass" in auth_results:
        dkim = "pass"
    elif "dkim=fail" in auth_results:
        dkim = "fail"
    elif "dkim=none" in auth_results:
        dkim = "none"

    if "dmarc=pass" in auth_results:
        dmarc = "pass"
    elif "dmarc=fail" in auth_results:
        dmarc = "fail"
    elif "dmarc=none" in auth_results:
        dmarc = "none"

    routing = {
        "received_chain": msg.get_all("Received") or [],
        "spf": spf,
        "dkim": dkim,
        "dmarc": dmarc,
        "originating_ip": msg.get("X-Originating-IP"),
        "mailer": msg.get("X-Mailer") or msg.get("User-Agent"),
    }

    # --- Threading ---
    references_raw = str(msg.get("References") or "")
    thread = {
        "message_id": msg.get("Message-ID"),
        "in_reply_to": msg.get("In-Reply-To"),
        "references": references_raw.split() if references_raw.strip() else [],
        "thread_index": msg.get("Thread-Index"),
    }

    # --- Calendar Events (ICS) ---
    calendar_events: list[dict[str, Any]] = []
    if ICALENDAR_AVAILABLE:
        try:
            for part in msg.walk():
                if part.get_content_type() == "text/calendar":
                    try:
                        ics_bytes = part.get_payload(decode=True)
                        if not ics_bytes:
                            continue
                        cal = ICalendar.from_ical(ics_bytes)
                        for component in cal.walk("VEVENT"):
                            dtstart = component.get("dtstart")
                            dtend = component.get("dtend")
                            calendar_events.append({
                                "uid": str(component.get("uid", "")),
                                "summary": str(component.get("summary", "")),
                                "start": str(dtstart.dt) if dtstart else None,
                                "end": str(dtend.dt) if dtend else None,
                                "location": str(component.get("location", "")),
                                "organizer": str(component.get("organizer", "")),
                                "status": str(component.get("status", "")),
                            })
                    except Exception as ics_err:
                        log.debug("ics_parse_failed", error=str(ics_err))
        except Exception as walk_err:
            log.debug("email_walk_failed", error=str(walk_err))

    return {
        "routing": routing,
        "thread": thread,
        "calendar_events": calendar_events,
    }


# =============================================================================
# T-MKIT-030: PPTX Hidden Slides + Embedded Excel aus Charts
# =============================================================================

def extract_pptx_hidden_info(file_data: bytes) -> Optional[dict[str, Any]]:
    """
    Analysiert eine PPTX-Datei auf versteckte Slides und eingebettete Excel-Objekte.

    PPTX ist intern ein ZIP-Archiv. Versteckte Slides haben das Attribut show="0"
    im p:sld-Element. Eingebettete Excel-Dateien liegen in ppt/embeddings/.

    Args:
        file_data: Rohe PPTX-Bytes

    Returns:
        Dict mit:
            - hidden_slide_count (int): Anzahl versteckter Slides
            - hidden_slide_numbers (list[int]): Slide-Nummern (1-basiert)
            - embedded_objects (list[str]): Namen eingebetteter Dateien
        Oder None bei Fehler.
    """
    try:
        hidden_slide_numbers: list[int] = []
        embedded_objects: list[str] = []

        with zipfile.ZipFile(io.BytesIO(file_data), "r") as zf:
            all_names = zf.namelist()

            # Slide XMLs parsen: ppt/slides/slide1.xml, slide2.xml, ...
            import re as _re
            slide_files = [
                n for n in all_names
                if _re.match(r"ppt/slides/slide\d+\.xml$", n)
            ]

            for slide_name in sorted(slide_files):
                # Slide-Nummer aus Dateinamen extrahieren
                m = _re.search(r"slide(\d+)\.xml$", slide_name)
                slide_num = int(m.group(1)) if m else 0

                try:
                    xml_bytes = zf.read(slide_name)
                    xml_text = xml_bytes.decode("utf-8", errors="replace")
                    # Prüfe auf show="0" im p:sld-Element
                    # Das Attribut kann mit oder ohne Namespace-Präfix erscheinen
                    if 'show="0"' in xml_text:
                        # Sicherstellen dass es am p:sld Element hängt
                        # (nicht in einem anderen Element)
                        import xml.etree.ElementTree as _ET
                        try:
                            root = _ET.fromstring(xml_bytes)
                            # p:sld Namespace
                            pml_ns = "http://schemas.openxmlformats.org/presentationml/2006/main"
                            tag_local = root.tag.split("}")[-1] if "}" in root.tag else root.tag
                            show_val = root.get(f"{{{pml_ns}}}show") or root.get("show")
                            if tag_local == "sld" and show_val == "0":
                                hidden_slide_numbers.append(slide_num)
                        except Exception:
                            # Fallback: einfache String-Suche
                            if 'show="0"' in xml_text:
                                hidden_slide_numbers.append(slide_num)
                except Exception:
                    pass

            # Eingebettete Objekte in ppt/embeddings/
            embedding_files = [
                n for n in all_names
                if n.startswith("ppt/embeddings/") and not n.endswith("/")
            ]
            for emb_path in sorted(embedding_files):
                emb_name = emb_path.split("/")[-1]
                # Nur Excel-Objekte: *.xlsx oder oleObject*.bin
                if emb_name.endswith(".xlsx") or (
                    emb_name.startswith("oleObject") and emb_name.endswith(".bin")
                ):
                    embedded_objects.append(emb_name)

        return {
            "hidden_slide_count": len(hidden_slide_numbers),
            "hidden_slide_numbers": sorted(hidden_slide_numbers),
            "embedded_objects": embedded_objects,
        }
    except Exception as e:
        log.warning("pptx_hidden_info_failed", error=str(e))
        return None


def convert_with_markitdown(file_path: Path, show_formulas: bool = False) -> dict[str, Any]:
    """
    Konvertiert eine Datei mit MarkItDown.

    Für Excel-Dateien (.xlsx/.xls) wird convert_excel_enhanced() verwendet (FR-MKIT-007):
    - Multi-Sheet Ausgabe, Merged Cells, optionale Formel-Annotationen, Chart-Extraktion.

    Für PDF-Dateien wird zusätzlich pdfplumber für Tabellen-Extraktion genutzt:
    - Wenn pdfplumber Tabellen findet, werden diese als Markdown-Tabellen in den
      MarkItDown-Text integriert.
    - Wenn pdfplumber keine Tabellen findet, wird reiner MarkItDown-Output genutzt.

    Für PDF und DOCX wird nach der Konvertierung Code-Block-Erkennung angewendet
    (FR-MKIT-005): Indentierte Blöcke werden erkannt und in Fences gewrappt.

    Für PDF werden zusätzlich PyMuPDF-Metadaten extrahiert (FR-MKIT-009):
    - Bookmarks/TOC werden als Inhaltsverzeichnis vorangestellt.
    - Annotationen werden als Blockquotes angehängt.
    - Formularfelder werden als Key-Value-Tabelle angehängt.

    Args:
        file_path: Pfad zur Datei.
        show_formulas: Excel-Formeln annotieren (nur für .xlsx/.xls, FR-MKIT-007).
    """
    suffix = file_path.suffix.lower()

    # FR-MKIT-007: Excel → enhanced converter
    if suffix in {".xlsx", ".xls"}:
        return convert_excel_enhanced(file_path, show_formulas=show_formulas)

    try:
        log.info("markitdown_convert", file=str(file_path))
        result = md.convert(str(file_path))
        markdown_text = result.text_content

        # PDF-spezifisch: pdfplumber für Tabellen-Extraktion
        if file_path.suffix.lower() == ".pdf" and PDFPLUMBER_AVAILABLE:
            page_tables = extract_tables_with_pdfplumber(file_path)
            if page_tables:
                merged_tables = merge_cross_page_tables(page_tables)
                table_markdown = tables_to_markdown(merged_tables)
                if table_markdown:
                    log.info(
                        "pdfplumber_tables_integrated",
                        file=str(file_path),
                        table_count=len(merged_tables),
                    )
                    markdown_text = markdown_text + "\n\n## Tabellen\n\n" + table_markdown
            elif IMG2TABLE_AVAILABLE:
                # AC-012-1: img2table als Fallback wenn pdfplumber keine Tabellen findet
                img2table_tables = extract_tables_with_img2table(file_path)
                if img2table_tables:
                    table_markdown = tables_to_markdown(img2table_tables)
                    if table_markdown:
                        log.info(
                            "img2table_fallback_integrated",
                            file=str(file_path),
                            table_count=len(img2table_tables),
                            tables_source="img2table",
                        )
                        markdown_text = markdown_text + "\n\n## Tabellen\n\n" + table_markdown

        # FR-MKIT-005: Code-Block-Erkennung für PDF und DOCX
        suffix = file_path.suffix.lower()
        if suffix in {".pdf", ".docx", ".doc"}:
            markdown_text = detect_and_fence_code_blocks(markdown_text)
            log.debug("code_fencing_applied", file=str(file_path))

        # FR-MKIT-008: DOCX Extras (Kommentare, Header/Footer, Track Changes)
        if suffix == ".docx":
            try:
                extras = extract_docx_extras(file_path)
                markdown_text = append_docx_extras_to_markdown(markdown_text, extras)
                log.debug("docx_extras_appended", file=str(file_path))
            except Exception as extras_err:
                log.warning(
                    "docx_extras_failed",
                    file=str(file_path),
                    error=str(extras_err),
                )

        # FR-MKIT-009: PDF-Metadaten (Bookmarks, Annotationen, Formularfelder)
        zugferd_data: Optional[dict[str, Any]] = None
        xmp_metadata_data: Optional[dict[str, Any]] = None
        embedded_files_data: Optional[list[dict[str, Any]]] = None
        if suffix == ".pdf":
            try:
                pdf_meta = extract_pdf_metadata(file_path)
                markdown_text = prepend_pdf_toc(markdown_text, pdf_meta.get("toc", []))
                markdown_text = append_pdf_annotations(markdown_text, pdf_meta.get("annotations", []))
                markdown_text = append_pdf_form_fields(markdown_text, pdf_meta.get("form_fields", []))
                log.debug("pdf_metadata_appended", file=str(file_path))
            except Exception as pdf_meta_err:
                log.warning(
                    "pdf_metadata_failed",
                    file=str(file_path),
                    error=str(pdf_meta_err),
                )

            # Datei-Bytes einmalig lesen für T-MKIT-024 + T-MKIT-025
            pdf_file_bytes: Optional[bytes] = None
            try:
                pdf_file_bytes = file_path.read_bytes()
            except Exception as read_err:
                log.warning("pdf_bytes_read_failed", file=str(file_path), error=str(read_err))

            # T-MKIT-024: ZUGFeRD/Factur-X E-Rechnung aus eingebettetem XML extrahieren
            if pdf_file_bytes is not None:
                try:
                    zugferd_xml = detect_zugferd(pdf_file_bytes)
                    if zugferd_xml is not None:
                        zugferd_data = parse_zugferd_xml(zugferd_xml)
                        log.info("zugferd_detected", file=str(file_path))
                except Exception as zugferd_err:
                    log.warning("zugferd_extract_failed", file=str(file_path), error=str(zugferd_err))

            # T-MKIT-025: XMP Metadata + Embedded Files extrahieren
            if pdf_file_bytes is not None:
                try:
                    xmp_metadata_data = extract_xmp_metadata(pdf_file_bytes)
                    if xmp_metadata_data:
                        log.debug("xmp_metadata_extracted", file=str(file_path))
                except Exception as xmp_err:
                    log.warning("xmp_metadata_failed", file=str(file_path), error=str(xmp_err))

                try:
                    embedded_list = list_embedded_files(pdf_file_bytes)
                    if embedded_list:
                        embedded_files_data = embedded_list
                        log.debug("embedded_files_listed", file=str(file_path), count=len(embedded_list))
                except Exception as emb_err:
                    log.warning("embedded_files_failed", file=str(file_path), error=str(emb_err))

        # T-MKIT-028: Office Document Properties (DOCX, XLSX, PPTX)
        document_properties_data: Optional[dict[str, Any]] = None
        if suffix in {".docx", ".xlsx", ".pptx", ".odt", ".ods", ".odp"}:
            try:
                file_bytes_for_props = file_path.read_bytes()
                document_properties_data = extract_document_properties(
                    file_bytes_for_props, file_path.name
                )
                if document_properties_data:
                    log.debug("document_properties_extracted", file=str(file_path))
            except Exception as props_err:
                log.warning(
                    "document_properties_failed",
                    file=str(file_path),
                    error=str(props_err),
                )

        # T-MKIT-029: Email Metadata (Routing, Threading, Calendar Events)
        email_metadata_data: Optional[dict[str, Any]] = None
        if suffix == ".eml":
            try:
                eml_bytes = file_path.read_bytes()
                email_metadata_data = extract_email_metadata(eml_bytes)
                if email_metadata_data:
                    log.debug("email_metadata_extracted", file=str(file_path))
            except Exception as eml_err:
                log.warning(
                    "email_metadata_failed",
                    file=str(file_path),
                    error=str(eml_err),
                )

        # T-MKIT-030: PPTX Hidden Slides + Embedded Excel aus Charts
        pptx_hidden_info: Optional[dict[str, Any]] = None
        if suffix in {".pptx", ".ppt"}:
            try:
                pptx_bytes = file_path.read_bytes()
                pptx_hidden_info = extract_pptx_hidden_info(pptx_bytes)
                if pptx_hidden_info and pptx_hidden_info.get("hidden_slide_count", 0) > 0:
                    count = pptx_hidden_info["hidden_slide_count"]
                    numbers = pptx_hidden_info["hidden_slide_numbers"]
                    slides_str = ", ".join(str(n) for n in numbers)
                    markdown_text += (
                        f"\n\n---\n*Note: This presentation contains {count} "
                        f"hidden slide(s) (slides: {slides_str}).*"
                    )
                    log.info(
                        "pptx_hidden_slides_detected",
                        file=str(file_path),
                        count=count,
                        slides=numbers,
                    )
            except Exception as pptx_err:
                log.warning(
                    "pptx_hidden_info_failed",
                    file=str(file_path),
                    error=str(pptx_err),
                )

        return {
            "success": True,
            "markdown": markdown_text,
            "title": getattr(result, "title", None),
            "zugferd": zugferd_data,
            "xmp_metadata": xmp_metadata_data,
            "embedded_files": embedded_files_data,
            "document_properties": document_properties_data,
            "email_metadata": email_metadata_data,
            "pptx_hidden_info": pptx_hidden_info,
        }
    except Exception as e:
        log.error("markitdown_error", file=str(file_path), error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": f"MarkItDown Fehler: {str(e)}"
        }


async def convert_url(url: str) -> dict[str, Any]:
    """Konvertiert eine URL zu Markdown (non-blocking via asyncio.to_thread)."""
    try:
        log.info("url_convert", url=url)
        result = await asyncio.to_thread(md.convert_url, url)
        return {
            "success": True,
            "markdown": result.text_content,
            "title": getattr(result, "title", None),
        }
    except Exception as e:
        log.error("url_convert_error", url=url, error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": f"URL-Konvertierung fehlgeschlagen: {str(e)}"
        }


# =============================================================================
# Embedded Image Extraction + Description
# =============================================================================

def extract_images_from_docx(file_path: Path) -> list[dict]:
    """
    Extrahiert eingebettete Bilder aus einer DOCX-Datei.

    DOCX ist intern ein ZIP-Archiv. Bilder liegen in word/media/*.

    Args:
        file_path: Pfad zur DOCX-Datei

    Returns:
        Liste von Dicts mit 'name', 'data' (bytes), 'position_hint' (Index)
    """
    images: list[dict] = []
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            media_files = [
                name for name in zf.namelist()
                if name.startswith("word/media/") and not name.endswith("/")
            ]
            for idx, media_name in enumerate(sorted(media_files)):
                try:
                    data = zf.read(media_name)
                    img = Image.open(io.BytesIO(data))
                    width, height = img.size
                    if width < MIN_IMAGE_SIZE_PX or height < MIN_IMAGE_SIZE_PX:
                        log.debug(
                            "skip_small_image_docx",
                            name=media_name,
                            width=width,
                            height=height,
                        )
                        continue
                    images.append({
                        "name": Path(media_name).name,
                        "data": data,
                        "position_hint": idx,
                    })
                except Exception as e:
                    log.warning("docx_image_read_error", name=media_name, error=str(e))
    except Exception as e:
        log.error("docx_open_error", file=str(file_path), error=str(e))
    log.info("docx_images_extracted", file=str(file_path), count=len(images))
    return images


def extract_images_from_pptx(file_path: Path) -> list[dict]:
    """
    Extrahiert eingebettete Bilder aus einer PPTX-Datei.

    PPTX ist intern ein ZIP-Archiv. Bilder liegen in ppt/media/*.

    Args:
        file_path: Pfad zur PPTX-Datei

    Returns:
        Liste von Dicts mit 'name', 'data' (bytes), 'slide_number' (Index)
    """
    images: list[dict] = []
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            media_files = [
                name for name in zf.namelist()
                if name.startswith("ppt/media/") and not name.endswith("/")
            ]
            for idx, media_name in enumerate(sorted(media_files)):
                try:
                    data = zf.read(media_name)
                    img = Image.open(io.BytesIO(data))
                    width, height = img.size
                    if width < MIN_IMAGE_SIZE_PX or height < MIN_IMAGE_SIZE_PX:
                        log.debug(
                            "skip_small_image_pptx",
                            name=media_name,
                            width=width,
                            height=height,
                        )
                        continue
                    images.append({
                        "name": Path(media_name).name,
                        "data": data,
                        "slide_number": idx + 1,
                    })
                except Exception as e:
                    log.warning("pptx_image_read_error", name=media_name, error=str(e))
    except Exception as e:
        log.error("pptx_open_error", file=str(file_path), error=str(e))
    log.info("pptx_images_extracted", file=str(file_path), count=len(images))
    return images


async def classify_image_type(image_data: bytes, mimetype: str) -> str:
    """
    Klassifiziert ein Bild via Mistral Vision in eine von fünf Kategorien.

    Schickt das Bild an die Vision-API mit einem präzisen Klassifizierungs-Prompt
    und gibt genau eine Kategorie zurück.

    Args:
        image_data: Rohe Bildbytes.
        mimetype: MIME-Typ des Bildes (z.B. 'image/png').

    Returns:
        Einer der Werte: 'photo', 'chart', 'diagram', 'text_scan', 'decorative'.
        Fallback: 'photo' bei Fehler oder unbekannter Antwort.
    """
    prompt = (
        "Classify this image into EXACTLY one category. Reply with ONE word only:\n\n"
        "photo      = photograph of real-world objects, people, places\n"
        "chart      = bar chart, line chart, pie chart, data visualization with axes/values\n"
        "diagram    = flowchart, org chart, mind map, network diagram, UML, architecture diagram\n"
        "text_scan  = image of a document, form, invoice, letter, or any image where text is the primary content\n"
        "decorative = logo, icon, background image, decorative graphic with no information value\n\n"
        "Reply with exactly one of: photo, chart, diagram, text_scan, decorative"
    )
    valid_types = {"photo", "chart", "diagram", "text_scan", "decorative"}

    log.info("classify_image_type_start", size=len(image_data), mimetype=mimetype)

    result = await analyze_with_mistral_vision(image_data, mimetype, prompt, language="en")
    if not result.get("success"):
        log.warning(
            "classify_image_type_failed",
            error=result.get("error"),
            fallback="photo",
        )
        return "photo"

    raw: str = result.get("markdown", "").strip().lower()
    # Extrahiere erstes Wort aus der Antwort (falls Modell mehr zurückgibt)
    first_word = raw.split()[0] if raw.split() else ""
    image_type = first_word if first_word in valid_types else "photo"

    log.info("classify_image_type_result", raw_response=raw, image_type=image_type)
    return image_type


async def convert_diagram_to_mermaid(image_data: bytes, mimetype: str) -> str:
    """
    Konvertiert ein Diagramm/Flowchart-Bild in Mermaid-Syntax via Vision-API.

    Nutzt einen spezialisierten Prompt um Flowcharts und andere Diagramme
    als valide Mermaid-Syntax zu extrahieren.

    Args:
        image_data: Rohe Bildbytes des Diagramms.
        mimetype: MIME-Typ des Bildes (z.B. 'image/png').

    Returns:
        Mermaid-Code-Block (```mermaid ... ```) oder Fallback-Text bei Fehler.
    """
    prompt = (
        "Convert this diagram image into valid Mermaid syntax.\n\n"
        "Choose the appropriate diagram type:\n"
        "- Flowchart/decision tree → graph TD\n"
        "- Sequence diagram → sequenceDiagram\n"
        "- Class diagram → classDiagram\n"
        "- Org chart → graph TD with descriptive node labels\n\n"
        "Rules:\n"
        "- Output ONLY the Mermaid code inside ```mermaid ... ``` fences\n"
        "- Use exact labels from the image — do not invent labels\n"
        "- If the image cannot be represented as Mermaid: output ```mermaid\\ngraph TD\\n    A[Nicht darstellbar]\\n```\n"
        "- No explanations, no text outside the code block"
    )

    log.info("convert_diagram_to_mermaid_start", size=len(image_data), mimetype=mimetype)

    result = await analyze_with_mistral_vision(image_data, mimetype, prompt, language="en")
    if not result.get("success"):
        log.warning(
            "convert_diagram_to_mermaid_failed",
            error=result.get("error"),
        )
        return "[Diagramm-Konvertierung nicht verfügbar]"

    mermaid_output: str = strip_llm_artifacts(result.get("markdown", "").strip())
    log.info("convert_diagram_to_mermaid_success", output_length=len(mermaid_output))
    return mermaid_output


async def extract_chart_data(image_data: bytes, mimetype: str) -> str:
    """
    Extrahiert Daten aus einem Chart/Diagramm-Bild als Markdown-Tabelle.

    Nutzt einen spezialisierten Prompt um Achsenbeschriftungen, Datenpunkte
    und Legenden aus Balken-, Linien- und Kreisdiagrammen zu extrahieren.

    Args:
        image_data: Rohe Bildbytes des Charts.
        mimetype: MIME-Typ des Bildes (z.B. 'image/png').

    Returns:
        Markdown-Tabelle mit den extrahierten Daten oder Fallback-Text bei Fehler.
    """
    prompt = (
        "Extract all data from this chart as a Markdown table.\n\n"
        "Instructions:\n"
        "- For bar/line charts: columns = X-axis label + one column per data series; rows = data points\n"
        "- For pie/donut charts: columns = Category | Value | Percentage\n"
        "- Use exact axis labels and legend entries as column headers\n"
        "- If exact values are not readable, use estimates with ~ prefix (e.g. ~42)\n"
        "- If no chart data found: output only [No chart data found]\n\n"
        "Output ONLY the Markdown table. No explanations, no introductions."
    )

    log.info("extract_chart_data_start", size=len(image_data), mimetype=mimetype)

    result = await analyze_with_mistral_vision(image_data, mimetype, prompt, language="en")
    if not result.get("success"):
        log.warning(
            "extract_chart_data_failed",
            error=result.get("error"),
        )
        return "[Daten-Extraktion nicht verfügbar]"

    table_output: str = strip_llm_artifacts(result.get("markdown", "").strip())
    log.info("extract_chart_data_success", output_length=len(table_output))
    return table_output


async def describe_embedded_images(
    images: list[dict],
    language: str = "de",
) -> list[dict]:
    """
    Beschreibt eine Liste extrahierter Bilder via Mistral Pixtral Vision.

    Bilder werden zuerst klassifiziert (AC-004-1):
    - 'diagram' → Mermaid-Syntax-Konvertierung (AC-004-2)
    - 'chart'   → Datentabellen-Extraktion (AC-004-3)
    - 'photo'   → generische Vision-Beschreibung
    - 'text_scan' → generische Vision-Beschreibung (Text-Extraktion)
    - 'decorative' → wird übersprungen

    Args:
        images: Liste von Dicts mit 'name' und 'data' (bytes) aus extract_images_from_*()
        language: Antwortsprache ('de' oder 'en')

    Returns:
        Liste von Dicts mit 'name', 'description', 'tokens', 'image_type'
    """
    results: list[dict] = []
    for image in images:
        name = image["name"]
        data = image["data"]
        mimetype = detect_mimetype_from_bytes(data) or "image/png"
        log.info("describing_embedded_image", name=name, size=len(data))

        # AC-004-1: Bild klassifizieren
        image_type = await classify_image_type(data, mimetype)
        log.info("image_classified", name=name, image_type=image_type)

        if image_type == "decorative":
            # Dekorative Bilder überspringen
            log.info("skip_decorative_image", name=name)
            continue

        if image_type == "diagram":
            # AC-004-2: Flowcharts/Organigramme → Mermaid
            description = await convert_diagram_to_mermaid(data, mimetype)
            results.append({
                "name": name,
                "description": description,
                "tokens": 0,
                "image_type": image_type,
            })
            continue

        if image_type == "chart":
            # AC-004-3: Balken-/Linien-/Kreisdiagramme → Datentabelle
            description = await extract_chart_data(data, mimetype)
            results.append({
                "name": name,
                "description": description,
                "tokens": 0,
                "image_type": image_type,
            })
            continue

        # 'photo' und 'text_scan': differenzierte Vision-Beschreibung
        if image_type == "text_scan":
            generic_prompt = (
                "Extrahiere den gesamten sichtbaren Text aus diesem Bild. "
                "Gib ihn strukturiert als Markdown wieder. Übersetze nicht."
                if language == "de"
                else "Extract all visible text from this image. "
                     "Return it as structured Markdown. Do not translate."
            )
        else:  # photo oder unbekannt
            generic_prompt = (
                "Beschreibe dieses Bild präzise: was ist zu sehen, relevante Beschriftungen, "
                "erkennbare Objekte und der Gesamtkontext. Format: kurzer Absatz."
                if language == "de"
                else "Describe this image precisely: what is shown, relevant labels, "
                     "recognizable objects and overall context. Format: short paragraph."
            )

        result = await analyze_with_mistral_vision(data, mimetype, generic_prompt, language)
        if result["success"]:
            results.append({
                "name": name,
                "description": result["markdown"],
                "tokens": result.get("tokens_total", 0),
                "image_type": image_type,
            })
        else:
            log.warning(
                "embedded_image_description_failed",
                name=name,
                error=result.get("error"),
            )
            results.append({
                "name": name,
                "description": f"[Bildbeschreibung nicht verfügbar: {result.get('error', 'Unbekannter Fehler')}]",
                "tokens": 0,
                "image_type": image_type,
            })
    return results


def insert_image_descriptions(markdown: str, descriptions: list[dict]) -> str:
    """
    Ersetzt Bild-Platzhalter im Markdown durch Pixtral-Beschreibungen.

    Markitdown erzeugt Platzhalter wie ![image](image1.png) für eingebettete Bilder.
    Diese Funktion ersetzt sie durch einen beschreibenden Blockquote.

    Args:
        markdown: Konvertierter Markdown-Text mit Bild-Platzhaltern
        descriptions: Liste von Dicts mit 'name' und 'description'

    Returns:
        Markdown mit eingefügten Bildbeschreibungen
    """
    # Baue Lookup: Dateiname → Beschreibung
    desc_map: dict[str, str] = {d["name"]: d["description"] for d in descriptions}

    def replace_placeholder(match: re.Match) -> str:
        alt_text = match.group(1)
        img_ref = match.group(2)
        # img_ref kann ein Dateiname oder Pfad sein
        img_name = Path(img_ref).name
        if img_name in desc_map:
            description = desc_map[img_name]
            return f"> **[Bild: {img_name}]** {description}"
        # Kein passender Eintrag → Original behalten
        return match.group(0)

    # Muster: ![alt](ref) — erfasst Bild-Platzhalter
    pattern = r"!\[([^\]]*)\]\(([^)]+)\)"
    return re.sub(pattern, replace_placeholder, markdown)


# =============================================================================
# Dokumenten-Klassifizierung via LLM
# =============================================================================

def get_classify_categories_from_db() -> list[str]:
    """Lädt alle Template-IDs + display_names aus der DB für den Classify-Prompt.

    Returns:
        Liste von Strings im Format "id: display_name", z.B. ["invoice: Rechnung", ...].
        Bei Fehler oder leerer DB wird auf DEFAULT_CLASSIFY_CATEGORIES zurückgefallen
        (im Format "id: id" für Rückwärtskompatibilität).
    """
    import time
    now = time.time()
    if _classify_categories_cache["categories"] is not None and (now - _classify_categories_cache["timestamp"]) < _CLASSIFY_CACHE_TTL:
        return _classify_categories_cache["categories"]

    try:
        conn = get_db_connection()
        rows = conn.execute(
            "SELECT id, display_name FROM template WHERE enabled = 1 ORDER BY priority DESC, id"
        ).fetchall()
        conn.close()
        if rows:
            categories = [f"{r['id']}: {r['display_name']}" for r in rows]
            _classify_categories_cache["categories"] = categories
            _classify_categories_cache["timestamp"] = now
            return categories
    except Exception:
        pass

    # Fallback auf hardcodierte Liste (Format: "id: id" für einheitliche Verarbeitung)
    return [f"{c.strip()}: {c.strip()}" for c in DEFAULT_CLASSIFY_CATEGORIES]


async def classify_document(
    markdown: str,
    categories: list[str] | None = None,
    language: str = "de",
) -> dict[str, Any]:
    """
    Klassifiziert ein Dokument anhand seines Markdown-Inhalts via Mistral API.

    Args:
        markdown: Konvertierter Markdown-Text des Dokuments.
        categories: Erlaubte Dokumenttypen. Wenn None, werden DEFAULT_CLASSIFY_CATEGORIES
                    verwendet.
        language: Sprache des Prompts ('de' oder 'en').

    Returns:
        Dict mit 'document_type' (str) und 'document_type_confidence' (float 0.0–1.0).
        Bei Fehlern wird {"document_type": "other", "document_type_confidence": 0.0}
        zurückgegeben (graceful degradation).
    """
    if not MISTRAL_API_KEY:
        log.warning("classify_document_no_api_key")
        return {"document_type": "other", "document_type_confidence": 0.0}

    # Template-Registry Kategorien laden (statt hardcodierte Liste)
    if categories is not None:
        # Explizite Kategorien vom Aufrufer → direkt als "id: id" nutzen (Rückwärtskompatibilität)
        db_categories = [f"{c}: {c}" for c in categories]
        # Erlaubte IDs für Validierung
        allowed_ids = list(categories)
    else:
        # Aus Template-Registry laden
        db_categories = get_classify_categories_from_db()
        # Erlaubte IDs aus "id: display_name" extrahieren
        allowed_ids = [c.split(":")[0].strip() for c in db_categories]

    # Kategorien-String für den Prompt: "id: Beschreibung" zeilenweise
    categories_lines = "\n".join(db_categories)

    # Markdown auf maximal CLASSIFY_MAX_CHARS Zeichen kürzen, um Token-Kosten zu begrenzen
    truncated_markdown = markdown[:CLASSIFY_MAX_CHARS] if len(markdown) > CLASSIFY_MAX_CHARS else markdown

    if language == "de":
        system_prompt = "Du bist ein Experte für Dokumentenklassifizierung. Antworte ausschließlich mit validem JSON."
        user_prompt = (
            f"Klassifiziere dieses Dokument. Wähle den spezifischsten Typ.\n"
            f"Antworte AUSSCHLIESSLICH mit JSON: {{\"type\": \"template_id\", \"confidence\": 0.95}}\n\n"
            f"Verfügbare Typen (ID: Beschreibung):\n{categories_lines}\n\n"
            f"\"confidence\": Zahl zwischen 0.0 (sehr unsicher) und 1.0 (sehr sicher)\n"
            f"Verwende GENAU eine der Typ-IDs. Bevorzuge den spezifischsten Typ.\n\n"
            f"Dokument:\n{truncated_markdown}"
        )
    else:
        system_prompt = "You are an expert document classifier. Respond exclusively with valid JSON."
        user_prompt = (
            f"Classify this document. Choose the most specific type.\n"
            f"Respond EXCLUSIVELY with JSON: {{\"type\": \"template_id\", \"confidence\": 0.95}}\n\n"
            f"Available types (ID: description):\n{categories_lines}\n\n"
            f"\"confidence\": number between 0.0 (very uncertain) and 1.0 (very certain)\n"
            f"Use EXACTLY one of the type IDs. Prefer the most specific type.\n\n"
            f"Document:\n{truncated_markdown}"
        )

    payload = {
        "model": MISTRAL_TEXT_MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "max_tokens": CLASSIFY_MAX_TOKENS,
        "temperature": 0.0,
    }

    try:
        log.info("classify_document_start", categories=allowed_ids, text_length=len(truncated_markdown))
        result = await call_mistral_vision_api(payload)
        content = result["choices"][0]["message"]["content"].strip()

        # JSON aus der Antwort extrahieren (Modell könnte Markdown-Code-Blöcke liefern)
        json_match = re.search(r"\{[^}]+\}", content, re.DOTALL)
        if not json_match:
            log.warning("classify_document_no_json", raw_content=content)
            return {"document_type": "other", "document_type_confidence": 0.0}

        parsed = json.loads(json_match.group(0))
        doc_type = str(parsed.get("type", "other")).strip()
        confidence = float(parsed.get("confidence", 0.0))

        # Nur erlaubte Typen (IDs) durchlassen
        if doc_type not in allowed_ids:
            log.warning("classify_document_unknown_type", doc_type=doc_type, allowed=allowed_ids)
            doc_type = "other"

        # Konfidenz auf gültigen Bereich begrenzen
        confidence = max(0.0, min(1.0, confidence))

        log.info("classify_document_done", document_type=doc_type, confidence=confidence)
        return {"document_type": doc_type, "document_type_confidence": confidence}

    except json.JSONDecodeError as exc:
        log.warning("classify_document_json_error", error=str(exc))
        return {"document_type": "other", "document_type_confidence": 0.0}
    except Exception as exc:
        log.warning("classify_document_api_error", error=str(exc))
        return {"document_type": "other", "document_type_confidence": 0.0}


# =============================================================================
# OCR-Nachkorrektur via LLM
# =============================================================================

async def correct_ocr_text(text: str, language: str = "de") -> dict[str, Any]:
    """
    Korrigiert typische OCR-Fehler in einem Text via Mistral API.

    Sendet den OCR-Text mit einem speziellen Korrektur-Prompt an die Mistral API.
    Typische OCR-Artefakte wie Zeichenverwechslungen (rn→m, 0→O, l→1, fi→fi),
    falsche Worttrennungen und fehlende Leerzeichen werden behoben.
    Inhaltliche Fakten, Zahlen, Namen und Daten werden NICHT verändert.

    Args:
        text: OCR-extrahierter Text (Markdown).
        language: Sprache des Textes ('de' oder 'en').

    Returns:
        Dict mit:
        - corrected_text (str): Korrigierter Text
        - corrections_count (int): Anzahl der Korrekturen
        - tokens (int): Verbrauchte Tokens
        - success (bool): War die Korrektur erfolgreich?
        - error (str, optional): Fehlermeldung bei Misserfolg
    """
    if not MISTRAL_API_KEY:
        log.warning("correct_ocr_text_no_api_key")
        return {
            "success": False,
            "error": "MISTRAL_API_KEY nicht konfiguriert",
            "corrected_text": text,
            "corrections_count": 0,
            "tokens": 0,
        }

    if language == "de":
        system_prompt = (
            "Du bist ein Experte für OCR-Fehlerkorrektur. "
            "Korrigiere ausschließlich offensichtliche OCR-Artefakte, verändere KEINE inhaltlichen Fakten."
        )
        user_prompt = (
            "Korrigiere OCR-Fehler in diesem Markdown-Text.\n\n"
            "Erlaubte Korrekturen (NUR diese):\n"
            "- Zeichen-Verwechslungen: rn→m, 0→O, l→1, fi-Ligaturen, Ü→U etc.\n"
            "- Zusammengeklebte Wörter: 'dasHaus' → 'das Haus'\n"
            "- Falsche Worttrennungen: 'Doku-\\nment' → 'Dokument'\n\n"
            "VERBOTEN:\n"
            "- Inhaltliche Korrekturen (Fakten, Zahlen, Namen)\n"
            "- Änderungen an Markdown-Formatierung (#, *, |, ```)\n"
            "- Umformulierungen\n\n"
            "Antworte mit dem korrigierten Text, dann genau eine abschließende Zeile:\n"
            "<<<CORRECTIONS:N>>>\n"
            "(wobei N die Anzahl der Korrekturen ist)\n\n"
            f"Text:\n{text}"
        )
    else:
        system_prompt = (
            "You are an expert in OCR error correction. "
            "Correct only obvious OCR artifacts, do NOT change any factual content."
        )
        user_prompt = (
            "Correct OCR errors in this Markdown text.\n\n"
            "Allowed corrections (ONLY these):\n"
            "- Character confusions: rn→m, 0→O, l→1, fi-ligatures, etc.\n"
            "- Glued-together words: 'theHouse' → 'the House'\n"
            "- Wrong hyphenation: 'docu-\\nment' → 'document'\n\n"
            "FORBIDDEN:\n"
            "- Content corrections (facts, numbers, names)\n"
            "- Changes to Markdown formatting (#, *, |, ```)\n"
            "- Rephrasing\n\n"
            "Reply with the corrected text, then exactly one closing line:\n"
            "<<<CORRECTIONS:N>>>\n"
            "(where N is the number of corrections)\n\n"
            f"Text:\n{text}"
        )

    payload = {
        "model": MISTRAL_TEXT_MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "max_tokens": OCR_CORRECT_MAX_TOKENS,
        "temperature": 0.0,
    }

    try:
        log.info("correct_ocr_text_start", text_length=len(text), language=language)
        result = await call_mistral_vision_api(payload)
        content = result["choices"][0]["message"]["content"]
        usage = result.get("usage", {})
        tokens_total = usage.get("total_tokens", 0)

        # Korrektur-Anzahl aus dem speziellen Marker parsen
        corrections_count = 0
        corrected_text = content

        marker_match = re.search(r"<<<CORRECTIONS:(\d+)>>>\s*$", content.strip(), re.MULTILINE)
        if not marker_match:
            # Fallback: altes Marker-Format unterstützen
            marker_match = re.search(r"---CORRECTIONS:\s*(\d+)\s*$", content.strip(), re.MULTILINE)
        if marker_match:
            corrections_count = int(marker_match.group(1))
            # Marker aus dem Text entfernen
            corrected_text = strip_llm_artifacts(content[: marker_match.start()].rstrip())
        else:
            log.warning("correct_ocr_text_no_corrections_marker", content_tail=content[-100:])
            corrected_text = strip_llm_artifacts(content)

        log.info(
            "correct_ocr_text_done",
            corrections_count=corrections_count,
            tokens=tokens_total,
        )
        return {
            "success": True,
            "corrected_text": corrected_text,
            "corrections_count": corrections_count,
            "tokens": tokens_total,
        }

    except Exception as exc:
        log.warning("correct_ocr_text_api_error", error=str(exc))
        return {
            "success": False,
            "error": str(exc),
            "corrected_text": text,
            "corrections_count": 0,
            "tokens": 0,
        }


# =============================================================================
# Schema-basierte strukturierte Extraktion
# =============================================================================

# EXTRACTION_TEMPLATES entfernt — Templates leben ausschließlich in der SQLite-DB.
# Seed-Daten: mcp/seed.sql (wird beim ersten Start geladen).


# ---------------------------------------------------------------------------
# _meta Block: Steuerrelevanz + Datentyp-Konventionen (T-MKIT-037)
# ---------------------------------------------------------------------------

_META_SCHEMA = {
    "absender": {
        "name": "string | null — Persönlicher Name des Absenders (z.B. 'Thomas Weber'). null wenn nur Firma erkennbar.",
        "firma": "string | null — Firmenname / Organisation (z.B. 'Telekom Deutschland GmbH', 'REWE Sascha Sieger oHG'). null wenn Privatperson.",
        "slug": "string — geläufigster Kurzname, Kleinbuchstaben, Bindestriche statt Leerzeichen",
        "adresse": {
            "strasse": "string | null",
            "plz": "string | null",
            "ort": "string | null",
        },
    },
    "empfaenger": {
        "name": "string | null — Name des Empfängers (z.B. 'Max Mustermann', 'Max und Maria Mustermann'). null wenn nicht erkennbar (z.B. Kassenbon).",
        "slug": "string — geläufigster Kurzname, Kleinbuchstaben, Bindestriche statt Leerzeichen",
        "adresse": {
            "strasse": "string | null",
            "plz": "string | null",
            "ort": "string | null",
        },
    },
    "steuerrelevant": "boolean — true wenn das Dokument steuerlich relevant ist",
    "steuerrelevanz_hinweis": "string | null — wörtliches Zitat aus dem Dokument falls vorhanden",
    "steuer_kategorie": "string | null — werbungskosten | sonderausgaben | aussergewoehnliche_belastungen | haushaltsnahe_dienstleistungen | handwerkerleistungen | vorsorgeaufwendungen | kapitalertraege | vermietung | kirchensteuer | spenden | kinderbetreuung | null",
    "steuerjahr": "string | null — YYYY",
    "mwst_ausgewiesen": "boolean",
    "mwst_betrag": "string | null — Decimal mit 2 Stellen",
    "mwst_satz": "string | null — z.B. '19' oder '7'",
    "aktenzeichen": "string | null — Aktenzeichen, Geschäftszeichen, Vorgangsnummer",
    "dokumenten_id": "string | null — eindeutige ID (Rechnungsnr, Policennr, Bescheidnr)"
}

_STEUER_SIGNALWOERTER = [
    "Finanzamt", "Steuererklärung", "steuerlich absetzbar", "steuerrelevant",
    "§10 EStG", "§10b EStG", "§35a EStG",
    "Werbungskosten", "Sonderausgaben", "außergewöhnliche Belastungen",
    "Vorsorgeaufwendungen", "Altersvorsorge", "Riester", "Rürup",
    "Spendenquittung", "Zuwendungsbestätigung",
    "Lohnanteil", "Arbeitskosten",
    "Arbeitsmittel", "Fortbildung", "Fachliteratur",
    "Bitte aufbewahren", "zur Vorlage"
]

_DATENTYP_KONVENTIONEN = """
Datentyp-Konventionen für die Extraktion:
- Datum: ISO 8601 YYYY-MM-DD (z.B. "2025-03-26")
- Betrag: IMMER Punkt als Dezimaltrenner, 2 Nachkommastellen (z.B. "49.99" NICHT "49,99")
- Währung: ISO 4217 (z.B. "EUR")
- IBAN: Ohne Leerzeichen (z.B. "DE89370400440532013000")
- Telefon: E.164 (z.B. "+4921611234567")
- PLZ: String mit führenden Nullen (z.B. "01234")
- Fehlende Werte: null (NICHT leerer String "")
- Boolean: true/false
- Tabellen/Positionen: JSON Array of Objects
- Positionsbezeichnungen: Wenn leer, aus Dokumentkontext ableiten (z.B. Seitenüberschrift, vorherige Position)
"""


# =============================================================================
# Template Registry — SQLite (T-MKIT-035)
# =============================================================================

def init_templates_db() -> None:
    """Erstellt die templates.db beim ersten Start und migriert bestehende Templates."""
    TEMPLATES_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(TEMPLATES_DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("""
        CREATE TABLE IF NOT EXISTS template (
            id TEXT PRIMARY KEY,
            category TEXT NOT NULL DEFAULT 'other',
            display_name TEXT NOT NULL,
            description TEXT,
            schema TEXT NOT NULL,
            field_descriptions TEXT,
            classify_keywords TEXT,
            typical_senders TEXT,
            steuer_relevanz TEXT,
            priority INTEGER DEFAULT 0,
            enabled BOOLEAN DEFAULT 1,
            version INTEGER DEFAULT 1,
            source TEXT DEFAULT 'manual',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_templates_category ON template(category)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_templates_enabled ON template(enabled)")

    # Seed aus seed.sql wenn DB leer
    cursor = conn.execute("SELECT COUNT(*) FROM template")
    if cursor.fetchone()[0] == 0:
        seed_path = Path(__file__).parent / "seed.sql"
        if seed_path.exists():
            seed_sql = seed_path.read_text(encoding="utf-8")
            conn.executescript(seed_sql)
            log.info("templates_seeded", source=str(seed_path))
    conn.commit()
    conn.close()


def get_db_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(str(TEMPLATES_DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def get_template_by_id(template_id: str) -> Optional[dict]:
    """Lädt ein Template aus der DB. Gibt None zurück wenn nicht gefunden oder disabled."""
    conn = get_db_connection()
    row = conn.execute("SELECT * FROM template WHERE id = ? AND enabled = 1", (template_id,)).fetchone()
    conn.close()
    if not row:
        return None
    result = dict(row)
    result["schema"] = json.loads(result["schema"])
    if result.get("field_descriptions"):
        result["field_descriptions"] = json.loads(result["field_descriptions"])
    return result


def get_all_template_ids() -> list[str]:
    """Gibt alle aktiven Template-IDs aus der DB zurück."""
    conn = get_db_connection()
    rows = conn.execute("SELECT id FROM template WHERE enabled = 1 ORDER BY category, display_name").fetchall()
    conn.close()
    return [r["id"] for r in rows]


def search_templates(query: str) -> list[dict]:
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT id, category, display_name, description, enabled FROM template "
        "WHERE id LIKE ? OR display_name LIKE ? OR description LIKE ? OR classify_keywords LIKE ?",
        (f"%{query}%", f"%{query}%", f"%{query}%", f"%{query}%")
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


async def _apply_auto_extract(
    response: "ConvertResponse",
    meta: dict,
    markdown: str,
    language: str,
    min_confidence: float,
    hints: list[str],
) -> "ConvertResponse":
    """Führt Auto-Extract nach der Konvertierung durch (T-MKIT-036).

    Klassifiziert (falls nicht bereits geschehen), sucht ein passendes Template
    und extrahiert strukturierte Daten — alles in einem Schritt.

    Mutiert meta und response in-place, gibt response zurück.
    """
    # Schritt 1: Klassifizierung wenn noch nicht vorhanden
    if not meta.get("document_type"):
        classify_result = await classify_document(markdown, None, language)
        meta.update(classify_result)
        response.meta = MetaData(**{k: v for k, v in meta.items()})

    # Schritt 2: Template finden nur wenn Konfidenz ausreichend
    confidence = meta.get("document_type_confidence", 0.0) or 0.0
    doc_type = meta.get("document_type", "other") or "other"
    meta["auto_extract"] = True

    if confidence >= min_confidence and doc_type and doc_type != "other":
        tmpl = find_matching_template(doc_type, markdown)
        if tmpl:
            schema = tmpl["schema"] if isinstance(tmpl["schema"], dict) else json.loads(tmpl["schema"])
            extraction = await extract_structured_data(
                markdown, schema, language,
                field_descriptions=tmpl.get("field_descriptions"),
                notes=tmpl.get("notes"),
            )
            if extraction["success"]:
                response.extracted = extraction["extracted"]
                meta["template_used"] = tmpl["id"]
                meta["template_version"] = tmpl.get("version", 1)
                response.meta = MetaData(**{k: v for k, v in meta.items()})
            else:
                log.warning("auto_extract_failed", template=tmpl["id"], error=extraction.get("error"))
                meta["template_used"] = None
                response.meta = MetaData(**{k: v for k, v in meta.items()})
        else:
            meta["template_used"] = None
            response.meta = MetaData(**{k: v for k, v in meta.items()})
            hints.append(
                f"No template registered for document_type '{doc_type}'. "
                f"Register one via POST /v1/templates."
            )
            meta["hints"] = hints
            response.meta = MetaData(**{k: v for k, v in meta.items()})
    else:
        meta["template_used"] = None
        response.meta = MetaData(**{k: v for k, v in meta.items()})

    return response


def find_matching_template(document_type: str, markdown: str) -> Optional[dict]:
    """Sucht ein passendes Template für einen Dokumenttyp (T-MKIT-036).

    Schritt 1: Exakter Match per Template-ID (document_type == template.id).
    Schritt 2: Keyword-Match — alle enabled Templates mit classify_keywords werden
               gegen den Markdown-Text geprüft. Sortierung: priority DESC, matches DESC.

    Returns:
        Template-Dict (mit geparstem 'schema') oder None.
    """
    # Schritt 1: Exakter Match
    tmpl = get_template_by_id(document_type)
    if tmpl is not None:
        return tmpl

    # Schritt 2: Keyword-Match
    try:
        conn = get_db_connection()
        rows = conn.execute(
            "SELECT * FROM template WHERE enabled = 1 AND classify_keywords IS NOT NULL"
        ).fetchall()
        conn.close()
    except Exception:
        return None

    markdown_lower = markdown.lower()
    candidates: list[tuple[dict, int, int]] = []
    for row in rows:
        keywords_raw = row["classify_keywords"] or ""
        keywords = [k.strip().lower() for k in keywords_raw.split(",") if k.strip()]
        if not keywords:
            continue
        matches = sum(1 for kw in keywords if kw in markdown_lower)
        if matches > 0:
            row_dict = dict(row)
            row_dict["schema"] = json.loads(row_dict["schema"]) if isinstance(row_dict.get("schema"), str) else row_dict.get("schema", {})
            if row_dict.get("field_descriptions") and isinstance(row_dict["field_descriptions"], str):
                try:
                    row_dict["field_descriptions"] = json.loads(row_dict["field_descriptions"])
                except Exception:
                    pass
            priority = row["priority"] if row["priority"] is not None else 0
            candidates.append((row_dict, matches, priority))

    if not candidates:
        return None

    # Sortiere: priority DESC, matches DESC
    candidates.sort(key=lambda x: (x[2], x[1]), reverse=True)
    return candidates[0][0]


def _make_null_tolerant(schema: dict) -> dict:
    """Macht ein JSON-Schema null-tolerant: 'type': 'string' → 'type': ['string', 'null'].

    LLMs geben für fehlende Felder korrekt null zurück. Standard JSON-Schema lehnt das ab
    wenn der Typ nur 'string' ist. Diese Funktion patcht rekursiv alle primitiven Typ-Deklarationen.
    """
    import copy
    s = copy.deepcopy(schema)

    def _patch(obj: dict) -> None:
        if not isinstance(obj, dict):
            return
        if "type" in obj and isinstance(obj["type"], str) and obj["type"] != "object" and obj["type"] != "array":
            obj["type"] = [obj["type"], "null"]
        if "properties" in obj and isinstance(obj["properties"], dict):
            for prop in obj["properties"].values():
                _patch(prop)
        if "items" in obj and isinstance(obj["items"], dict):
            _patch(obj["items"])

    _patch(s)
    return s


async def extract_structured_data(
    markdown: str,
    schema: dict,
    language: str = "de",
    field_descriptions: Optional[dict | str] = None,
    notes: Optional[str] = None,
) -> dict:
    """
    Extrahiert strukturierte Daten aus Markdown gemäß einem JSON-Schema via Mistral API.

    Sendet den Markdown-Inhalt zusammen mit dem Schema an die Mistral API.
    Die Antwort wird als JSON geparst und optional gegen das Schema validiert.

    Args:
        markdown: Konvertierter Markdown-Text des Dokuments.
        schema: JSON-Schema für die gewünschten extrahierten Felder.
        language: Sprache des Prompts ('de' oder 'en').
        field_descriptions: Kontext/Beschreibung pro Feld aus der Template-DB (optional).
        notes: Zusätzliche Hinweise aus der Template-DB (optional).

    Returns:
        Dict mit:
        - success (bool)
        - extracted (dict): Extrahierte Daten passend zum Schema
        - tokens (int): Verbrauchte Tokens
        - error (str): Fehlermeldung (nur bei success=False)
    """
    if not MISTRAL_API_KEY:
        log.warning("extract_structured_data_no_api_key")
        return {
            "success": False,
            "error": "MISTRAL_API_KEY nicht konfiguriert",
            "extracted": None,
            "tokens": 0,
        }

    schema_str = json.dumps(schema, ensure_ascii=False)

    # Template-spezifische Hints aus der DB
    template_hints = ""
    if field_descriptions:
        fd = field_descriptions if isinstance(field_descriptions, str) else json.dumps(field_descriptions, indent=2, ensure_ascii=False)
        template_hints += f"\n\nHinweise zu den Feldern:\n{fd}"
    if notes:
        template_hints += f"\n\nZusätzliche Hinweise:\n{notes}"

    # _meta Block Instruktion — wird an beide Sprach-Prompts angehängt
    meta_instruction = (
        "\n\nZUSÄTZLICH zu den Schema-Feldern extrahiere IMMER einen '_meta' Block mit folgenden Feldern:\n"
        + json.dumps(_META_SCHEMA, indent=2, ensure_ascii=False)
        + "\n\nAbsender/Empfänger-Regeln:"
        + "\n- absender: Wer hat das Dokument erstellt/verschickt? Bei Firmen: firma='Telekom Deutschland GmbH', name=null. Bei Personen mit Firma: name='Thomas Weber', firma='Schornsteinfeger Weber'."
        + "\n- empfaenger: An wen ist das Dokument gerichtet? Bei Kassenbons ohne Empfänger: name=null."
        + "\n- Adresse nur befüllen wenn im Dokument erkennbar."
        + "\n- slug: Der geläufigste, kürzeste Name unter dem man den Absender/Empfänger kennt. "
        + "Wie ein Mensch ihn im Alltag nennen würde. Kleinbuchstaben, Bindestriche statt Leerzeichen, "
        + "keine Rechtsformen (GmbH, AG, e.V.), keine Titel (Dr., Dipl.-Ing.), "
        + "keine Abteilungen (Amt für...) — nur der Kern. "
        + "Beispiele: 'Telekom Deutschland GmbH' → 'telekom', "
        + "'Debeka Krankenversicherungsverein a. G.' → 'debeka', "
        + "'Landeshauptstadt Düsseldorf, Amt für Zentrale Dienste, CC Beihilfe' → 'beihilfe', "
        + "'Landesamt für Besoldung und Versorgung NRW' → 'lbv', "
        + "'Thomas Weber, Bevollmächtigter Bezirksschornsteinfeger' → 'weber', "
        + "'Dr. med. Anna Schmidt' → 'dr-schmidt', "
        + "'REWE Sascha Sieger oHG' → 'rewe', "
        + "'Deutsche Rentenversicherung' → 'rentenversicherung', "
        + "'Stadtsparkasse Mönchengladbach' → 'sparkasse-mg', "
        + "'ING-DiBa AG' → 'ing'."
        + "\n\nSteuerrelevanz-Signalwörter (aktiv suchen): "
        + ", ".join(_STEUER_SIGNALWOERTER[:10]) + ", ..."
        + "\nAuch implizit steuerrelevant: Rechnungen mit MwSt, Gehaltsabrechnungen, Versicherungsbeiträge, Handwerkerleistungen."
        + "\n" + _DATENTYP_KONVENTIONEN
    )

    if language == "de":
        system_prompt = (
            "Du bist ein Experte für Dokumentenanalyse und Datenextraktion. "
            "Antworte ausschließlich mit validem JSON."
        )
        user_prompt = (
            "Extrahiere strukturierte Daten aus diesem Dokument gemäß dem JSON-Schema.\n\n"
            "Regeln:\n"
            "- Extrahiere NUR Werte die explizit im Dokument stehen — erfinde KEINE Werte\n"
            "- Fehlende Felder: null (niemals raten oder interpolieren)\n"
            "- Arrays: leeres Array [] wenn keine Einträge vorhanden\n"
            "- Zahlen: exakt wie im Dokument (keine Umrechnung, keine Rundung)\n\n"
            "Antworte AUSSCHLIESSLICH mit dem JSON-Objekt. Kein Markdown, keine Erklärungen.\n\n"
            f"Schema:\n{json.dumps(schema, indent=2, ensure_ascii=False)}"
            f"{template_hints}"
            f"{meta_instruction}\n\n"
            f"Dokument:\n{markdown[:EXTRACT_MAX_CHARS]}"
        )
    else:
        system_prompt = (
            "You are an expert in document analysis and data extraction. "
            "Respond exclusively with valid JSON."
        )
        user_prompt = (
            "Extract structured data from this document according to the JSON schema.\n\n"
            "Rules:\n"
            "- Extract ONLY values explicitly stated in the document — do NOT invent values\n"
            "- Missing fields: null (never guess or interpolate)\n"
            "- Arrays: empty array [] if no entries present\n"
            "- Numbers: exactly as in the document (no conversion, no rounding)\n\n"
            "Respond EXCLUSIVELY with the JSON object. No Markdown, no explanations.\n\n"
            f"Schema:\n{json.dumps(schema, indent=2, ensure_ascii=False)}"
            f"{template_hints}"
            f"{meta_instruction}\n\n"
            f"Document:\n{markdown[:EXTRACT_MAX_CHARS]}"
        )

    payload = {
        "model": MISTRAL_TEXT_MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "max_tokens": EXTRACT_MAX_TOKENS,
        "temperature": 0.0,
    }

    try:
        log.info(
            "extract_structured_data_start",
            schema_keys=list(schema.get("properties", {}).keys()),
            text_length=len(markdown),
        )
        result = await call_mistral_vision_api(payload)
        content = result["choices"][0]["message"]["content"].strip()
        usage = result.get("usage", {})
        tokens = usage.get("total_tokens", 0)

        # JSON aus der Antwort extrahieren (Modell könnte Markdown-Code-Blöcke liefern)
        json_match = re.search(r"\{[\s\S]*\}", content)
        if not json_match:
            log.warning("extract_structured_data_no_json", raw_content=content[:200])
            return {
                "success": False,
                "error": f"Kein JSON in der API-Antwort gefunden: {content[:100]}",
                "extracted": None,
                "tokens": tokens,
            }

        extracted = json.loads(json_match.group(0))

        # Schema-Validierung (AC-014-7) — null-tolerant
        if JSONSCHEMA_AVAILABLE:
            try:
                # Schema null-tolerant machen: "type": "string" → "type": ["string", "null"]
                tolerant_schema = _make_null_tolerant(schema)
                jsonschema.validate(instance=extracted, schema=tolerant_schema)
                log.info("extract_structured_data_valid", tokens=tokens)
            except jsonschema.ValidationError as ve:
                log.warning(
                    "extract_structured_data_schema_violation",
                    error=str(ve.message),
                    tokens=tokens,
                )
                # Graceful: trotzdem zurückgeben, aber Warnung im Log
        else:
            log.debug("extract_structured_data_no_jsonschema")

        log.info("extract_structured_data_success", tokens=tokens)
        return {
            "success": True,
            "extracted": extracted,
            "tokens": tokens,
        }

    except json.JSONDecodeError as exc:
        log.warning("extract_structured_data_json_decode_error", error=str(exc))
        return {
            "success": False,
            "error": f"JSON-Parsing fehlgeschlagen: {str(exc)}",
            "extracted": None,
            "tokens": 0,
        }
    except Exception as exc:
        log.error("extract_structured_data_api_error", error=str(exc))
        return {
            "success": False,
            "error": f"API-Fehler bei Extraktion: {str(exc)}",
            "extracted": None,
            "tokens": 0,
        }


# =============================================================================
# Quality Scoring (FR-MKIT-010)
# =============================================================================

def calculate_quality_score(markdown: str, meta: dict) -> dict:
    """
    Berechnet einen Qualitäts-Score für konvertierten Markdown-Text.

    Scoring-Komponenten:
    - Zeichendichte (0-0.3): Verhältnis sinnvoller Zeichen zu Whitespace
    - Wort-Qualität (0-0.3): Anteil erkennbarer Wörter (>= 3 Buchstaben)
    - Struktur-Elemente (0-0.2): Headings, Listen, Tabellen, Code-Blöcke
    - OCR/Vision Confidence (0-0.2): Token-Effizienz als Proxy bei vision_used

    Args:
        markdown: Konvertierter Markdown-Text
        meta: Metadaten-Dictionary (kann vision_used, tokens_* enthalten)

    Returns:
        Dict mit quality_score (0.0-1.0) und quality_grade ('poor'/'fair'/'good'/'excellent')
    """
    if not markdown or not markdown.strip():
        return {"quality_score": 0.0, "quality_grade": "poor"}

    text = markdown.strip()
    total_chars = len(text)

    # --- Komponente 1: Zeichendichte (0-0.3) ---
    # Verhältnis von Nicht-Whitespace zu Gesamt-Zeichen
    non_ws_chars = len(re.sub(r'\s', '', text))
    if total_chars > 0:
        density_ratio = non_ws_chars / total_chars
        # Gute Dichte ist 0.4-0.8; sehr niedrig (<0.2) oder sehr hoch (>0.9) ist suspicious
        if density_ratio < 0.1:
            density_score = 0.0
        elif density_ratio < 0.2:
            density_score = density_ratio * 1.0  # linear bis 0.2
        elif density_ratio <= 0.8:
            density_score = 0.3  # Optimal
        else:
            density_score = max(0.1, 0.3 * (1.0 - density_ratio))
    else:
        density_score = 0.0

    # --- Komponente 2: Wort-Qualität (0-0.3) ---
    # Anteil erkennbarer Wörter (mindestens 3 Buchstaben, keine Gibberish-Sequenzen)
    words = re.findall(r'[a-zA-ZäöüÄÖÜß]{3,}', text)
    total_word_tokens = re.findall(r'\S+', text)

    if total_word_tokens:
        word_ratio = len(words) / len(total_word_tokens)
        word_score = min(0.3, word_ratio * 0.3)
    else:
        word_score = 0.0

    # Bonus: Mindestlänge — sehr kurzer Text bekommt Abzug
    if total_chars < 50:
        word_score *= 0.5

    # --- Komponente 3: Struktur-Elemente (0-0.2) ---
    structure_score = 0.0
    lines = text.split('\n')

    has_headings = any(line.strip().startswith('#') for line in lines)
    has_lists = any(re.match(r'^\s*[-*+]\s', line) or re.match(r'^\s*\d+\.\s', line) for line in lines)
    has_tables = any('|' in line and line.count('|') >= 2 for line in lines)
    has_codeblocks = '```' in text

    structure_elements = sum([has_headings, has_lists, has_tables, has_codeblocks])
    structure_score = min(0.2, structure_elements * 0.05)

    # --- Komponente 4: OCR/Vision Confidence (0-0.2) ---
    vision_score = 0.0
    vision_used = meta.get("vision_used", False)
    scanned = meta.get("scanned", False)

    if vision_used or scanned:
        tokens_prompt = meta.get("tokens_prompt") or 0
        tokens_completion = meta.get("tokens_completion") or 0

        if tokens_prompt > 0 and tokens_completion > 0:
            # Token-Effizienz: mehr Output-Tokens relativ zu Input → guter Inhalt extrahiert
            efficiency = tokens_completion / tokens_prompt
            if efficiency >= 0.5:
                vision_score = 0.2
            elif efficiency >= 0.2:
                vision_score = 0.15
            elif efficiency >= 0.05:
                vision_score = 0.1
            else:
                vision_score = 0.05
        elif tokens_completion > 0:
            # Nur Completion bekannt → Mindest-Score
            vision_score = 0.1
        else:
            # Vision wurde verwendet, aber keine Token-Daten → neutraler Wert
            vision_score = 0.1
    else:
        # Kein Vision → volle 0.2 als Baseline (keine Unsicherheit durch OCR)
        vision_score = 0.2

    # --- Gesamt-Score ---
    raw_score = density_score + word_score + structure_score + vision_score
    quality_score = round(min(1.0, max(0.0, raw_score)), 4)

    # --- Grade Mapping ---
    if quality_score < 0.3:
        quality_grade = "poor"
    elif quality_score < 0.6:
        quality_grade = "fair"
    elif quality_score < 0.8:
        quality_grade = "good"
    else:
        quality_grade = "excellent"

    log.debug(
        "quality_score_calculated",
        score=quality_score,
        grade=quality_grade,
        density=density_score,
        words=word_score,
        structure=structure_score,
        vision=vision_score,
    )

    return {"quality_score": quality_score, "quality_grade": quality_grade}


# =============================================================================
# Smart Chunking (FR-MKIT-011)
# =============================================================================

def chunk_markdown(markdown: str, chunk_size: int = 512, source: str = "") -> list[dict]:
    """
    Splittet Markdown intelligent an Heading-Grenzen für RAG-Anwendungen.

    Algorithmus:
    1. Identifiziert alle Headings (# ## ### etc.) mit ihren Positionen
    2. Schützt "atomare" Blöcke: Tabellen (| ... |) und Code-Blöcke (``` ... ```)
    3. Splittet an Headings; wenn ein Chunk > chunk_size Tokens, werden Absätze
       (doppelte Newlines) als sekundäre Split-Punkte genutzt
    4. Tabellen und Code-Blöcke werden niemals zerstückelt

    Args:
        markdown: Der zu chunkende Markdown-Text
        chunk_size: Maximale Chunk-Größe in Tokens (Heuristik: len(text) / 4)
        source: Quelldatei-Name (wird in Chunk-Metadaten eingebettet)

    Returns:
        Liste von Chunk-Dicts mit: index, heading, source, token_count, text
    """
    if not markdown or not markdown.strip():
        return []

    def _token_count(text: str) -> int:
        """Heuristische Token-Schätzung: len / 4."""
        return max(1, len(text) // 4)

    def _is_atomic_block_start(line: str) -> tuple[bool, str]:
        """Prüft ob eine Zeile den Start eines atomaren Blocks markiert."""
        stripped = line.strip()
        if stripped.startswith("```"):
            return True, "code"
        if stripped.startswith("|") and stripped.endswith("|"):
            return True, "table"
        return False, ""

    def _split_into_sections(text: str) -> list[tuple[str, str]]:
        """
        Teilt Text in (heading, content)-Paare auf.
        Heading ist leer für Inhalt vor dem ersten Heading.
        Atomare Blöcke (Code/Tabellen) werden nicht zerstückelt.
        """
        lines = text.split("\n")
        sections: list[tuple[str, str]] = []
        current_heading = ""
        current_lines: list[str] = []
        in_code_block = False
        in_table = False

        for line in lines:
            stripped = line.strip()

            # Code-Block tracking
            if stripped.startswith("```"):
                if in_code_block:
                    # Ende des Code-Blocks
                    current_lines.append(line)
                    in_code_block = False
                    continue
                else:
                    in_code_block = True
                    current_lines.append(line)
                    continue

            if in_code_block:
                current_lines.append(line)
                continue

            # Tabellen-Tracking (Zeilen die mit | anfangen und enden)
            if stripped.startswith("|") and stripped.endswith("|"):
                in_table = True
                current_lines.append(line)
                continue
            else:
                in_table = False

            # Heading-Erkennung (nur außerhalb atomarer Blöcke)
            heading_match = re.match(r"^(#{1,6})\s+(.+)$", line)
            if heading_match:
                # Speichert bisherigen Abschnitt
                if current_lines or current_heading:
                    sections.append((current_heading, "\n".join(current_lines)))
                current_heading = line.strip()
                current_lines = []
            else:
                current_lines.append(line)

        # Letzten Abschnitt speichern
        if current_lines or current_heading:
            sections.append((current_heading, "\n".join(current_lines)))

        return sections

    def _split_at_paragraphs(heading: str, content: str, chunk_size: int) -> list[tuple[str, str]]:
        """
        Splittet langen Inhalt an Absatz-Grenzen (doppelte Newlines).
        Respektiert Code-Blöcke und Tabellen als atomare Einheiten.
        """
        # Wenn unter Größenlimit: direkt zurückgeben
        full_text = (heading + "\n\n" + content).strip() if heading else content.strip()
        if _token_count(full_text) <= chunk_size:
            return [(heading, content)]

        # Teile an doppelten Newlines auf, aber behalte Code/Tabellen zusammen
        paragraphs: list[str] = []
        current_para: list[str] = []
        in_code = False
        in_table = False

        for line in content.split("\n"):
            stripped = line.strip()

            if stripped.startswith("```"):
                if in_code:
                    current_para.append(line)
                    in_code = False
                else:
                    in_code = True
                    current_para.append(line)
                continue

            if in_code:
                current_para.append(line)
                continue

            if stripped.startswith("|") and stripped.endswith("|"):
                in_table = True
                current_para.append(line)
                continue
            else:
                if in_table and stripped == "":
                    # Ende der Tabelle: speichern als atomaren Absatz
                    paragraphs.append("\n".join(current_para))
                    current_para = []
                    in_table = False
                    continue
                in_table = False

            if stripped == "" and current_para and not in_code and not in_table:
                # Paragraph-Grenze
                paragraphs.append("\n".join(current_para))
                current_para = []
            else:
                current_para.append(line)

        if current_para:
            paragraphs.append("\n".join(current_para))

        # Paragraphen zu Chunks zusammenfassen (so viel wie möglich unter chunk_size)
        result_sections: list[tuple[str, str]] = []
        current_chunk_lines: list[str] = []
        current_tokens = _token_count(heading) if heading else 0

        for para in paragraphs:
            if not para.strip():
                continue
            para_tokens = _token_count(para)

            if current_chunk_lines and (current_tokens + para_tokens > chunk_size):
                # Aktuellen Chunk abschließen
                result_sections.append((heading, "\n\n".join(current_chunk_lines)))
                current_chunk_lines = [para]
                current_tokens = ((_token_count(heading) if heading else 0) + para_tokens)
            else:
                current_chunk_lines.append(para)
                current_tokens += para_tokens

        if current_chunk_lines:
            result_sections.append((heading, "\n\n".join(current_chunk_lines)))

        return result_sections if result_sections else [(heading, content)]

    # Schritt 1: Text in Heading-Sektionen aufteilen
    sections = _split_into_sections(markdown)

    # Schritt 2: Lange Sektionen an Absatz-Grenzen weiter aufteilen
    fine_sections: list[tuple[str, str]] = []
    for heading, content in sections:
        sub_sections = _split_at_paragraphs(heading, content, chunk_size)
        fine_sections.extend(sub_sections)

    # Schritt 3: Chunks mit Metadaten erzeugen
    chunks: list[dict] = []
    for idx, (heading, content) in enumerate(fine_sections):
        # Chunk-Text zusammensetzen
        if heading and content.strip():
            chunk_text = heading + "\n\n" + content.strip()
        elif heading:
            chunk_text = heading
        else:
            chunk_text = content.strip()

        if not chunk_text:
            continue

        chunks.append({
            "index": len(chunks),
            "heading": heading,
            "source": source,
            "token_count": _token_count(chunk_text),
            "text": chunk_text,
        })

    log.debug(
        "chunk_markdown_done",
        source=source,
        chunk_count=len(chunks),
        chunk_size=chunk_size,
        total_tokens=sum(c["token_count"] for c in chunks),
    )

    return chunks


# =============================================================================
# Hilfsfunktion: PDF-Seite als Bild rendern (Dual-Pass Validierung)
# =============================================================================

def render_first_page_as_image(file_path: Path) -> Optional[tuple[bytes, str]]:
    """Rendert die erste Seite eines PDFs als Bild für Dual-Pass Validierung.

    Args:
        file_path: Pfad zur PDF-Datei.

    Returns:
        Tuple (image_bytes, mimetype) oder None wenn nicht verfügbar/fehlgeschlagen.
    """
    try:
        if not PDF2IMAGE_AVAILABLE:
            return None
        images = convert_from_path(str(file_path), dpi=PDF_RENDER_DPI, first_page=1, last_page=1)
        if not images:
            return None
        first_page = images[0]
        if first_page.width > IMAGE_MAX_WIDTH:
            ratio = IMAGE_MAX_WIDTH / first_page.width
            new_height = int(first_page.height * ratio)
            first_page = first_page.resize(
                (IMAGE_MAX_WIDTH, new_height), Image.Resampling.LANCZOS
            )
        buf = io.BytesIO()
        first_page.save(buf, format="PNG")
        return buf.getvalue(), "image/png"
    except Exception as e:
        log.warning("render_first_page_failed", error=str(e))
        return None


# =============================================================================
# OCR Embed (T-MKIT-033)
# =============================================================================

def embed_ocr_in_pdf(
    file_data: bytes,
    ocr_text: str,
    pages_text: list[str] | None = None,
) -> bytes | None:
    """
    Embeds OCR text as an invisible text layer into a PDF, making it searchable.

    Args:
        file_data: Raw PDF bytes.
        ocr_text: Full OCR text to embed. Used when pages_text is None.
        pages_text: Optional per-page OCR text. If provided, each element is
                    embedded on its corresponding page. Falls back to ocr_text
                    distributed equally across pages when None.

    Returns:
        PDF bytes with embedded invisible text layer, or None on failure.
    """
    if not PYMUPDF_AVAILABLE:
        log.warning("embed_ocr_in_pdf_skipped_no_pymupdf")
        return None
    try:
        doc = fitz.open(stream=file_data, filetype="pdf")
        num_pages = len(doc)

        for page_num, page in enumerate(doc):
            # Determine text for this page
            if pages_text is not None and page_num < len(pages_text):
                page_text = pages_text[page_num]
            else:
                # Distribute full OCR text equally across pages
                if num_pages > 0:
                    chars_per_page = max(1, len(ocr_text) // num_pages)
                    start = page_num * chars_per_page
                    end = start + chars_per_page if page_num < num_pages - 1 else len(ocr_text)
                    page_text = ocr_text[start:end]
                else:
                    page_text = ocr_text

            if not page_text.strip():
                continue

            # Insert invisible text layer: render_mode=3 makes text invisible
            # but still selectable/searchable by PDF viewers and search engines.
            page.insert_text(
                fitz.Point(0, 12),
                page_text,
                fontsize=1,
                render_mode=3,
            )

        result_bytes = doc.tobytes()
        doc.close()
        return result_bytes
    except Exception as e:
        log.warning("embed_ocr_in_pdf_failed", error=str(e))
        return None


# =============================================================================
# Core Konvertierungs-Logik
# =============================================================================

async def convert_auto(
    file_data: bytes,
    filename: str,
    source: str,
    source_type: str,
    input_meta: dict[str, Any],
    prompt: Optional[str] = None,
    language: str = "de",
    describe_images: bool = False,
    classify: bool = False,
    classify_categories: list[str] | None = None,
    extract_schema: Optional[dict] = None,
    ocr_correct: bool = False,
    show_formulas: bool = False,
    chunk: bool = False,
    chunk_size: int = 512,
    accuracy: str = "standard",
    ocr_embed: bool = False,
    auto_extract: bool = False,
    min_confidence: float = 0.7,
) -> ConvertResponse:
    """
    Intelligente Konvertierung basierend auf Dateityp.

    Args:
        file_data: Rohe Datei-Bytes
        filename: Dateiname (wird für Extension-Erkennung genutzt)
        source: Quell-Pfad oder -Bezeichnung (für Metadaten)
        source_type: 'file', 'base64' oder 'url'
        input_meta: Beliebige Pass-through-Metadaten
        prompt: Optionaler Custom-Prompt für Vision
        language: Antwortsprache ('de' oder 'en')
        describe_images: Eingebettete Bilder in DOCX/PPTX durch Pixtral beschreiben
        classify: Dokumenttyp nach Konvertierung via LLM klassifizieren
        classify_categories: Erlaubte Dokumenttypen (überschreibt DEFAULT_CLASSIFY_CATEGORIES)
        extract_schema: JSON-Schema für strukturierte Daten-Extraktion (AC-014-1)
        ocr_correct: OCR-Nachkorrektur via LLM aktivieren (AC-015-5)
        show_formulas: Excel-Formeln im Output annotieren (FR-MKIT-007)
        chunk: Smart Chunking für RAG aktivieren (FR-MKIT-011)
        chunk_size: Maximale Chunk-Größe in Tokens (Default: 512)
        accuracy: Accuracy-Modus: 'standard' (Default) oder 'high'. High aktiviert
                  automatische OCR-Correction und Dual-Pass Vision-Validierung (T-MKIT-020).
        ocr_embed: Wenn True und das Dokument ein gescanntes PDF ist, wird der OCR-Text
                   als unsichtbare Textschicht eingebettet (T-MKIT-033).
    """
    start_time = time.time()
    ext = get_file_extension(filename)
    mimetype = detect_mimetype_from_bytes(file_data) or get_mimetype(Path(filename))

    meta = {
        **input_meta,
        "source": source,
        "source_type": source_type,
        "format": ext.lstrip("."),
        "size_bytes": len(file_data),
    }

    # Größenprüfung
    if len(file_data) > MAX_FILE_SIZE_BYTES:
        meta["duration_ms"] = int((time.time() - start_time) * 1000)
        return create_error_response(
            ErrorCode.FILE_TOO_LARGE,
            f"Datei zu groß: {len(file_data) / 1024 / 1024:.1f}MB (Max: {MAX_FILE_SIZE_MB}MB)",
            meta=meta
        )

    # T-MKIT-020: Accuracy-Modus immer in Meta dokumentieren
    meta["accuracy_mode"] = accuracy

    # T-MKIT-023: High-Accuracy ohne API-Key → Warning + Degradierung auf Standard
    # Nur für Dateitypen die Mistral API nutzen (Bilder, PDFs) — nicht für Audio/Video (Whisper)
    _needs_mistral_api = ext in IMAGE_EXTENSIONS or ext in MARKITDOWN_EXTENSIONS
    if accuracy == "high" and not MISTRAL_API_KEY and _needs_mistral_api:
        log.warning("high_accuracy_degraded_no_api_key")
        meta["accuracy_warning"] = "High accuracy requested but MISTRAL_API_KEY not set — running in standard mode"
        accuracy = "standard"
        meta["accuracy_mode"] = accuracy

    # Bild → Vision
    if ext in IMAGE_EXTENSIONS or (mimetype and mimetype.startswith("image/")):
        processed_data, resize_meta = resize_image_if_needed(file_data)
        meta.update(resize_meta)
        meta["vision_used"] = True

        vision_prompt = prompt or (
            "Analysiere dieses Bild und gib den Inhalt als Markdown zurück.\n\n"
            "- Wenn Text sichtbar ist: extrahiere ihn vollständig und strukturiert "
            "(Überschriften, Listen, Tabellen in Markdown-Syntax)\n"
            "- Wenn es ein Diagramm, Chart oder Grafik ist: beschreibe die dargestellten Daten präzise\n"
            "- Wenn es ein Foto ohne Text ist: beschreibe den Bildinhalt in einem kurzen Absatz\n"
            "- Wenn die Bildqualität zu schlecht ist: schreibe [UNLESERLICH]\n\n"
            "Antworte ausschließlich mit dem Markdown-Ergebnis."
        )

        result = await analyze_with_mistral_vision(
            processed_data,
            mimetype or "image/jpeg",
            vision_prompt,
            language
        )

        meta["duration_ms"] = int((time.time() - start_time) * 1000)

        if result["success"]:
            meta["vision_model"] = result.get("vision_model")
            meta["tokens_prompt"] = result.get("tokens_prompt")
            meta["tokens_completion"] = result.get("tokens_completion")
            meta["tokens_total"] = result.get("tokens_total")

            # T-MKIT-027: EXIF/GPS/IPTC Metadaten aus Original-Bilddaten extrahieren
            try:
                img_metadata = extract_image_metadata(file_data)
                if img_metadata.get("exif"):
                    meta["exif"] = img_metadata["exif"]
                if img_metadata.get("iptc"):
                    meta["iptc"] = img_metadata["iptc"]
            except Exception as _exif_err:
                log.debug("exif_extraction_skipped", error=str(_exif_err))

            markdown = result["markdown"]
            pipeline_steps: list[str] = ["vision"]

            # AC-015: OCR-Nachkorrektur via LLM (T-MKIT-022: auch bei accuracy="high" automatisch aktiv)
            if ocr_correct or accuracy == "high":
                log.info("ocr_correct_start", path="vision", accuracy=accuracy)
                correction = await correct_ocr_text(markdown, language=language)
                if correction["success"]:
                    markdown = correction["corrected_text"]
                    meta["ocr_corrected"] = True
                    meta["ocr_corrections_count"] = correction["corrections_count"]
                    log.info("ocr_correct_done_vision", corrections=correction["corrections_count"])
                else:
                    log.warning("ocr_correct_failed_vision", error=correction.get("error"))
                    meta["ocr_corrected"] = False

            # T-MKIT-020: High-Accuracy-Pipeline für Bilder
            if accuracy == "high":
                log.info("high_accuracy_image_dual_pass_start", file=filename)
                effective_mimetype = mimetype or "image/jpeg"
                markdown = await dual_pass_validate(
                    markdown=markdown,
                    file_data=processed_data,
                    mimetype=effective_mimetype,
                    language=language,
                )
                pipeline_steps.append("dual_pass_validation")

            meta["pipeline_steps"] = pipeline_steps

            if classify:
                classify_result = await classify_document(markdown, classify_categories, language)
                meta.update(classify_result)
            # AC-010: Quality Scoring
            meta.update(calculate_quality_score(markdown, meta))
            # T-MKIT-032: Response hints for LLM consumers
            hints: list[str] = []
            if not extract_schema and not auto_extract and meta.get("document_type") == "invoice":
                hints.append("This document was classified as an invoice. Add template='invoice' to extract structured fields (invoice_number, total, line_items, etc.).")
            if not extract_schema and not auto_extract and meta.get("document_type") and meta.get("document_type") != "other":
                hints.append("Add auto_extract=true to automatically extract structured data based on document type.")
            if meta.get("quality_grade") == "poor" and accuracy != "high":
                hints.append("Low quality score detected. Try accuracy='high' for better results on scanned or complex documents.")
            if hints:
                meta["hints"] = hints
            response = create_success_response(markdown, meta=meta)
            # T-MKIT-036: Auto-Extract (classify → template lookup → extraction)
            if auto_extract and not extract_schema:
                response = await _apply_auto_extract(response, meta, markdown, language, min_confidence, hints)
            # AC-014-2/AC-014-3: Strukturierte Extraktion falls Schema gesetzt
            elif extract_schema:
                extraction = await extract_structured_data(markdown, extract_schema, language)
                if extraction["success"]:
                    response.extracted = extraction["extracted"]
                    if accuracy == "high":
                        pipeline_steps_updated = list(meta.get("pipeline_steps", pipeline_steps))
                        if "schema_extraction" not in pipeline_steps_updated:
                            pipeline_steps_updated.append("schema_extraction")
                        meta["pipeline_steps"] = pipeline_steps_updated
                        response.meta = MetaData(**{
                            k: v for k, v in meta.items()
                        })
                else:
                    log.warning("extract_structured_data_failed_vision", error=extraction.get("error"))
            # FR-MKIT-011: Smart Chunking für RAG
            if chunk:
                response.chunks = chunk_markdown(markdown, chunk_size=chunk_size, source=source)
            return response
        else:
            return create_error_response(
                result.get("error_code", ErrorCode.VISION_FAILED),
                result["error"],
                meta=meta
            )

    # Audio/Video → faster-whisper Transkription (FR-MKIT-006)
    elif ext in AUDIO_EXTENSIONS or ext in VIDEO_EXTENSIONS:
        temp_media_path = TEMP_DIR / f"{hashlib.md5(file_data).hexdigest()}_{filename}"
        extracted_wav: Optional[Path] = None
        try:
            temp_media_path.write_bytes(file_data)
            audio_path = temp_media_path

            # Video: erst Audio-Track extrahieren
            if ext in VIDEO_EXTENSIONS:
                log.info("video_audio_extract_start", file=filename)
                try:
                    extracted_wav = extract_audio_from_video(temp_media_path)
                    audio_path = extracted_wav
                except RuntimeError as exc:
                    meta["duration_ms"] = int((time.time() - start_time) * 1000)
                    return create_error_response(
                        ErrorCode.CONVERSION_FAILED,
                        f"Audio-Extraktion fehlgeschlagen: {str(exc)}",
                        meta=meta,
                    )

            # Transkribieren
            transcription = transcribe_audio(audio_path)
            meta["duration_ms"] = int((time.time() - start_time) * 1000)

            if not transcription["success"]:
                return create_error_response(
                    ErrorCode.CONVERSION_FAILED,
                    transcription.get("error", "Transkription fehlgeschlagen"),
                    meta=meta,
                )

            # Meta-Daten setzen (AC-006-6)
            meta["language"] = transcription.get("language", "unknown")
            meta["duration_seconds"] = transcription.get("duration", 0.0)
            meta["whisper_model"] = transcription.get("model_size", WHISPER_MODEL_SIZE)
            meta["accuracy_mode"] = accuracy  # T-MKIT-022: accuracy_mode auch bei Audio/Video

            transcript_text = transcription.get("text", "")
            markdown = f"# Transkription\n\n{transcript_text}"

            log.info(
                "audio_transcription_done",
                file=filename,
                language=meta["language"],
                duration=meta["duration_seconds"],
                chars=len(transcript_text),
            )

            # AC-010: Quality Scoring
            meta.update(calculate_quality_score(markdown, meta))
            response = create_success_response(markdown, meta=meta)
            # FR-MKIT-011: Smart Chunking für RAG
            if chunk:
                response.chunks = chunk_markdown(markdown, chunk_size=chunk_size, source=source)
            return response

        finally:
            temp_media_path.unlink(missing_ok=True)
            if extracted_wav is not None:
                extracted_wav.unlink(missing_ok=True)

    # Dokument → MarkItDown (mit optionalem Scanned-PDF-Routing)
    elif ext in MARKITDOWN_EXTENSIONS or ext:
        temp_path = TEMP_DIR / f"{hashlib.md5(file_data).hexdigest()}_{filename}"
        try:
            temp_path.write_bytes(file_data)

            # Scanned PDF Detection: VOR dem normalen markitdown-Pfad prüfen
            if ext == ".pdf" and is_scanned_pdf(temp_path):
                log.info("scanned_pdf_detected", file=filename)
                result = await convert_scanned_pdf(temp_path, language=language)
                meta["duration_ms"] = int((time.time() - start_time) * 1000)

                if result["success"]:
                    meta["scanned"] = True
                    meta["vision_model"] = result.get("vision_model")
                    meta["ocr_model"] = result.get("ocr_model")
                    meta["tokens_prompt"] = result.get("tokens_prompt")
                    meta["tokens_completion"] = result.get("tokens_completion")
                    meta["tokens_total"] = result.get("tokens_total")
                    meta["tokens_per_page"] = result.get("tokens_per_page")
                    meta["pages_processed"] = result.get("pages_processed") or result.get("pages")
                    if result.get("ocr_model"):
                        # OCR3-Pfad: kein Vision
                        meta["vision_used"] = False
                    else:
                        meta["vision_used"] = True

                    scanned_markdown = result["markdown"]
                    scanned_pipeline_steps: list[str] = ["ocr"]

                    # AC-015: OCR-Nachkorrektur via LLM (nur wenn explizit aktiviert ODER accuracy=high)
                    if ocr_correct or accuracy == "high":
                        log.info("ocr_correct_start", path="scanned_pdf", accuracy=accuracy)
                        correction = await correct_ocr_text(scanned_markdown, language=language)
                        if correction["success"]:
                            scanned_markdown = correction["corrected_text"]
                            meta["ocr_corrected"] = True
                            meta["ocr_corrections_count"] = correction["corrections_count"]
                            log.info("ocr_correct_done_scanned_pdf", corrections=correction["corrections_count"])
                            scanned_pipeline_steps.append("ocr_correction")
                        else:
                            log.warning("ocr_correct_failed_scanned_pdf", error=correction.get("error"))
                            meta["ocr_corrected"] = False

                    # T-MKIT-020: High-Accuracy → Dual-Pass Validation für gescannte PDFs
                    if accuracy == "high":
                        log.info("high_accuracy_scanned_pdf_dual_pass_start", file=filename)
                        rendered = render_first_page_as_image(temp_path)
                        if rendered is not None:
                            page_image_bytes, page_mimetype = rendered
                            scanned_markdown = await dual_pass_validate(
                                markdown=scanned_markdown,
                                file_data=page_image_bytes,
                                mimetype=page_mimetype,
                                language=language,
                            )
                            scanned_pipeline_steps.append("dual_pass_validation")
                        else:
                            log.warning("high_accuracy_scanned_pdf_dual_pass_skipped_no_pdf2image")

                    meta["pipeline_steps"] = scanned_pipeline_steps

                    if classify:
                        classify_result = await classify_document(scanned_markdown, classify_categories, language)
                        meta.update(classify_result)
                    # AC-010: Quality Scoring
                    meta.update(calculate_quality_score(scanned_markdown, meta))
                    # T-MKIT-032: Response hints for LLM consumers
                    scanned_hints: list[str] = []
                    if not extract_schema and not auto_extract and meta.get("document_type") == "invoice":
                        scanned_hints.append("This document was classified as an invoice. Add template='invoice' to extract structured fields (invoice_number, total, line_items, etc.).")
                    if not extract_schema and not auto_extract and meta.get("document_type") and meta.get("document_type") != "other":
                        scanned_hints.append("Add auto_extract=true to automatically extract structured data based on document type.")
                    if meta.get("quality_grade") == "poor" and accuracy != "high":
                        scanned_hints.append("Low quality score detected. Try accuracy='high' for better results on scanned or complex documents.")
                    if not ocr_correct and accuracy != "high":
                        scanned_hints.append("This is a scanned document. Enable ocr_correct=true or accuracy='high' to fix common OCR errors.")
                    if meta.get("scanned") and not ocr_embed:
                        scanned_hints.append("This scanned PDF has no searchable text layer. Add ocr_embed=true to create a searchable PDF with embedded OCR text.")
                    if scanned_hints:
                        meta["hints"] = scanned_hints
                    response = create_success_response(scanned_markdown, meta=meta)
                    # T-MKIT-033: OCR Embed — invisible text layer in scanned PDF
                    if ocr_embed:
                        pages_text_list: list[str] | None = None
                        if result.get("pages_text"):
                            pages_text_list = result["pages_text"]
                        embedded = embed_ocr_in_pdf(file_data, scanned_markdown, pages_text_list)
                        if embedded is not None:
                            response.enriched_pdf = base64.b64encode(embedded).decode("utf-8")
                            log.info("ocr_embed_done", size_bytes=len(embedded))
                        else:
                            log.warning("ocr_embed_failed_no_output")
                    # T-MKIT-036: Auto-Extract (classify → template lookup → extraction)
                    if auto_extract and not extract_schema:
                        response = await _apply_auto_extract(response, meta, scanned_markdown, language, min_confidence, scanned_hints)
                    # AC-014-2/AC-014-3: Strukturierte Extraktion falls Schema gesetzt
                    elif extract_schema:
                        extraction = await extract_structured_data(scanned_markdown, extract_schema, language)
                        if extraction["success"]:
                            response.extracted = extraction["extracted"]
                            if accuracy == "high":
                                updated_steps = list(meta.get("pipeline_steps", scanned_pipeline_steps))
                                if "schema_extraction" not in updated_steps:
                                    updated_steps.append("schema_extraction")
                                meta["pipeline_steps"] = updated_steps
                                response.meta = MetaData(**{
                                    k: v for k, v in meta.items()
                                })
                        else:
                            log.warning("extract_structured_data_failed_scanned_pdf", error=extraction.get("error"))
                    # FR-MKIT-011: Smart Chunking für RAG
                    if chunk:
                        response.chunks = chunk_markdown(scanned_markdown, chunk_size=chunk_size, source=source)
                    return response
                else:
                    return create_error_response(
                        result.get("error_code", ErrorCode.CONVERSION_FAILED),
                        result["error"],
                        meta=meta
                    )

            result = convert_with_markitdown(temp_path, show_formulas=show_formulas)
            meta["duration_ms"] = int((time.time() - start_time) * 1000)

            if result["success"]:
                if result.get("title"):
                    meta["title"] = result["title"]

                # FR-MKIT-007: Excel-spezifische Metadaten
                if ext in {".xlsx", ".xls"}:
                    if result.get("sheets_count") is not None:
                        meta["sheets_count"] = result["sheets_count"]
                    if result.get("charts_count") is not None:
                        meta["charts_count"] = result["charts_count"]
                    # T-MKIT-026: Hidden sheets
                    if result.get("hidden_sheets"):
                        meta["hidden_sheets"] = result["hidden_sheets"]

                # T-MKIT-024: ZUGFeRD/Factur-X Daten aus convert_with_markitdown übernehmen
                zugferd_extracted = result.get("zugferd")
                if zugferd_extracted is not None:
                    meta["zugferd"] = zugferd_extracted

                # T-MKIT-025: XMP Metadata + Embedded Files übernehmen
                xmp_extracted = result.get("xmp_metadata")
                if xmp_extracted is not None:
                    meta["xmp_metadata"] = xmp_extracted

                embedded_extracted = result.get("embedded_files")
                if embedded_extracted is not None:
                    meta["embedded_files"] = embedded_extracted

                # T-MKIT-028: Office Document Properties übernehmen
                doc_props = result.get("document_properties")
                if doc_props is not None:
                    meta["document_properties"] = doc_props

                # T-MKIT-029: Email Metadata übernehmen
                email_meta = result.get("email_metadata")
                if email_meta is not None:
                    meta["email_routing"] = email_meta.get("routing")
                    meta["email_thread"] = email_meta.get("thread")
                    meta["calendar_events"] = email_meta.get("calendar_events")

                # T-MKIT-030: PPTX Hidden Slides übernehmen
                pptx_hidden = result.get("pptx_hidden_info")
                if pptx_hidden is not None and pptx_hidden.get("hidden_slide_count", 0) > 0:
                    meta["hidden_slides"] = pptx_hidden["hidden_slide_count"]

                markdown = result["markdown"]
                markitdown_pipeline_steps: list[str] = ["markitdown"]

                # Eingebettete Bilder beschreiben (nur für DOCX/PPTX, nur wenn aktiviert)
                if describe_images and ext in {".docx", ".doc", ".pptx", ".ppt"}:
                    log.info("embedded_images_describe_start", file=filename, ext=ext)
                    if ext in {".docx", ".doc"}:
                        images = extract_images_from_docx(temp_path)
                    else:
                        images = extract_images_from_pptx(temp_path)

                    if images:
                        descriptions = await describe_embedded_images(images, language=language)
                        markdown = insert_image_descriptions(markdown, descriptions)
                        meta["images_described"] = len(descriptions)
                        log.info(
                            "embedded_images_described",
                            file=filename,
                            count=len(descriptions),
                        )

                # T-MKIT-020: High-Accuracy → Dual-Pass Validation für PDFs und Bilder
                if accuracy == "high" and ext == ".pdf":
                    log.info("high_accuracy_pdf_dual_pass_start", file=filename)
                    rendered = render_first_page_as_image(temp_path)
                    if rendered is not None:
                        page_image_bytes, page_mimetype = rendered
                        markdown = await dual_pass_validate(
                            markdown=markdown,
                            file_data=page_image_bytes,
                            mimetype=page_mimetype,
                            language=language,
                        )
                        markitdown_pipeline_steps.append("dual_pass_validation")
                    else:
                        log.warning("high_accuracy_pdf_dual_pass_skipped_no_pdf2image")

                meta["pipeline_steps"] = markitdown_pipeline_steps

                if classify:
                    classify_result = await classify_document(markdown, classify_categories, language)
                    meta.update(classify_result)

                # AC-010: Quality Scoring
                meta.update(calculate_quality_score(markdown, meta))
                # T-MKIT-032: Response hints for LLM consumers
                doc_hints: list[str] = []
                if not extract_schema and not auto_extract and meta.get("document_type") == "invoice":
                    doc_hints.append("This document was classified as an invoice. Add template='invoice' to extract structured fields (invoice_number, total, line_items, etc.).")
                if not extract_schema and not auto_extract and meta.get("document_type") and meta.get("document_type") != "other":
                    doc_hints.append("Add auto_extract=true to automatically extract structured data based on document type.")
                if meta.get("quality_grade") == "poor" and accuracy != "high":
                    doc_hints.append("Low quality score detected. Try accuracy='high' for better results on scanned or complex documents.")
                if meta.get("scanned") and not ocr_correct and accuracy != "high":
                    doc_hints.append("This is a scanned document. Enable ocr_correct=true or accuracy='high' to fix common OCR errors.")
                if not describe_images and ext in (".docx", ".pptx") and source_type == "file":
                    doc_hints.append("This document may contain embedded images. Add describe_images=true to describe them via Vision AI.")
                # T-MKIT-024: ZUGFeRD hint wenn erkannt aber kein template/extract_schema
                if meta.get("zugferd") is not None and not extract_schema and not auto_extract:
                    doc_hints.append("This PDF contains embedded ZUGFeRD/Factur-X e-invoice data. Add template='invoice' to get the structured data in the extracted field — no LLM needed, 100% accurate.")
                if doc_hints:
                    meta["hints"] = doc_hints
                response = create_success_response(markdown, meta=meta)
                # T-MKIT-036: Auto-Extract (classify → template lookup → extraction)
                if auto_extract and not extract_schema:
                    response = await _apply_auto_extract(response, meta, markdown, language, min_confidence, doc_hints)
                # T-MKIT-024: ZUGFeRD Daten direkt als extracted verwenden (kein LLM nötig)
                elif extract_schema and meta.get("zugferd") is not None:
                    response.extracted = meta["zugferd"]
                    meta["zugferd_source"] = "embedded_xml"
                    meta["extraction_method"] = "zugferd"
                    response.meta = MetaData(**{k: v for k, v in meta.items()})
                    log.info("zugferd_used_as_extracted", file=filename)
                elif extract_schema:
                    # AC-014-2/AC-014-3: Strukturierte Extraktion falls Schema gesetzt (kein ZUGFeRD)
                    extraction = await extract_structured_data(markdown, extract_schema, language)
                    if extraction["success"]:
                        response.extracted = extraction["extracted"]
                        if accuracy == "high":
                            updated_steps = list(meta.get("pipeline_steps", markitdown_pipeline_steps))
                            if "schema_extraction" not in updated_steps:
                                updated_steps.append("schema_extraction")
                            meta["pipeline_steps"] = updated_steps
                            response.meta = MetaData(**{
                                k: v for k, v in meta.items()
                            })
                    else:
                        log.warning("extract_structured_data_failed_markitdown", error=extraction.get("error"))
                # FR-MKIT-011: Smart Chunking für RAG
                if chunk:
                    response.chunks = chunk_markdown(markdown, chunk_size=chunk_size, source=source)
                return response
            else:
                return create_error_response(
                    result.get("error_code", ErrorCode.CONVERSION_FAILED),
                    result["error"],
                    meta=meta
                )
        finally:
            temp_path.unlink(missing_ok=True)

    else:
        meta["duration_ms"] = int((time.time() - start_time) * 1000)
        return create_error_response(
            ErrorCode.UNSUPPORTED_FORMAT,
            f"Nicht unterstütztes Format: {ext}",
            meta=meta
        )


async def convert_folder_contents(
    folder_path: Path,
    input_meta: dict[str, Any],
    language: str = "de",
) -> ConvertResponse:
    """
    Konvertiert alle Dateien in einem Ordner zu einem zusammengeführten Markdown.
    """
    start_time = time.time()
    log.info("folder_convert_start", folder=str(folder_path))

    if not folder_path.exists():
        return create_error_response(
            ErrorCode.FILE_NOT_FOUND,
            f"Ordner nicht gefunden: {folder_path}",
            meta=input_meta
        )

    if not folder_path.is_dir():
        return create_error_response(
            ErrorCode.INVALID_INPUT,
            f"Kein Ordner: {folder_path}",
            meta=input_meta
        )

    # Dateien sammeln
    files = sorted([
        f for f in folder_path.iterdir()
        if f.is_file() and not should_skip_file(f.name)
    ], key=lambda f: f.name.lower())

    if not files:
        return create_error_response(
            ErrorCode.INVALID_INPUT,
            f"Keine Dateien im Ordner: {folder_path}",
            meta=input_meta
        )

    # Ergebnisse sammeln
    markdown_parts = []
    file_results = []
    total_tokens = 0
    files_processed = 0
    files_failed = 0

    for file_path in files:
        file_meta = {"filename": file_path.name}

        try:
            file_data = file_path.read_bytes()
            result = await convert_auto(
                file_data=file_data,
                filename=file_path.name,
                source=str(file_path),
                source_type="file",
                input_meta={},
                language=language,  # T-MKIT-022: language durchgereicht
            )

            if result.success:
                files_processed += 1
                markdown_parts.append(f"\n\n## {file_path.name}\n\n{result.markdown}")
                file_meta["success"] = True
                file_meta["size_bytes"] = len(file_data)

                if result.meta:
                    if hasattr(result.meta, 'tokens_total') and result.meta.tokens_total:
                        total_tokens += result.meta.tokens_total
                        file_meta["tokens"] = result.meta.tokens_total
                    if hasattr(result.meta, 'vision_used') and result.meta.vision_used:
                        file_meta["vision_used"] = True
            else:
                files_failed += 1
                file_meta["success"] = False
                file_meta["error"] = result.error.message if result.error else "Unknown error"
                markdown_parts.append(f"\n\n## {file_path.name}\n\n*Konvertierung fehlgeschlagen*")

        except Exception as e:
            files_failed += 1
            file_meta["success"] = False
            file_meta["error"] = str(e)
            log.error("folder_file_error", file=file_path.name, error=str(e))

        file_results.append(file_meta)

    # Zusammenführen
    combined_markdown = f"# {folder_path.name}\n" + "".join(markdown_parts)

    meta = {
        **input_meta,
        "source": str(folder_path),
        "source_type": "folder",
        "files_processed": files_processed,
        "files_failed": files_failed,
        "files_total": len(files),
        "tokens_total": total_tokens,
        "files": file_results,
        "duration_ms": int((time.time() - start_time) * 1000),
    }

    log.info("folder_convert_complete",
             folder=str(folder_path),
             processed=files_processed,
             failed=files_failed,
             duration_ms=meta["duration_ms"])

    return create_success_response(combined_markdown, meta=meta)


# =============================================================================
# Tips Helper (T-MKIT-032)
# =============================================================================

_EXAMPLE_URL = "https://example.com"  # Example URL for documentation purposes only

def _build_tips_dict() -> dict:
    """Builds the tips dictionary used by both MCP get_tips tool and REST /v1/tips endpoint."""
    data_dir = str(DATA_DIR)
    return {
        "service": f"Daigestr — Document Intelligence Service v{VERSION}",
        "quick_reference": {
            "convert_file": {"endpoint": "POST /v1/convert", "mcp_tool": "convert", "params": {"path": f"{data_dir}/file.pdf"}},
            "convert_url": {"endpoint": "POST /v1/convert", "mcp_tool": "convert", "params": {"url": _EXAMPLE_URL}},
            "convert_base64": {"endpoint": "POST /v1/convert", "mcp_tool": "convert", "params": {"base64_data": "...", "filename": "file.pdf"}},
            "extract_invoice": {"endpoint": "POST /v1/extract", "mcp_tool": "extract", "params": {"path": f"{data_dir}/invoice.pdf", "template": "invoice"}},
            "extract_custom": {"endpoint": "POST /v1/convert", "mcp_tool": "convert", "params": {"path": f"{data_dir}/doc.pdf", "extract_schema": {"type": "object", "properties": {"title": {"type": "string"}}}}},
        },
        "common_mistakes": [
            {"problem": "extracted is null", "cause": "Missing extract_schema or template parameter", "fix": "Add template='invoice' or extract_schema={...} to get structured JSON in the extracted field"},
            {"problem": "extracted is null with auto_extract=true", "cause": "No template registered for this document type", "fix": "Register a template via POST /v1/templates or check meta.document_type and meta.template_used for details"},
            {"problem": "chunks is null", "cause": "Missing chunk parameter", "fix": "Add chunk=true to get RAG-ready chunks in the chunks field"},
            {"problem": "Poor quality on scanned PDFs", "cause": "Using standard accuracy", "fix": "Set accuracy='high' for OCR correction + dual-pass validation"},
            {"problem": "Images in DOCX/PPTX not described", "cause": "describe_images defaults to false", "fix": "Set describe_images=true (costs extra API calls)"},
            {"problem": "Document type not detected", "cause": "classify defaults to false", "fix": "Set classify=true to get document_type in meta"},
            {"problem": "OCR errors in output", "cause": "No post-correction", "fix": "Set ocr_correct=true or accuracy='high' (auto-enables correction)"},
        ],
        "available_templates": get_all_template_ids(),
        "available_formats": {
            "documents": ["pdf", "docx", "doc", "pptx", "ppt", "xlsx", "xls", "odt", "ods", "odp", "html", "csv", "txt", "md", "rtf", "json", "xml"],
            "images": ["jpg", "jpeg", "png", "gif", "webp", "bmp"],
            "audio": ["mp3", "wav", "ogg", "flac", "m4a"],
            "video": ["mp4", "mkv", "webm", "avi", "mov"],
        },
        "optional_features": {
            "accuracy": {"values": ["standard", "high"], "default": "standard", "description": "high activates OCR correction + dual-pass vision validation for scanned documents"},
            "classify": {"type": "bool", "default": False, "description": "Detect document type (invoice, contract, cv, etc.) with confidence score in meta"},
            "extract_schema": {"type": "dict", "default": None, "description": "JSON Schema for structured extraction — result in 'extracted' field. Without this, extracted is always null."},
            "template": {"type": "str", "default": None, "description": "Shortcut for extract_schema. Available: invoice, cv, contract"},
            "auto_extract": {"type": "bool", "default": False, "description": "Automatically classify document, find matching template, and extract structured data — all in one call. No template or extract_schema needed."},
            "min_confidence": {"type": "float", "default": 0.7, "description": "Minimum classification confidence for auto_extract to use a template (0.0-1.0)"},
            "describe_images": {"type": "bool", "default": False, "description": "Extract and describe embedded images in DOCX/PPTX via Vision AI"},
            "ocr_correct": {"type": "bool", "default": False, "description": "LLM post-correction for OCR errors"},
            "ocr_embed": {"type": "bool", "default": False, "description": "Embed OCR text layer into PDF — creates searchable PDF output"},
            "show_formulas": {"type": "bool", "default": False, "description": "Show Excel formulas in output"},
            "chunk": {"type": "bool", "default": False, "description": "Split output into RAG-ready chunks"},
            "chunk_size": {"type": "int", "default": 512, "description": "Approximate chunk size in tokens"},
            "language": {"type": "str", "default": "de", "description": "Language for Vision/OCR responses"},
            "zugferd": {"type": "bool", "default": "auto", "description": "Extract ZUGFeRD/Factur-X e-invoice data from PDF (auto-detected, 100% accuracy, no LLM required)"},
        },
        "response_fields": {
            "markdown": "Always present — the converted document as Markdown",
            "extracted": "Only present when extract_schema or template is set — structured JSON",
            "chunks": "Only present when chunk=true — list of text segments with metadata",
            "meta": "Always present — processing metadata (quality_score, duration, pipeline_steps, etc.)",
        },
        "v3_meta_fields": {
            "zugferd": "ZUGFeRD/Factur-X structured invoice data (auto-extracted from PDF when present)",
            "xmp_metadata": "PDF XMP metadata (title, creator, keywords, custom properties)",
            "embedded_files": "Files embedded inside the PDF (e.g. original XML in ZUGFeRD)",
            "hidden_sheets": "XLSX sheets with visibility=hidden or xlSheetVeryHidden — extracted regardless",
            "exif": "Image EXIF data (camera, GPS coordinates, exposure, focal length)",
            "iptc": "Image IPTC metadata (creator, copyright, keywords, caption)",
            "document_properties": "Office document core/app/custom properties (author, company, revision, etc.)",
            "email_routing": "Email routing headers (Received chain, SPF, DKIM, DMARC results)",
            "email_thread": "Email thread/conversation metadata (In-Reply-To, References)",
            "calendar_events": "Calendar events extracted from .ics attachments",
            "hidden_slides": "PPTX hidden slides — extracted with hidden=true marker in meta",
        },
        "note_mcp_vs_rest": "MCP tool 'convert' uses 'base64_data' parameter (not 'base64' like REST API)",
    }


# =============================================================================
# REST API Endpoints
# =============================================================================

@app.post("/v1/convert", response_model=ConvertResponse)
async def api_convert(request: ConvertRequest) -> ConvertResponse:
    """
    Konvertiert eine Datei zu Markdown.
    """
    inputs = [request.path, request.base64, request.url]
    if sum(1 for x in inputs if x) != 1:
        return create_error_response(
            ErrorCode.INVALID_INPUT,
            "Genau einer von 'path', 'base64' oder 'url' muss angegeben werden",
            meta=request.meta
        )

    # Template → Schema Auflösung (AC-014-4)
    effective_schema = request.extract_schema
    if request.template and not effective_schema:
        tmpl = get_template_by_id(request.template)
        if tmpl is None:
            return create_error_response(
                ErrorCode.INVALID_INPUT,
                f"Unbekanntes Template: '{request.template}'. Verfügbar: {get_all_template_ids()}",
                meta=request.meta
            )
        effective_schema = tmpl["schema"]

    # Pfad-basiert
    if request.path:
        file_path = resolve_path(request.path)
        if not file_path.exists():
            return create_error_response(
                ErrorCode.FILE_NOT_FOUND,
                f"Datei nicht gefunden: {file_path}",
                meta=request.meta
            )
        file_data = file_path.read_bytes()
        return await convert_auto(
            file_data=file_data,
            filename=file_path.name,
            source=str(file_path),
            source_type="file",
            input_meta=request.meta,
            prompt=request.prompt,
            language=request.language,
            describe_images=request.describe_images,
            classify=request.classify,
            classify_categories=request.classify_categories,
            extract_schema=effective_schema,
            ocr_correct=request.ocr_correct,
            show_formulas=request.show_formulas,
            chunk=request.chunk,
            chunk_size=request.chunk_size,
            accuracy=request.accuracy,
            ocr_embed=request.ocr_embed,
            auto_extract=request.auto_extract,
            min_confidence=request.min_confidence,
        )

    # Base64
    if request.base64:
        if not request.filename:
            return create_error_response(
                ErrorCode.INVALID_INPUT,
                "'filename' ist erforderlich bei Base64-Upload",
                meta=request.meta
            )
        try:
            file_data = base64.b64decode(request.base64)
        except Exception as e:
            return create_error_response(
                ErrorCode.INVALID_BASE64,
                f"Ungültiges Base64: {str(e)}",
                meta=request.meta
            )
        return await convert_auto(
            file_data=file_data,
            filename=request.filename,
            source="base64",
            source_type="base64",
            input_meta=request.meta,
            prompt=request.prompt,
            language=request.language,
            describe_images=request.describe_images,
            classify=request.classify,
            classify_categories=request.classify_categories,
            extract_schema=effective_schema,
            ocr_correct=request.ocr_correct,
            show_formulas=request.show_formulas,
            chunk=request.chunk,
            chunk_size=request.chunk_size,
            accuracy=request.accuracy,
            ocr_embed=request.ocr_embed,
            auto_extract=request.auto_extract,
            min_confidence=request.min_confidence,
        )

    # URL — T-MKIT-022: durch convert_auto() routen für vollständige Pipeline
    if request.url:
        import tempfile as _tempfile

        # Content-Type → Extension ermitteln; HTML-Seiten direkt mit convert_url() (markitdown-native)
        try:
            async with httpx.AsyncClient(timeout=float(MISTRAL_TIMEOUT)) as client:
                head_resp = await client.head(request.url, follow_redirects=True)
                url_content_type = head_resp.headers.get("content-type", "text/html")
        except Exception:
            url_content_type = "text/html"

        # Prüfen ob die URL eine HTML-Seite ist → alter Pfad (markitdown convert_url ist besser für HTML)
        is_html = url_content_type.split(";")[0].strip() in (
            "text/html", "application/xhtml+xml", "text/plain",
        )

        if is_html:
            # Fallback: HTML-Seiten mit markitdown convert_url (strukturierter)
            start_time = time.time()
            result = await convert_url(request.url)
            meta = {
                **request.meta,
                "source": request.url,
                "source_type": "url",
                "url": request.url,
                "content_type": url_content_type,
                "duration_ms": int((time.time() - start_time) * 1000),
                "accuracy_mode": request.accuracy,
            }
            if result["success"]:
                if result.get("title"):
                    meta["title"] = result["title"]
                if request.classify:
                    classify_result = await classify_document(
                        result["markdown"], request.classify_categories, request.language
                    )
                    meta.update(classify_result)
                response = create_success_response(result["markdown"], meta=meta)
                # T-MKIT-036: Auto-Extract
                if request.auto_extract and not effective_schema:
                    response = await _apply_auto_extract(response, meta, result["markdown"], request.language, request.min_confidence, [])
                elif effective_schema:
                    extraction = await extract_structured_data(result["markdown"], effective_schema, request.language)
                    if extraction["success"]:
                        response.extracted = extraction["extracted"]
                    else:
                        log.warning("extract_structured_data_failed_url_html", error=extraction.get("error"))
                if request.chunk:
                    response.chunks = chunk_markdown(result["markdown"], chunk_size=request.chunk_size, source=request.url)
                return response
            else:
                return create_error_response(
                    result.get("error_code", ErrorCode.CONVERSION_FAILED),
                    result["error"],
                    meta=meta
                )
        else:
            # Nicht-HTML (PDF, DOCX, Bilder, …) → Inhalt laden und durch convert_auto() schicken
            try:
                async with httpx.AsyncClient(timeout=float(MISTRAL_TIMEOUT)) as client:
                    dl_resp = await client.get(request.url, follow_redirects=True)
                    dl_resp.raise_for_status()
            except Exception as exc:
                return create_error_response(
                    ErrorCode.CONVERSION_FAILED,
                    f"URL-Download fehlgeschlagen: {str(exc)}",
                    meta=request.meta
                )

            raw_content_type = dl_resp.headers.get("content-type", url_content_type)
            ct_base = raw_content_type.split(";")[0].strip()
            guessed_ext = mimetypes.guess_extension(ct_base) or ".bin"
            # mimetypes liefert manchmal .jpe statt .jpg — normalisieren
            _ext_map = {".jpe": ".jpg", ".jpeg": ".jpg"}
            guessed_ext = _ext_map.get(guessed_ext, guessed_ext)

            url_hash = hashlib.md5(request.url.encode()).hexdigest()
            temp_path = TEMP_DIR / f"url_{url_hash}{guessed_ext}"
            temp_path.write_bytes(dl_resp.content)
            try:
                return await convert_auto(
                    file_data=dl_resp.content,
                    filename=temp_path.name,
                    source=request.url,
                    source_type="url",
                    input_meta={**request.meta, "url": request.url, "content_type": raw_content_type},
                    prompt=request.prompt,
                    language=request.language,
                    describe_images=request.describe_images,
                    classify=request.classify,
                    classify_categories=request.classify_categories,
                    ocr_correct=request.ocr_correct,
                    show_formulas=request.show_formulas,
                    chunk=request.chunk,
                    chunk_size=request.chunk_size,
                    extract_schema=effective_schema,
                    accuracy=request.accuracy,
                    auto_extract=request.auto_extract,
                    min_confidence=request.min_confidence,
                )
            finally:
                temp_path.unlink(missing_ok=True)

    return create_error_response(
        ErrorCode.INTERNAL_ERROR,
        "Unerwarteter Zustand",
        meta=request.meta
    )


@app.post("/v1/convert/folder", response_model=ConvertResponse)
async def api_convert_folder(request: ConvertFolderRequest) -> ConvertResponse:
    """
    Konvertiert alle Dateien in einem Ordner zu Markdown.
    """
    folder_path = resolve_path(request.path)
    return await convert_folder_contents(
        folder_path=folder_path,
        input_meta=request.meta,
        language=request.language,
    )


@app.post("/v1/extract", response_model=ConvertResponse)
async def api_extract(request: ExtractRequest) -> ConvertResponse:
    """
    Konvertiert ein Dokument zu Markdown UND extrahiert strukturierte Daten in einem Schritt.

    Kombiniert convert + schema-basierte Extraktion via LLM (AC-014-5).
    """
    inputs = [request.path, request.base64, request.url]
    if sum(1 for x in inputs if x) != 1:
        return create_error_response(
            ErrorCode.INVALID_INPUT,
            "Genau einer von 'path', 'base64' oder 'url' muss angegeben werden",
            meta=request.meta
        )

    # Template → Schema Auflösung (AC-014-4)
    effective_schema = request.extract_schema
    if request.template and not effective_schema:
        tmpl = get_template_by_id(request.template)
        if tmpl is None:
            return create_error_response(
                ErrorCode.INVALID_INPUT,
                f"Unbekanntes Template: '{request.template}'. Verfügbar: {get_all_template_ids()}",
                meta=request.meta
            )
        effective_schema = tmpl["schema"]

    # Wenn auto_extract=true ist extract_schema/template optional
    if not effective_schema and not request.auto_extract:
        return create_error_response(
            ErrorCode.INVALID_INPUT,
            "Entweder 'extract_schema', 'template' oder 'auto_extract=true' muss angegeben werden",
            meta=request.meta
        )

    # Konvertierungs-Request aus ExtractRequest aufbauen (T-MKIT-022: neue Felder durchreichen)
    convert_req = ConvertRequest(
        path=request.path,
        base64=request.base64,
        filename=request.filename,
        url=request.url,
        language=request.language,
        meta=request.meta,
        extract_schema=effective_schema,
        accuracy=request.accuracy,
        ocr_correct=request.ocr_correct,
        describe_images=request.describe_images,
        classify=request.classify,
        auto_extract=request.auto_extract,
        min_confidence=request.min_confidence,
    )
    return await api_convert(convert_req)


@app.get("/v1/templates/categories")
async def api_template_categories() -> dict:
    """Gibt alle Template-Kategorien mit Anzahl zurück (T-MKIT-035)."""
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT category, COUNT(*) as count FROM template WHERE enabled = 1 GROUP BY category ORDER BY category"
    ).fetchall()
    conn.close()
    return {"categories": [{"name": r["category"], "count": r["count"]} for r in rows]}


@app.get("/v1/templates/search")
async def api_search_templates(q: str = "") -> dict:
    """Sucht Templates nach Stichwort (T-MKIT-035)."""
    if not q:
        return {"templates": []}
    results = search_templates(q)
    return {"templates": results}


@app.get("/v1/templates", response_model=TemplateResponse)
async def api_templates() -> TemplateResponse:
    """
    Gibt alle vordefinierten Extraktions-Templates zurück (AC-014-4).

    Templates können direkt als 'template' Parameter in /v1/convert oder /v1/extract
    übergeben werden.
    """
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT id, category, display_name, enabled FROM template ORDER BY category, display_name"
    ).fetchall()
    conn.close()
    # Build legacy-compatible dict with full schemas for TemplateResponse
    templates: dict = {}
    for row in rows:
        tmpl = get_template_by_id(row["id"])
        if tmpl:
            templates[row["id"]] = tmpl["schema"]
    return TemplateResponse(templates=templates)


@app.get("/v1/templates/{template_id}")
async def api_get_template(template_id: str) -> dict:
    """Gibt ein einzelnes Template mit allen Feldern zurück (T-MKIT-035)."""
    tmpl = get_template_by_id(template_id)
    if not tmpl:
        raise HTTPException(status_code=404, detail=f"Template '{template_id}' not found")
    return tmpl


@app.post("/v1/templates/bulk")
async def api_bulk_templates(request: dict) -> dict:
    """Bulk-Upsert für Templates (T-MKIT-035)."""
    templates = request.get("templates", [])
    mode = request.get("mode", "upsert")
    conn = get_db_connection()
    created = 0
    updated = 0
    for tmpl in templates:
        if "id" not in tmpl or "schema" not in tmpl:
            continue
        existing = conn.execute("SELECT id FROM template WHERE id = ?", (tmpl["id"],)).fetchone()
        if existing and mode == "upsert":
            conn.execute(
                """UPDATE template SET category=?, display_name=?, description=?, schema=?,
                   field_descriptions=?, classify_keywords=?, typical_senders=?, steuer_relevanz=?,
                   priority=?, enabled=?, version=?, source=?, notes=?, updated_at=CURRENT_TIMESTAMP
                   WHERE id=?""",
                (tmpl.get("category", "other"), tmpl.get("display_name", tmpl["id"]),
                 tmpl.get("description"), json.dumps(tmpl["schema"]),
                 json.dumps(tmpl.get("field_descriptions")) if tmpl.get("field_descriptions") else None,
                 tmpl.get("classify_keywords"), tmpl.get("typical_senders"),
                 tmpl.get("steuer_relevanz"), tmpl.get("priority", 0),
                 tmpl.get("enabled", True), tmpl.get("version", 1),
                 tmpl.get("source", "manual"), tmpl.get("notes"), tmpl["id"])
            )
            updated += 1
        elif not existing:
            conn.execute(
                """INSERT INTO template (id, category, display_name, description, schema, field_descriptions,
                   classify_keywords, typical_senders, steuer_relevanz, priority, enabled, version, source, notes)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (tmpl["id"], tmpl.get("category", "other"), tmpl.get("display_name", tmpl["id"]),
                 tmpl.get("description"), json.dumps(tmpl["schema"]),
                 json.dumps(tmpl.get("field_descriptions")) if tmpl.get("field_descriptions") else None,
                 tmpl.get("classify_keywords"), tmpl.get("typical_senders"),
                 tmpl.get("steuer_relevanz"), tmpl.get("priority", 0),
                 tmpl.get("enabled", True), tmpl.get("version", 1),
                 tmpl.get("source", "manual"), tmpl.get("notes"))
            )
            created += 1
    conn.commit()
    conn.close()
    return {"success": True, "created": created, "updated": updated, "total": len(templates)}


@app.post("/v1/templates")
async def api_create_template(request: dict) -> dict:
    """Erstellt ein neues Template (T-MKIT-035)."""
    if "id" not in request or "schema" not in request:
        raise HTTPException(status_code=400, detail="'id' and 'schema' are required")
    conn = get_db_connection()
    try:
        conn.execute(
            """INSERT INTO template (id, category, display_name, description, schema, field_descriptions,
               classify_keywords, typical_senders, steuer_relevanz, priority, enabled, version, source, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (request["id"], request.get("category", "other"), request.get("display_name", request["id"]),
             request.get("description"), json.dumps(request["schema"]),
             json.dumps(request.get("field_descriptions")) if request.get("field_descriptions") else None,
             request.get("classify_keywords"), request.get("typical_senders"),
             request.get("steuer_relevanz"), request.get("priority", 0),
             request.get("enabled", True), request.get("version", 1),
             request.get("source", "manual"), request.get("notes"))
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=409, detail=f"Template '{request['id']}' already exists")
    conn.close()
    return {"success": True, "id": request["id"]}


@app.put("/v1/templates/{template_id}")
async def api_update_template(template_id: str, request: dict) -> dict:
    """Aktualisiert ein Template (partial update, T-MKIT-035)."""
    conn = get_db_connection()
    existing = conn.execute("SELECT id FROM template WHERE id = ?", (template_id,)).fetchone()
    if not existing:
        conn.close()
        raise HTTPException(status_code=404, detail=f"Template '{template_id}' not found")
    # Fixed UPDATE with COALESCE — no dynamic SQL construction
    row = conn.execute("SELECT * FROM template WHERE id = ?", (template_id,)).fetchone()
    current = dict(row)
    conn.execute(
        """UPDATE template SET
            category = ?, display_name = ?, description = ?,
            schema = ?, field_descriptions = ?,
            classify_keywords = ?, typical_senders = ?,
            steuer_relevanz = ?, priority = ?, enabled = ?,
            version = ?, source = ?, notes = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?""",
        (
            request.get("category", current["category"]),
            request.get("display_name", current["display_name"]),
            request.get("description", current["description"]),
            json.dumps(request["schema"]) if "schema" in request else current["schema"],
            json.dumps(request["field_descriptions"]) if "field_descriptions" in request else current["field_descriptions"],
            request.get("classify_keywords", current["classify_keywords"]),
            request.get("typical_senders", current["typical_senders"]),
            request.get("steuer_relevanz", current["steuer_relevanz"]),
            request.get("priority", current["priority"]),
            request.get("enabled", current["enabled"]),
            request.get("version", current["version"]),
            request.get("source", current["source"]),
            request.get("notes", current["notes"]),
            template_id,
        )
    )
    conn.commit()
    conn.close()
    return {"success": True, "id": template_id}


@app.delete("/v1/templates/{template_id}")
async def api_delete_template(template_id: str) -> dict:
    """Löscht ein Template (T-MKIT-035)."""
    conn = get_db_connection()
    cursor = conn.execute("DELETE FROM template WHERE id = ?", (template_id,))
    conn.commit()
    conn.close()
    if cursor.rowcount == 0:
        raise HTTPException(status_code=404, detail=f"Template '{template_id}' not found")
    return {"success": True, "id": template_id}


@app.post("/v1/analyze", response_model=ConvertResponse)
async def api_analyze(request: AnalyzeRequest) -> ConvertResponse:
    """Analysiert ein Bild explizit mit Mistral Vision."""
    if not request.path and not request.base64:
        return create_error_response(
            ErrorCode.INVALID_INPUT,
            "'path' oder 'base64' muss angegeben werden",
            meta=request.meta
        )

    if request.path:
        file_path = resolve_path(request.path)
        if not file_path.exists():
            return create_error_response(
                ErrorCode.FILE_NOT_FOUND,
                f"Datei nicht gefunden: {file_path}",
                meta=request.meta
            )
        file_data = file_path.read_bytes()
        filename = file_path.name
        source = str(file_path)
        source_type = "file"
    else:
        if not request.filename:
            return create_error_response(
                ErrorCode.INVALID_INPUT,
                "'filename' ist erforderlich bei Base64",
                meta=request.meta
            )
        try:
            file_data = base64.b64decode(request.base64)
        except Exception as e:
            return create_error_response(
                ErrorCode.INVALID_BASE64,
                f"Ungültiges Base64: {str(e)}",
                meta=request.meta
            )
        filename = request.filename
        source = "base64"
        source_type = "base64"

    start_time = time.time()
    processed_data, resize_meta = resize_image_if_needed(file_data)
    mimetype = detect_mimetype_from_bytes(file_data) or get_mimetype(Path(filename))

    result = await analyze_with_mistral_vision(
        processed_data,
        mimetype,
        request.prompt,
        request.language
    )

    meta = {
        **request.meta,
        **resize_meta,
        "source": source,
        "source_type": source_type,
        "format": get_file_extension(filename).lstrip("."),
        "size_bytes": len(file_data),
        "vision_used": True,
        "duration_ms": int((time.time() - start_time) * 1000),
    }

    if result["success"]:
        meta["vision_model"] = result.get("vision_model")
        meta["tokens_prompt"] = result.get("tokens_prompt")
        meta["tokens_completion"] = result.get("tokens_completion")
        meta["tokens_total"] = result.get("tokens_total")
        return create_success_response(result["markdown"], meta=meta)
    else:
        return create_error_response(
            result.get("error_code", ErrorCode.VISION_FAILED),
            result["error"],
            meta=meta
        )


@app.get("/v1/health", response_model=HealthResponse)
async def api_health() -> HealthResponse:
    """Health-Check Endpoint."""
    uptime = int(time.time() - START_TIME)
    mistral_configured = bool(MISTRAL_API_KEY)

    return HealthResponse(
        status="ok",
        version=VERSION,
        meta={
            "mistral_api_configured": mistral_configured,
            "vision_model": MISTRAL_VISION_MODEL,
            "mistral_ocr_configured": mistral_configured and MISTRAL_OCR_ENABLED,
            "ocr_model": MISTRAL_OCR_MODEL if MISTRAL_OCR_ENABLED else None,
            "max_file_size_mb": MAX_FILE_SIZE_MB,
            "image_max_width": IMAGE_MAX_WIDTH,
            "max_retries": MAX_RETRIES,
            "uptime_seconds": uptime,
            "mcp_port": MCP_PORT,
            "rest_port": REST_PORT,
        }
    )


@app.get("/v1/formats")
async def api_formats() -> dict:
    """Listet unterstützte Formate auf."""
    return {
        "markitdown": sorted(MARKITDOWN_EXTENSIONS),
        "vision": sorted(IMAGE_EXTENSIONS),
        "audio": sorted(AUDIO_EXTENSIONS),
        "video": sorted(VIDEO_EXTENSIONS),
        "all": sorted(MARKITDOWN_EXTENSIONS | IMAGE_EXTENSIONS | AUDIO_EXTENSIONS | VIDEO_EXTENSIONS),
    }


@app.get("/v1/tips")
async def api_tips() -> dict:
    """Usage tips and common patterns for Daigestr."""
    return _build_tips_dict()


# =============================================================================
# MCP Tools
# =============================================================================

@mcp.tool(name="convert")
async def mcp_convert(
    path: Optional[str] = None,
    base64_data: Optional[str] = None,
    filename: Optional[str] = None,
    url: Optional[str] = None,
    meta: Optional[dict] = None,
    accuracy: str = "standard",
    classify: bool = False,
    classify_categories: Optional[list] = None,
    describe_images: bool = False,
    ocr_correct: bool = False,
    ocr_embed: bool = False,
    show_formulas: bool = False,
    chunk: bool = False,
    chunk_size: int = 512,
    extract_schema: Optional[dict] = None,
    template: Optional[str] = None,
    language: str = "de",
    prompt: Optional[str] = None,
    auto_extract: bool = False,
    min_confidence: float = 0.7,
) -> str:
    """
    Converts a file, URL, or base64 payload to Markdown.

    Default output is always Markdown in the 'markdown' field.

    To get STRUCTURED JSON, add extract_schema or template parameter.
    To get AUTO-EXTRACTED JSON (no template needed), add auto_extract=true.
    To get RAG CHUNKS, add chunk=true.
    To CLASSIFY the document type, add classify=true.
    To DESCRIBE IMAGES in DOCX/PPTX, add describe_images=true.
    For SCANNED PDFs, use accuracy='high' for best results.
    To embed OCR text as searchable layer in scanned PDFs, add ocr_embed=true.

    Examples:
      convert(path="<data_dir>/invoice.pdf", template="invoice")  -> extracted JSON
      convert(path="<data_dir>/invoice.pdf", auto_extract=true)  -> classify + auto template lookup + extracted JSON
      convert(path="<data_dir>/scan.pdf", accuracy="high")  -> OCR + correction
      convert(path="<data_dir>/scan.pdf", ocr_embed=true)  -> searchable PDF in enriched_pdf
      convert(url="<url>", chunk=true)  -> RAG chunks
      convert(path="<data_dir>/report.docx", describe_images=true, classify=true)

    Note: MCP uses 'base64_data', REST API uses 'base64'.

    Args:
        path: Dateipfad im Container.
        base64_data: Base64-kodierte Datei (erfordert filename).
        filename: Dateiname (erforderlich bei base64_data).
        url: URL zu Datei oder Webseite.
        meta: Beliebige Metadaten (werden durchgereicht).
        accuracy: Accuracy-Modus: 'standard' (Default) oder 'high'.
        classify: Dokumenttyp via LLM klassifizieren.
        classify_categories: Erlaubte Dokumenttypen (überschreibt Default).
        describe_images: Eingebettete Bilder in DOCX/PPTX beschreiben.
        ocr_correct: OCR-Nachkorrektur via LLM aktivieren.
        ocr_embed: Wenn True, OCR-Text als unsichtbare Textschicht in gescannte PDFs einbetten.
        show_formulas: Excel-Formeln im Output annotieren.
        chunk: Smart Chunking für RAG aktivieren.
        chunk_size: Maximale Chunk-Größe in Tokens (Default: 512).
        extract_schema: JSON-Schema für strukturierte Daten-Extraktion.
        template: Vordefinierter Template-Name als Alternative zu extract_schema.
        language: Antwortsprache ('de' oder 'en').
        prompt: Optionaler Custom-Prompt für Vision.
        auto_extract: Wenn True, wird der Dokumenttyp klassifiziert, ein passendes Template gesucht und strukturierte Daten extrahiert — alles in einem Schritt.
        min_confidence: Minimale Klassifizierungs-Konfidenz für auto_extract (Default: 0.7).
    """
    # Template → Schema Auflösung
    effective_schema = extract_schema
    if template and not effective_schema:
        tmpl = get_template_by_id(template)
        if tmpl is None:
            return json.dumps({
                "success": False,
                "error": f"Unbekanntes Template: '{template}'. Verfügbar: {get_all_template_ids()}"
            })
        effective_schema = tmpl["schema"]

    if path:
        file_path = resolve_path(path)
        if not file_path.exists():
            return json.dumps({"success": False, "error": f"Datei nicht gefunden: {file_path}"})
        file_data = file_path.read_bytes()
        response = await convert_auto(
            file_data=file_data,
            filename=file_path.name,
            source=str(file_path),
            source_type="file",
            input_meta=meta or {},
            prompt=prompt,
            language=language,
            describe_images=describe_images,
            classify=classify,
            classify_categories=classify_categories,
            extract_schema=effective_schema,
            ocr_correct=ocr_correct,
            show_formulas=show_formulas,
            chunk=chunk,
            chunk_size=chunk_size,
            accuracy=accuracy,
            ocr_embed=ocr_embed,
            auto_extract=auto_extract,
            min_confidence=min_confidence,
        )
    elif base64_data and filename:
        try:
            file_data = base64.b64decode(base64_data)
        except Exception as e:
            return json.dumps({"success": False, "error": f"Ungültiges Base64: {e}"})
        response = await convert_auto(
            file_data=file_data,
            filename=filename,
            source="base64",
            source_type="base64",
            input_meta=meta or {},
            prompt=prompt,
            language=language,
            describe_images=describe_images,
            classify=classify,
            classify_categories=classify_categories,
            extract_schema=effective_schema,
            ocr_correct=ocr_correct,
            show_formulas=show_formulas,
            chunk=chunk,
            chunk_size=chunk_size,
            accuracy=accuracy,
            ocr_embed=ocr_embed,
            auto_extract=auto_extract,
            min_confidence=min_confidence,
        )
    elif url:
        result = await convert_url(url)
        if result["success"]:
            url_meta: dict[str, Any] = {"source": url, "source_type": "url"}
            if result.get("title"):
                url_meta["title"] = result["title"]
            if classify:
                classify_result = await classify_document(result["markdown"], classify_categories, language)
                url_meta.update(classify_result)
            mcp_url_meta = {**(meta or {}), **url_meta}
            response = create_success_response(result["markdown"], meta=mcp_url_meta)
            # T-MKIT-036: Auto-Extract für URL
            if auto_extract and not effective_schema:
                response = await _apply_auto_extract(response, mcp_url_meta, result["markdown"], language, min_confidence, [])
            elif effective_schema:
                extraction = await extract_structured_data(result["markdown"], effective_schema, language)
                if extraction["success"]:
                    response.extracted = extraction["extracted"]
                else:
                    log.warning("mcp_convert_extract_failed_url", error=extraction.get("error"))
            if chunk:
                response.chunks = chunk_markdown(result["markdown"], chunk_size=chunk_size, source=url)
        else:
            response = create_error_response(result.get("error_code", "ERROR"), result["error"])
    else:
        return json.dumps({"success": False, "error": "path, url oder (base64_data + filename) erforderlich"})

    return response.model_dump_json()


@mcp.tool(name="extract")
async def mcp_extract(
    extract_schema: Optional[dict] = None,
    path: Optional[str] = None,
    base64_data: Optional[str] = None,
    filename: Optional[str] = None,
    url: Optional[str] = None,
    template: Optional[str] = None,
    language: str = "de",
    meta: Optional[dict] = None,
    accuracy: str = "standard",
    ocr_correct: bool = False,
    classify: bool = False,
    auto_extract: bool = False,
    min_confidence: float = 0.7,
) -> str:
    """
    Converts a file and extracts structured data in one step.

    REQUIRES either extract_schema (JSON Schema dict), template name, or auto_extract=true.
    Available templates: 'invoice', 'cv', 'contract'.
    See /v1/templates for template schemas.

    The response contains both 'markdown' (full text) and 'extracted' (structured JSON).

    Examples:
      extract(path="<data_dir>/invoice.pdf", template="invoice")
      extract(path="<data_dir>/cv.pdf", template="cv")
      extract(path="<data_dir>/doc.pdf", auto_extract=true)  -> classify + auto template lookup + extracted JSON
      extract(path="<data_dir>/doc.pdf", extract_schema={"type": "object", "properties": {"title": {"type": "string"}}})

    Args:
        extract_schema: JSON-Schema für die gewünschten extrahierten Felder (optional wenn template oder auto_extract gesetzt).
        path: Dateipfad im Container (alternativ zu base64_data oder url).
        base64_data: Base64-kodierte Datei (erfordert filename).
        filename: Dateiname (erforderlich bei base64_data).
        url: URL zu Datei oder Webseite (alternativ zu path/base64_data).
        template: Vordefinierter Template-Name ('invoice', 'cv', 'contract') als
                  Alternative zu extract_schema.
        language: Antwortsprache ('de' oder 'en').
        meta: Beliebige Metadaten (werden durchgereicht).
        accuracy: Accuracy-Modus: 'standard' (Default) oder 'high'.
        ocr_correct: OCR-Nachkorrektur via LLM aktivieren.
        classify: Dokumenttyp via LLM klassifizieren.
        auto_extract: Wenn True, wird Dokumenttyp klassifiziert, ein passendes Template gesucht und Daten extrahiert.
        min_confidence: Minimale Klassifizierungs-Konfidenz für auto_extract (Default: 0.7).
    """
    # Template → Schema Auflösung
    effective_schema = extract_schema
    if template and not effective_schema:
        tmpl = get_template_by_id(template)
        if tmpl is None:
            return json.dumps({
                "success": False,
                "error": f"Unbekanntes Template: '{template}'. Verfügbar: {get_all_template_ids()}"
            })
        effective_schema = tmpl["schema"]

    # auto_extract=true macht extract_schema/template optional
    if not effective_schema and not auto_extract:
        return json.dumps({
            "success": False,
            "error": "Entweder 'extract_schema', 'template' oder 'auto_extract=true' muss angegeben werden"
        })

    if path:
        file_path = resolve_path(path)
        if not file_path.exists():
            return json.dumps({"success": False, "error": f"Datei nicht gefunden: {file_path}"})
        file_data = file_path.read_bytes()
        response = await convert_auto(
            file_data=file_data,
            filename=file_path.name,
            source=str(file_path),
            source_type="file",
            input_meta=meta or {},
            language=language,
            extract_schema=effective_schema,
            accuracy=accuracy,
            ocr_correct=ocr_correct,
            classify=classify,
            auto_extract=auto_extract,
            min_confidence=min_confidence,
        )
    elif base64_data and filename:
        try:
            file_data = base64.b64decode(base64_data)
        except Exception as e:
            return json.dumps({"success": False, "error": f"Ungültiges Base64: {e}"})
        response = await convert_auto(
            file_data=file_data,
            filename=filename,
            source="base64",
            source_type="base64",
            input_meta=meta or {},
            language=language,
            extract_schema=effective_schema,
            accuracy=accuracy,
            ocr_correct=ocr_correct,
            classify=classify,
            auto_extract=auto_extract,
            min_confidence=min_confidence,
        )
    elif url:
        result = await convert_url(url)
        if result["success"]:
            mcp_extract_url_meta: dict[str, Any] = {"source": url, "source_type": "url"}
            if classify:
                classify_result = await classify_document(result["markdown"], None, language)
                mcp_extract_url_meta.update(classify_result)
            combined_meta = {**(meta or {}), **mcp_extract_url_meta}
            response = create_success_response(result["markdown"], meta=combined_meta)
            # T-MKIT-036: Auto-Extract für URL
            if auto_extract and not effective_schema:
                response = await _apply_auto_extract(response, combined_meta, result["markdown"], language, min_confidence, [])
            elif effective_schema:
                extraction = await extract_structured_data(result["markdown"], effective_schema, language)
                if extraction["success"]:
                    response.extracted = extraction["extracted"]
        else:
            response = create_error_response(result.get("error_code", "ERROR"), result["error"])
    else:
        return json.dumps({"success": False, "error": "path, url oder (base64_data + filename) erforderlich"})

    return response.model_dump_json()


@mcp.tool(name="convert_folder")
async def mcp_convert_folder(
    path: str,
    meta: Optional[dict] = None,
    language: str = "de",
) -> str:
    """
    Converts all files in a directory to a single merged Markdown document.
    Each file becomes a ## filename section. Per-file metadata is tracked.

    Args:
        path: Ordnerpfad im Container.
        meta: Beliebige Metadaten (werden durchgereicht).
        language: Antwortsprache ('de' oder 'en').
    """
    folder_path = resolve_path(path)
    response = await convert_folder_contents(
        folder_path=folder_path,
        input_meta=meta or {},
        language=language,
    )
    return response.model_dump_json()


@mcp.tool(name="health")
async def mcp_health() -> str:
    """Health-Check (MCP-Version)."""
    response = await api_health()
    return response.model_dump_json()


@mcp.tool(name="list_files")
async def mcp_list_files(subdir: str = "") -> str:
    """Listet Dateien im /data Verzeichnis auf."""
    target_dir = DATA_DIR / subdir if subdir else DATA_DIR
    if not target_dir.exists():
        return json.dumps({"error": f"Verzeichnis nicht gefunden: {target_dir}"})

    files = []
    for item in sorted(target_dir.iterdir()):
        if item.is_file():
            files.append({
                "name": item.name,
                "size": item.stat().st_size,
                "type": item.suffix.lower()
            })
        elif item.is_dir():
            files.append({"name": item.name + "/", "type": "directory"})

    return json.dumps({"path": str(target_dir), "files": files}, ensure_ascii=False)


@mcp.tool(name="get_tips")
async def mcp_get_tips() -> str:
    """Returns usage tips and common patterns for Daigestr. Call this first to understand available features."""
    return json.dumps(_build_tips_dict(), indent=2, ensure_ascii=False)


# =============================================================================
# Server Start
# =============================================================================

def run_rest_server():
    """Startet den REST-Server in einem separaten Thread."""
    uvicorn.run(
        app,
        host=os.getenv("BIND_HOST", "0.0.0.0"),
        port=REST_PORT,
        log_level=LOG_LEVEL.lower(),
    )


if __name__ == "__main__":
    # Initialize Template Registry DB (T-MKIT-035)
    try:
        init_templates_db()
        log.info("template_registry_initialized", db_path=str(TEMPLATES_DB_PATH))
    except Exception as _db_init_err:
        log.warning("template_registry_init_failed", error=str(_db_init_err))

    log.info("server_starting",
             version=VERSION,
             data_dir=str(DATA_DIR),
             vision_model=MISTRAL_VISION_MODEL,
             vision_enabled=bool(MISTRAL_API_KEY),
             mcp_port=MCP_PORT,
             rest_port=REST_PORT,
             max_file_size_mb=MAX_FILE_SIZE_MB,
             image_max_width=IMAGE_MAX_WIDTH,
             max_retries=MAX_RETRIES)

    print(f"Daigestr v{VERSION}")
    print(f"Data-Verzeichnis: {DATA_DIR}")
    print(f"Vision-Modell: {MISTRAL_VISION_MODEL}")
    print(f"Vision aktiviert: {'Ja' if MISTRAL_API_KEY else 'Nein'}")
    print(f"MCP-Port: {MCP_PORT}")
    print(f"REST-Port: {REST_PORT}")
    print(f"Max Dateigröße: {MAX_FILE_SIZE_MB}MB")
    print(f"Max Bildbreite: {IMAGE_MAX_WIDTH}px")
    print(f"Max Retries: {MAX_RETRIES}")
    print("-" * 50)

    rest_thread = threading.Thread(target=run_rest_server, daemon=True)
    rest_thread.start()
    print(f"REST-API gestartet auf Port {REST_PORT}")

    transport = os.getenv("MCP_TRANSPORT", "sse")
    if transport == "stdio":
        mcp.run()
    else:
        print(f"MCP-Server startet auf Port {MCP_PORT}")
        mcp.run(transport="sse", host=os.getenv("BIND_HOST", "0.0.0.0"), port=MCP_PORT)
